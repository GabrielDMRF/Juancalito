import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import os
import sys
import requests
import json
from datetime import datetime, date, timedelta
from PIL import Image, ImageTk
import io
import base64
import threading
import time

# Agregar path para importar modelos
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.database import get_db, Empleado
from utils.qr_server import QRServer

class AsistenciaQRView:
    def __init__(self, parent):
        self.parent = parent
        self.db = get_db()
        self.qr_server = None
        self.qr_image = None
        self.qr_label = None
        self.status_label = None
        self.refresh_thread = None
        self.is_running = False
        self.qr_port = 5000  # Puerto por defecto
        
        # Colores modernos
        self.colors = {
            'primary': '#3498db',
            'secondary': '#2c3e50',
            'success': '#27ae60',
            'warning': '#f39c12',
            'danger': '#e74c3c',
            'light': '#ecf0f1',
            'dark': '#2c3e50',
            'white': '#ffffff',
            'gray': '#95a5a6'
        }
        
        # Obtener ruta de la base de datos
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        self.db_path = os.path.join(base_dir, 'empleados.db')
        
        self.setup_window()
        self.create_widgets()
        
        # Iniciar servidor en hilo separado para no bloquear la UI
        threading.Thread(target=self.start_qr_server_thread, daemon=True).start()
        
        # Iniciar sincronización automática (optimizada)
        # self.iniciar_sincronizacion_automatica()
        
        # Iniciar sincronización desde Railway cada 30 segundos (optimizada)
        # self.iniciar_sincronizacion_desde_railway()
    
    def setup_window(self):
        """Configurar ventana principal moderna"""
        self.window = tk.Toplevel(self.parent)
        self.window.title("📱 Sistema de Asistencia QR")
        
        # Configurar ventana
        self.window.configure(bg=self.colors['light'])
        self.window.geometry("1000x700")
        self.window.minsize(900, 600)
        
        # Centrar ventana
        self.window.transient(self.parent)
        self.window.grab_set()
        
        # Configurar cierre
        self.window.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Configurar estilo moderno
        self.setup_styles()
    
    def setup_styles(self):
        """Configurar estilos modernos para ttk"""
        style = ttk.Style()
        
        # Configurar tema
        style.theme_use('clam')
        
        # Configurar colores para diferentes widgets
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 18, 'bold'), 
                       foreground=self.colors['dark'],
                       background=self.colors['light'])
        
        style.configure('Subtitle.TLabel', 
                       font=('Segoe UI', 12), 
                       foreground=self.colors['secondary'],
                       background=self.colors['light'])
        
        style.configure('Status.TLabel', 
                       font=('Segoe UI', 10), 
                       background=self.colors['light'])
        
        style.configure('Info.TLabel', 
                       font=('Segoe UI', 9), 
                       background=self.colors['light'])
        
        # Configurar frames
        style.configure('Card.TFrame', 
                       background=self.colors['white'],
                       relief='flat',
                       borderwidth=1)
        
        # Configurar botones
        style.configure('Primary.TButton', 
                       font=('Segoe UI', 10, 'bold'),
                       background=self.colors['primary'],
                       foreground=self.colors['white'])
        
        style.configure('Success.TButton', 
                       font=('Segoe UI', 10, 'bold'),
                       background=self.colors['success'],
                       foreground=self.colors['white'])
        
        style.configure('Warning.TButton', 
                       font=('Segoe UI', 10, 'bold'),
                       background=self.colors['warning'],
                       foreground=self.colors['white'])
    
    def create_widgets(self):
        """Crear widgets de la interfaz moderna"""
        # Frame principal con padding
        main_frame = ttk.Frame(self.window, style='Card.TFrame', padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header con título y estado
        self.create_header(main_frame)
        
        # Contenido principal en dos columnas
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))
        
        # Columna izquierda - QR y controles
        left_frame = ttk.Frame(content_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        self.create_qr_section(left_frame)
        self.create_controls_section(left_frame)
        
        # Columna derecha - Registros recientes
        right_frame = ttk.Frame(content_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        self.create_records_section(right_frame)
        
        # Iniciar actualización automática (deshabilitada temporalmente)
        # self.iniciar_actualizacion_automatica()
    
    def create_header(self, parent):
        """Crear header con título y estado"""
        header_frame = ttk.Frame(parent, style='Card.TFrame')
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Título principal
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(fill=tk.X, padx=20, pady=15)
        
        title_label = ttk.Label(title_frame, 
                               text="📱 Sistema de Asistencia QR", 
                               style='Title.TLabel')
        title_label.pack(side=tk.LEFT)
        
        # Estado del servidor
        status_frame = ttk.Frame(header_frame)
        status_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        status_icon = ttk.Label(status_frame, text="🔄", font=('Segoe UI', 12))
        status_icon.pack(side=tk.LEFT, padx=(0, 8))
        
        self.status_label = ttk.Label(status_frame, 
                                     text="Iniciando servidor...", 
                                     style='Status.TLabel')
        self.status_label.pack(side=tk.LEFT)
    
    def create_qr_section(self, parent):
        """Crear sección del QR"""
        qr_frame = ttk.LabelFrame(parent, text="📱 QR del Día", padding="20")
        qr_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Contenedor del QR con borde y sombra
        qr_container = ttk.Frame(qr_frame, style='Card.TFrame')
        qr_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Label para mostrar el QR
        self.qr_label = ttk.Label(qr_container, 
                                 text="🔄 Generando QR...", 
                                 font=('Segoe UI', 12),
                                 background=self.colors['white'],
                                 relief='solid',
                                 borderwidth=1)
        self.qr_label.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)
        
        # Información del QR
        info_frame = ttk.Frame(qr_frame)
        info_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.info_label = ttk.Label(info_frame, 
                                   text="", 
                                   style='Info.TLabel',
                                   justify=tk.CENTER)
        self.info_label.pack()
        
    def sincronizar_con_railway(self):
        """Sincronizar datos con Railway"""
        try:
            import sys
            import os
            # Agregar el directorio raíz al path para poder importar src
            sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
            from src.utils.railway_sync import RailwaySync
            
            # Crear instancia de sincronización
            sync = RailwaySync()
            
            # Verificar estado de Railway
            status = sync.get_railway_status()
            if status.get('status') != 'healthy':
                messagebox.showerror("❌ Error", f"Railway no está disponible:\n{status.get('message', 'Error desconocido')}")
                return
            
            # Mostrar progreso
            progress_window = tk.Toplevel(self.window)
            progress_window.title("🔄 Sincronizando con Railway")
            progress_window.geometry("400x200")
            progress_window.resizable(False, False)
            progress_window.transient(self.window)
            progress_window.grab_set()
            
            # Centrar ventana
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (400 // 2)
            y = (progress_window.winfo_screenheight() // 2) - (200 // 2)
            progress_window.geometry(f"400x200+{x}+{y}")
            
            # Contenido de la ventana
            ttk.Label(progress_window, text="🔄 Sincronizando datos...", 
                     style='Title.TLabel').pack(pady=20)
            
            progress_label = ttk.Label(progress_window, text="Enviando empleados...", 
                                     style='Info.TLabel')
            progress_label.pack(pady=10)
            
            progress_bar = ttk.Progressbar(progress_window, mode='indeterminate')
            progress_bar.pack(pady=10, padx=20, fill='x')
            progress_bar.start()
            
            def sync_process():
                try:
                    # Sincronizar empleados
                    progress_label.config(text="Enviando empleados...")
                    progress_window.update()
                    
                    if sync.sync_empleados_to_railway():
                        # Sincronizar asistencias
                        progress_label.config(text="Enviando asistencias...")
                        progress_window.update()
                        
                        if sync.sync_asistencias_to_railway():
                            progress_label.config(text="✅ Sincronización completada")
                            progress_bar.stop()
                            progress_window.after(2000, progress_window.destroy)
                            messagebox.showinfo("✅ Éxito", "Datos sincronizados correctamente con Railway")
                        else:
                            progress_window.destroy()
                            messagebox.showerror("❌ Error", "Error sincronizando asistencias")
                    else:
                        progress_window.destroy()
                        messagebox.showerror("❌ Error", "Error sincronizando empleados")
                        
                except Exception as e:
                    progress_window.destroy()
                    messagebox.showerror("❌ Error", f"Error en sincronización: {str(e)}")
            
            # Ejecutar sincronización en hilo separado
            import threading
            sync_thread = threading.Thread(target=sync_process, daemon=True)
            sync_thread.start()
            
        except Exception as e:
            messagebox.showerror("❌ Error", f"Error iniciando sincronización: {str(e)}")

    def create_controls_section(self, parent):
        """Crear sección de controles"""
        controls_frame = ttk.LabelFrame(parent, text="⚙️ Controles", padding="15")
        controls_frame.pack(fill=tk.X)
        
        # Botones en grid
        buttons_frame = ttk.Frame(controls_frame)
        buttons_frame.pack(fill=tk.X)
        
        # Primera fila de botones
        row1 = ttk.Frame(buttons_frame)
        row1.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(row1, text="🔄 Actualizar QR", 
                  style='Primary.TButton',
                  command=self.actualizar_qr).pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        ttk.Button(row1, text="📊 Ver Reportes", 
                  style='Success.TButton',
                  command=self.abrir_reportes).pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        # Segunda fila de botones
        row2 = ttk.Frame(buttons_frame)
        row2.pack(fill=tk.X)
        
        ttk.Button(row2, text="🌐 Abrir en Navegador", 
                  style='Warning.TButton',
                  command=self.abrir_navegador).pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        ttk.Button(row2, text="❌ Cerrar", 
                  style='Warning.TButton',
                  command=self.on_closing).pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Tercera fila de botones
        row3 = ttk.Frame(buttons_frame)
        row3.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(row3, text="🔄 Sincronizar con Railway", 
                  style='Primary.TButton',
                  command=self.sincronizar_con_railway).pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    def create_records_section(self, parent):
        """Crear sección de registros recientes"""
        records_frame = ttk.LabelFrame(parent, text="📋 Registros Recientes", padding="15")
        records_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header de registros
        records_header = ttk.Frame(records_frame)
        records_header.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(records_header, 
                 text="Registros de hoy", 
                 style='Subtitle.TLabel').pack(side=tk.LEFT)
        
        # Contador de registros
        self.records_count_label = ttk.Label(records_header, 
                                            text="0 registros", 
                                            style='Info.TLabel')
        self.records_count_label.pack(side=tk.RIGHT)
        
        # Treeview para registros con estilo moderno
        tree_frame = ttk.Frame(records_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('empleado', 'fecha', 'hora_entrada', 'hora_salida', 'tipo')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=8)
        
        # Configurar columnas con mejor formato
        self.tree.heading('empleado', text='👤 Empleado')
        self.tree.heading('fecha', text='📅 Fecha')
        self.tree.heading('hora_entrada', text='⏰ Entrada')
        self.tree.heading('hora_salida', text='🚪 Salida')
        self.tree.heading('tipo', text='📝 Tipo')
        
        self.tree.column('empleado', width=180, minwidth=150)
        self.tree.column('fecha', width=80, minwidth=80)
        self.tree.column('hora_entrada', width=80, minwidth=80)
        self.tree.column('hora_salida', width=80, minwidth=80)
        self.tree.column('tipo', width=80, minwidth=80)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Configurar colores alternados para las filas
        self.tree.tag_configure('oddrow', background=self.colors['light'])
        self.tree.tag_configure('evenrow', background=self.colors['white'])
    
    def start_qr_server_thread(self):
        """Iniciar servidor QR en hilo separado"""
        try:
            # Verificar si el puerto 5000 está disponible
            import socket
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            result = sock.connect_ex(('localhost', 5000))
            sock.close()
            
            if result == 0:
                # Puerto 5000 está ocupado, usar 5001
                self.qr_port = 5001
                self.qr_server = QRServer(self.db_path, port=5001)
            else:
                # Puerto 5000 está libre
                self.qr_port = 5000
                self.qr_server = QRServer(self.db_path, port=5000)
            
            self.qr_server.start_server()
            
            # Programar verificación del servidor después de un breve delay
            self.window.after(2000, self.verificar_servidor)
            
        except Exception as e:
            # Usar after para actualizar la UI desde el hilo principal
            error_msg = str(e)
            self.window.after(0, lambda: self.status_label.config(text=f"❌ Error: {error_msg}"))
            # No mostrar error modal para evitar bloqueos
            print(f"Error iniciando servidor QR: {error_msg}")
    
    def verificar_servidor(self):
        """Verificar si el servidor está activo y actualizar QR"""
        try:
            # Intentar conectar al servidor en el puerto configurado
            port = getattr(self, 'qr_port', 5000)
            response = requests.get(f'http://localhost:{port}/', timeout=3)
            if response.status_code == 200:
                self.status_label.config(text=f"✅ Servidor activo en puerto {port}")
                self.actualizar_qr()
                return
        except requests.exceptions.RequestException:
            pass
        
        # Si aún no está listo, intentar de nuevo en 3 segundos (menos agresivo)
        self.window.after(3000, self.verificar_servidor)
    
    def actualizar_qr(self):
        """Actualizar QR del día con mejor manejo de errores"""
        try:
            if not self.qr_server or not self.qr_server.is_running:
                self.qr_label.config(text="❌ Servidor no disponible")
                return
            
            # Mostrar indicador de carga
            self.qr_label.config(text="🔄 Generando QR...", image="")
            self.window.update_idletasks()
            
            # Usar el puerto correcto
            port = getattr(self, 'qr_port', 5000)
            
            # Obtener QR del servidor con timeout más largo
            response = requests.get(f'http://localhost:{port}/qr_diario', timeout=10)
            if response.status_code == 200:
                qr_data = response.json()
                
                # Decodificar imagen QR
                img_data = qr_data['qr_image'].split(',')[1]
                img_bytes = base64.b64decode(img_data)
                img = Image.open(io.BytesIO(img_bytes))
                
                # Tamaño responsive
                qr_size = min(300, 250)  # Tamaño máximo más pequeño
                
                # Redimensionar imagen
                img = img.resize((qr_size, qr_size), Image.Resampling.LANCZOS)
                self.qr_image = ImageTk.PhotoImage(img)
                
                # Mostrar QR
                self.qr_label.config(image=self.qr_image, text="")
                
                # Actualizar información con mejor formato
                ip_local = qr_data.get('ip_local', 'localhost')
                info_text = f"📅 Fecha: {qr_data['fecha']}\n🔑 Token: {qr_data['token'][:20]}...\n🌐 IP: {ip_local}:{port}"
                self.info_label.config(text=info_text)
                
            else:
                self.qr_label.config(text="❌ Error obteniendo QR")
                
        except requests.exceptions.Timeout:
            self.qr_label.config(text="⏰ Timeout - Reintentando...")
            # Reintentar después de 5 segundos (menos agresivo)
            self.window.after(5000, self.actualizar_qr)
        except requests.exceptions.ConnectionError:
            self.qr_label.config(text="🔌 Error de conexión")
        except Exception as e:
            self.qr_label.config(text=f"❌ Error: {str(e)[:30]}...")
            print(f"Error actualizando QR: {e}")
    
    def abrir_reportes(self):
        """Abrir ventana de reportes de asistencia"""
        ReportesAsistenciaView(self.window)
    
    def abrir_navegador(self):
        """Abrir formulario de asistencia en navegador"""
        try:
            import webbrowser
            port = getattr(self, 'qr_port', 5000)
            webbrowser.open(f'http://localhost:{port}')
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el navegador: {str(e)}")
    
    def cargar_registros_recientes(self):
        """Cargar registros de asistencia recientes con mejor formato"""
        try:
            # Limpiar treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Obtener registros de hoy
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT e.nombre_completo, a.fecha, a.hora_entrada, a.hora_salida, a.tipo_registro
                FROM asistencias a
                JOIN empleados e ON a.empleado_id = e.id
                WHERE a.fecha >= date('now', '-7 days')
                ORDER BY a.fecha DESC, a.hora_entrada DESC
                LIMIT 20
            """)
            
            registros = cursor.fetchall()
            conn.close()
            
            # Actualizar contador
            self.records_count_label.config(text=f"{len(registros)} registros")
            
            # Insertar en treeview con colores alternados
            for i, registro in enumerate(registros):
                nombre, fecha, entrada, salida, tipo = registro
                
                # Formatear fechas con mejor manejo
                try:
                    if isinstance(fecha, str):
                        fecha_str = fecha
                    else:
                        fecha_str = fecha.strftime('%d/%m/%Y')
                except:
                    fecha_str = str(fecha)
                
                try:
                    if isinstance(entrada, str):
                        # Formatear entrada para mostrar solo hora
                        try:
                            entrada_dt = datetime.strptime(entrada, '%Y-%m-%d %H:%M:%S.%f')
                            entrada_str = entrada_dt.strftime('%H:%M:%S')
                        except ValueError:
                            try:
                                entrada_dt = datetime.strptime(entrada, '%Y-%m-%d %H:%M:%S')
                                entrada_str = entrada_dt.strftime('%H:%M:%S')
                            except ValueError:
                                entrada_str = entrada
                    elif entrada:
                        entrada_str = entrada.strftime('%H:%M:%S')
                    else:
                        entrada_str = '-'
                except:
                    entrada_str = str(entrada) if entrada else '-'
                
                try:
                    if isinstance(salida, str):
                        # Formatear salida para mostrar solo hora
                        try:
                            salida_dt = datetime.strptime(salida, '%Y-%m-%d %H:%M:%S.%f')
                            salida_str = salida_dt.strftime('%H:%M:%S')
                        except ValueError:
                            try:
                                salida_dt = datetime.strptime(salida, '%Y-%m-%d %H:%M:%S')
                                salida_str = salida_dt.strftime('%H:%M:%S')
                            except ValueError:
                                salida_str = salida
                    elif salida:
                        salida_str = salida.strftime('%H:%M:%S')
                    else:
                        salida_str = '-'
                except:
                    salida_str = str(salida) if salida else '-'
                
                # Aplicar colores alternados
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                
                self.tree.insert('', 'end', 
                               values=(nombre, fecha_str, entrada_str, salida_str, tipo),
                               tags=(tag,))
                
        except Exception as e:
            print(f"Error cargando registros: {e}")
    
    def iniciar_actualizacion_automatica(self):
        """Iniciar actualización automática de registros"""
        self.is_running = True
        
        def actualizar_loop():
            last_count = 0
            while self.is_running:
                try:
                    # Obtener conteo actual de registros
                    conn = sqlite3.connect(self.db_path)
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT COUNT(*) FROM asistencias 
                        WHERE fecha >= date('now', '-7 days')
                    """)
                    current_count = cursor.fetchone()[0]
                    conn.close()
                    
                    # Si hay nuevos registros, actualizar la interfaz
                    if current_count != last_count:
                        print(f"🔄 Nuevos registros detectados: {current_count - last_count}")
                        self.cargar_registros_recientes()
                        last_count = current_count
                        
                        # Actualizar contador en la interfaz
                        if hasattr(self, 'records_count_label'):
                            self.records_count_label.config(text=f"{current_count} registros")
                    
                    time.sleep(10)  # Verificar cada 10 segundos
                except Exception as e:
                    print(f"Error en actualización automática: {e}")
                    time.sleep(30)  # Esperar más tiempo si hay error
        
        self.refresh_thread = threading.Thread(target=actualizar_loop, daemon=True)
        self.refresh_thread.start()
    
    def iniciar_sincronizacion_automatica(self):
        """Iniciar sincronización automática de empleados"""
        def sincronizar_empleados():
            try:
                # Obtener datos de SQLAlchemy
                db = get_db()
                empleados_sqlalchemy = db.query(Empleado).filter(Empleado.estado == True).all()
                db.close()
                
                # Actualizar SQLite
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Limpiar empleados existentes
                cursor.execute("DELETE FROM empleados")
                
                # Insertar empleados actualizados
                for emp in empleados_sqlalchemy:
                    cursor.execute("""
                        INSERT INTO empleados (
                            id, nombre_completo, cedula, telefono, email, 
                            area_trabajo, cargo, salario_base, estado, fecha_creacion
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        emp.id,
                        emp.nombre_completo,
                        emp.cedula,
                        emp.telefono or '',
                        emp.email or '',
                        emp.area_trabajo or '',
                        emp.cargo or '',
                        emp.salario_base or 0,
                        emp.estado,
                        emp.fecha_creacion or '2024-01-01'
                    ))
                
                conn.commit()
                conn.close()
                
                print(f"Sincronizacion automatica: {len(empleados_sqlalchemy)} empleados")
                
            except Exception as e:
                print(f"Error en sincronizacion automatica: {e}")
        
        # Sincronización inicial
        sincronizar_empleados()
        
        # Sincronización automática cada 60 segundos
        def loop_sincronizacion():
            while self.is_running:
                try:
                    time.sleep(60)  # Sincronizar cada 60 segundos
                    if self.is_running:
                        sincronizar_empleados()
                except Exception as e:
                    print(f"Error en loop de sincronizacion: {e}")
                    time.sleep(120)  # Esperar más tiempo si hay error
        
        # Iniciar hilo de sincronización
        sync_thread = threading.Thread(target=loop_sincronizacion, daemon=True)
        sync_thread.start()
        print("Sincronizacion automatica de empleados iniciada")
    
    def iniciar_sincronizacion_desde_railway(self):
        """Iniciar sincronización automática desde Railway"""
        def sincronizar_desde_railway():
            try:
                from utils.railway_sync import RailwaySync
                sync = RailwaySync()
                if sync.sync_from_railway():
                    print("✅ Sincronización desde Railway completada")
                    # Recargar registros después de sincronizar
                    self.cargar_registros_recientes()
                else:
                    print("❌ Error sincronizando desde Railway")
            except Exception as e:
                print(f"Error en sincronización desde Railway: {e}")
        
        # Sincronización inicial
        sincronizar_desde_railway()
        
        # Sincronización automática cada 30 segundos
        def loop_sincronizacion_railway():
            while self.is_running:
                try:
                    time.sleep(30)  # Sincronizar cada 30 segundos
                    if self.is_running:
                        sincronizar_desde_railway()
                except Exception as e:
                    print(f"Error en loop de sincronización Railway: {e}")
                    time.sleep(60)  # Esperar más tiempo si hay error
        
        # Iniciar hilo de sincronización desde Railway
        sync_railway_thread = threading.Thread(target=loop_sincronizacion_railway, daemon=True)
        sync_railway_thread.start()
        print("Sincronización automática desde Railway iniciada (cada 30 segundos)")
    
    def on_closing(self):
        """Manejar cierre de ventana"""
        self.is_running = False
        
        if self.qr_server:
            self.qr_server.stop_server()
        
        self.window.destroy()


class ReportesAsistenciaView:
    def __init__(self, parent):
        self.parent = parent
        self.db = get_db()
        
        # Colores modernos
        self.colors = {
            'primary': '#3498db',
            'secondary': '#2c3e50',
            'success': '#27ae60',
            'warning': '#f39c12',
            'danger': '#e74c3c',
            'light': '#ecf0f1',
            'dark': '#2c3e50',
            'white': '#ffffff',
            'gray': '#95a5a6'
        }
        
        # Obtener ruta de la base de datos
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        self.db_path = os.path.join(base_dir, 'empleados.db')
        
        self.setup_window()
        self.setup_styles()
        self.create_widgets()
    
    def setup_window(self):
        """Configurar ventana de reportes moderna"""
        self.window = tk.Toplevel(self.parent)
        self.window.title("📊 Reportes de Asistencia")
        
        # Configurar ventana
        self.window.configure(bg=self.colors['light'])
        self.window.geometry("1200x700")
        self.window.minsize(1000, 600)
        
        # Centrar ventana
        self.window.transient(self.parent)
        self.window.grab_set()
    
    def setup_styles(self):
        """Configurar estilos modernos"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configurar estilos
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 16, 'bold'), 
                       foreground=self.colors['dark'],
                       background=self.colors['light'])
        
        style.configure('Subtitle.TLabel', 
                       font=('Segoe UI', 11), 
                       foreground=self.colors['secondary'],
                       background=self.colors['light'])
        
        style.configure('Card.TFrame', 
                       background=self.colors['white'],
                       relief='flat',
                       borderwidth=1)
        
        style.configure('Primary.TButton', 
                       font=('Segoe UI', 10, 'bold'),
                       background=self.colors['primary'],
                       foreground=self.colors['white'])
        
        style.configure('Success.TButton', 
                       font=('Segoe UI', 10, 'bold'),
                       background=self.colors['success'],
                       foreground=self.colors['white'])
    
    def create_widgets(self):
        """Crear widgets de reportes modernos"""
        # Frame principal
        main_frame = ttk.Frame(self.window, style='Card.TFrame', padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header
        self.create_header(main_frame)
        
        # Filtros
        self.create_filters_section(main_frame)
        
        # Estadísticas
        self.create_stats_section(main_frame)
        
        # Resultados
        self.create_results_section(main_frame)
        
        # Cargar datos iniciales
        self.buscar_registros()
    
    def create_header(self, parent):
        """Crear header moderno"""
        header_frame = ttk.Frame(parent, style='Card.TFrame')
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = ttk.Label(header_frame, 
                               text="📊 Reportes de Asistencia", 
                               style='Title.TLabel')
        title_label.pack(pady=15)
        
        subtitle_label = ttk.Label(header_frame, 
                                  text="Consulta y exporta registros de asistencia", 
                                  style='Subtitle.TLabel')
        subtitle_label.pack(pady=(0, 15))
    
    def create_filters_section(self, parent):
        """Crear sección de filtros moderna"""
        filters_frame = ttk.LabelFrame(parent, text="🔍 Filtros de Búsqueda", padding="15")
        filters_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Contenedor de filtros
        filters_container = ttk.Frame(filters_frame)
        filters_container.pack(fill=tk.X)
        
        # Filtros en una fila
        ttk.Label(filters_container, text="📅 Fecha desde:", 
                 font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        
        self.fecha_desde = tk.StringVar(value=date.today().strftime('%Y-%m-%d'))
        fecha_desde_entry = ttk.Entry(filters_container, textvariable=self.fecha_desde, 
                                     width=12, font=('Segoe UI', 10))
        fecha_desde_entry.pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Label(filters_container, text="📅 Fecha hasta:", 
                 font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        
        self.fecha_hasta = tk.StringVar(value=date.today().strftime('%Y-%m-%d'))
        fecha_hasta_entry = ttk.Entry(filters_container, textvariable=self.fecha_hasta, 
                                     width=12, font=('Segoe UI', 10))
        fecha_hasta_entry.pack(side=tk.LEFT, padx=(0, 20))
        
        # Botones de acción
        ttk.Button(filters_container, text="🔍 Buscar", 
                  style='Primary.TButton',
                  command=self.buscar_registros).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(filters_container, text="📄 Exportar Excel", 
                  style='Success.TButton',
                  command=self.exportar_excel).pack(side=tk.LEFT, padx=(0, 10))
        
        # Botones de períodos rápidos
        periods_frame = ttk.Frame(filters_container)
        periods_frame.pack(side=tk.LEFT, padx=(20, 0))
        
        ttk.Button(periods_frame, text="📅 Hoy", 
                  style='Warning.TButton',
                  command=self.mostrar_hoy).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(periods_frame, text="📊 Esta Semana", 
                  style='Warning.TButton',
                  command=self.mostrar_semana).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(periods_frame, text="📈 Este Mes", 
                  style='Warning.TButton',
                  command=self.mostrar_mes).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(periods_frame, text="📋 Todos", 
                  style='Warning.TButton',
                  command=self.mostrar_todos).pack(side=tk.LEFT)
    
    def create_stats_section(self, parent):
        """Crear sección de estadísticas"""
        stats_frame = ttk.LabelFrame(parent, text="📊 Estadísticas", padding="15")
        stats_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Frame para estadísticas
        stats_container = ttk.Frame(stats_frame)
        stats_container.pack(fill=tk.X)
        
        # Estadísticas por empleado
        self.stats_label = ttk.Label(stats_container, 
                                   text="Selecciona un período para ver estadísticas", 
                                   style='Info.TLabel')
        self.stats_label.pack()
    
    def actualizar_estadisticas(self, registros):
        """Actualizar estadísticas basadas en los registros"""
        try:
            if not registros:
                self.stats_label.config(text="No hay registros para mostrar estadísticas")
                return
            
            # Calcular estadísticas
            total_horas = 0
            registros_completos = 0
            empleados_unicos = set()
            
            for registro in registros:
                nombre, fecha, entrada, salida = registro
                empleados_unicos.add(nombre)
                
                if entrada and salida:
                    try:
                        # Convertir a datetime
                        if isinstance(entrada, str):
                            entrada_dt = datetime.strptime(entrada, '%Y-%m-%d %H:%M:%S.%f')
                        else:
                            entrada_dt = entrada
                        
                        if isinstance(salida, str):
                            salida_dt = datetime.strptime(salida, '%Y-%m-%d %H:%M:%S.%f')
                        else:
                            salida_dt = salida
                        
                        tiempo_trabajado = salida_dt - entrada_dt
                        horas = tiempo_trabajado.total_seconds() / 3600
                        total_horas += horas
                        registros_completos += 1
                    except:
                        pass
            
            # Mostrar estadísticas
            stats_text = f"""
📊 Estadísticas del Período:
👥 Empleados: {len(empleados_unicos)}
📅 Registros completos: {registros_completos}
⏱️ Total horas trabajadas: {total_horas:.2f} horas
📈 Promedio por empleado: {total_horas/len(empleados_unicos):.2f} horas
            """.strip()
            
            self.stats_label.config(text=stats_text)
            
        except Exception as e:
            self.stats_label.config(text=f"Error calculando estadísticas: {str(e)}")
    
    def create_results_section(self, parent):
        """Crear sección de resultados moderna"""
        results_frame = ttk.LabelFrame(parent, text="📋 Resultados", padding="15")
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header de resultados
        results_header = ttk.Frame(results_frame)
        results_header.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(results_header, text="Registros encontrados", 
                 style='Subtitle.TLabel').pack(side=tk.LEFT)
        
        # Contador de resultados
        self.results_count_label = ttk.Label(results_header, 
                                            text="0 registros", 
                                            font=('Segoe UI', 10),
                                            foreground=self.colors['secondary'])
        self.results_count_label.pack(side=tk.RIGHT)
        
        # Treeview moderno
        tree_frame = ttk.Frame(results_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('empleado', 'fecha', 'entrada', 'salida', 'horas_trabajadas', 'estado')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        # Configurar columnas con iconos
        self.tree.heading('empleado', text='👤 Empleado')
        self.tree.heading('fecha', text='📅 Fecha')
        self.tree.heading('entrada', text='⏰ Entrada')
        self.tree.heading('salida', text='🚪 Salida')
        self.tree.heading('horas_trabajadas', text='⏱️ Horas Trabajadas')
        self.tree.heading('estado', text='📊 Estado')
        
        self.tree.column('empleado', width=220, minwidth=200)
        self.tree.column('fecha', width=100, minwidth=100)
        self.tree.column('entrada', width=100, minwidth=100)
        self.tree.column('salida', width=100, minwidth=100)
        self.tree.column('horas_trabajadas', width=140, minwidth=120)
        self.tree.column('estado', width=120, minwidth=100)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Configurar colores alternados
        self.tree.tag_configure('oddrow', background=self.colors['light'])
        self.tree.tag_configure('evenrow', background=self.colors['white'])
        self.tree.tag_configure('complete', background='#d5f4e6')
        self.tree.tag_configure('incomplete', background='#fadbd8')
    
    def buscar_registros(self):
        """Buscar registros según filtros con mejor formato"""
        try:
            # Limpiar treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            fecha_desde = self.fecha_desde.get()
            fecha_hasta = self.fecha_hasta.get()
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT e.nombre_completo, a.fecha, a.hora_entrada, a.hora_salida
                FROM asistencias a
                JOIN empleados e ON a.empleado_id = e.id
                WHERE a.fecha BETWEEN ? AND ?
                ORDER BY a.fecha DESC, e.nombre_completo
            """, (fecha_desde, fecha_hasta))
            
            registros = cursor.fetchall()
            conn.close()
            
            # Actualizar contador
            self.results_count_label.config(text=f"{len(registros)} registros")
            
            # Actualizar estadísticas
            self.actualizar_estadisticas(registros)
            
            # Insertar en treeview
            for i, registro in enumerate(registros):
                nombre, fecha, entrada, salida = registro
                
                # Calcular horas trabajadas
                horas_trabajadas = "-"
                estado = "Incompleto"
                
                try:
                    if entrada and salida:
                        # Convertir strings a datetime si es necesario
                        if isinstance(entrada, str):
                            # Intentar diferentes formatos de fecha
                            try:
                                entrada_dt = datetime.strptime(entrada, '%Y-%m-%d %H:%M:%S.%f')
                            except ValueError:
                                try:
                                    entrada_dt = datetime.strptime(entrada, '%Y-%m-%d %H:%M:%S')
                                except ValueError:
                                    entrada_dt = datetime.fromisoformat(entrada.replace('Z', '+00:00'))
                        else:
                            entrada_dt = entrada
                        
                        if isinstance(salida, str):
                            # Intentar diferentes formatos de fecha
                            try:
                                salida_dt = datetime.strptime(salida, '%Y-%m-%d %H:%M:%S.%f')
                            except ValueError:
                                try:
                                    salida_dt = datetime.strptime(salida, '%Y-%m-%d %H:%M:%S')
                                except ValueError:
                                    salida_dt = datetime.fromisoformat(salida.replace('Z', '+00:00'))
                        else:
                            salida_dt = salida
                        
                        tiempo_trabajado = salida_dt - entrada_dt
                        horas = tiempo_trabajado.total_seconds() / 3600
                        horas_trabajadas = f"{horas:.2f} horas"
                        estado = "✅ Completo"
                    elif entrada:
                        estado = "⏰ Solo entrada"
                except Exception as e:
                    print(f"Error calculando horas: {e}")
                    print(f"Entrada: {entrada}, Salida: {salida}")
                    horas_trabajadas = "❌ Error"
                
                # Formatear fechas
                try:
                    if isinstance(fecha, str):
                        fecha_str = fecha
                    else:
                        fecha_str = fecha.strftime('%d/%m/%Y')
                except:
                    fecha_str = str(fecha)
                
                try:
                    if isinstance(entrada, str):
                        # Formatear entrada para mostrar solo hora
                        try:
                            entrada_dt = datetime.strptime(entrada, '%Y-%m-%d %H:%M:%S.%f')
                            entrada_str = entrada_dt.strftime('%H:%M:%S')
                        except ValueError:
                            try:
                                entrada_dt = datetime.strptime(entrada, '%Y-%m-%d %H:%M:%S')
                                entrada_str = entrada_dt.strftime('%H:%M:%S')
                            except ValueError:
                                entrada_str = entrada
                    elif entrada:
                        entrada_str = entrada.strftime('%H:%M:%S')
                    else:
                        entrada_str = '-'
                except:
                    entrada_str = str(entrada) if entrada else '-'
                
                try:
                    if isinstance(salida, str):
                        # Formatear salida para mostrar solo hora
                        try:
                            salida_dt = datetime.strptime(salida, '%Y-%m-%d %H:%M:%S.%f')
                            salida_str = salida_dt.strftime('%H:%M:%S')
                        except ValueError:
                            try:
                                salida_dt = datetime.strptime(salida, '%Y-%m-%d %H:%M:%S')
                                salida_str = salida_dt.strftime('%H:%M:%S')
                            except ValueError:
                                salida_str = salida
                    elif salida:
                        salida_str = salida.strftime('%H:%M:%S')
                    else:
                        salida_str = '-'
                except:
                    salida_str = str(salida) if salida else '-'
                
                # Aplicar tags según el estado
                if "Completo" in estado:
                    tags = ('complete',)
                elif "Solo entrada" in estado:
                    tags = ('incomplete',)
                else:
                    tags = ('evenrow' if i % 2 == 0 else 'oddrow',)
                
                self.tree.insert('', 'end', 
                               values=(nombre, fecha_str, entrada_str, salida_str, horas_trabajadas, estado),
                               tags=tags)
                
        except Exception as e:
            messagebox.showerror("Error", f"Error buscando registros: {str(e)}")
    
    def mostrar_hoy(self):
        """Mostrar registros de hoy"""
        from datetime import date
        hoy = date.today()
        self.fecha_desde.set(hoy.strftime('%Y-%m-%d'))
        self.fecha_hasta.set(hoy.strftime('%Y-%m-%d'))
        self.buscar_registros()
    
    def mostrar_semana(self):
        """Mostrar registros de esta semana"""
        from datetime import date, timedelta
        hoy = date.today()
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        fin_semana = inicio_semana + timedelta(days=6)
        self.fecha_desde.set(inicio_semana.strftime('%Y-%m-%d'))
        self.fecha_hasta.set(fin_semana.strftime('%Y-%m-%d'))
        self.buscar_registros()
    
    def mostrar_mes(self):
        """Mostrar registros de este mes"""
        from datetime import date
        hoy = date.today()
        inicio_mes = date(hoy.year, hoy.month, 1)
        if hoy.month == 12:
            fin_mes = date(hoy.year + 1, 1, 1) - timedelta(days=1)
        else:
            fin_mes = date(hoy.year, hoy.month + 1, 1) - timedelta(days=1)
        self.fecha_desde.set(inicio_mes.strftime('%Y-%m-%d'))
        self.fecha_hasta.set(fin_mes.strftime('%Y-%m-%d'))
        self.buscar_registros()
    
    def mostrar_todos(self):
        """Mostrar todos los registros"""
        self.fecha_desde.set('2020-01-01')  # Fecha muy antigua
        self.fecha_hasta.set('2030-12-31')  # Fecha muy futura
        self.buscar_registros()
    
    def exportar_excel(self):
        """Exportar reporte a Excel con mejor formato"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Asistencia"
            
            # Encabezados con mejor formato
            headers = ['Empleado', 'Fecha', 'Entrada', 'Salida', 'Horas Trabajadas', 'Estado']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['primary'], 
                                       end_color=self.colors['primary'], 
                                       fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Obtener datos del treeview
            row = 2
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
            
            # Ajustar ancho de columnas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Guardar archivo
            from tkinter import filedialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar reporte como"
            )
            
            if filename:
                wb.save(filename)
                messagebox.showinfo("✅ Éxito", f"Reporte exportado a:\n{filename}")
                
        except Exception as e:
            messagebox.showerror("❌ Error", f"Error exportando reporte: {str(e)}")
