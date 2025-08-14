#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Inventario de Almac√©n - M√≥dulo Independiente
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import os
from datetime import datetime, date
import sqlite3
import csv

class InventarioAlmacenWindow:
    def __init__(self, parent, main_window=None):
        self.parent = parent
        self.main_window = main_window
        self.setup_database()
        
        # Crear ventana principal
        self.window = tk.Toplevel(parent)
        self.window.title("Sistema de Inventario - ALMAC√âN")
        self.window.geometry("1200x800")
        self.window.configure(bg='#f0f0f0')
        
        # Configurar ventana
        self.window.resizable(True, True)
        self.window.minsize(1000, 600)
        
        # Centrar ventana
        self.center_window()
        
        # Hacer modal
        self.window.transient(parent)
        self.window.grab_set()
        
        # Configurar estilos
        self.configurar_estilos()
        
        # Crear interfaz
        self.create_interface()
        
        # Cargar datos autom√°ticamente
        total_cargados = self.cargar_inventario()
        
        # Mostrar informaci√≥n de carga si es exitoso
        if total_cargados and total_cargados > 5:
            self.window.after(2000, lambda: self.mostrar_info_carga(total_cargados))
    
    def configurar_estilos(self):
        """Configurar estilos para almac√©n"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Estilo espec√≠fico para almac√©n
        style.configure('Almacen.Treeview',
                       background='#ffffff',
                       foreground='#2c3e50',
                       fieldbackground='#ffffff',
                       font=('Arial', 9),
                       rowheight=25)
        
        # Headers con color espec√≠fico para almac√©n
        style.configure('Almacen.Treeview.Heading',
                       background='#3498db',  # Azul para almac√©n
                       foreground='white',
                       font=('Arial', 9, 'bold'),
                       relief='flat')
        
        style.map('Almacen.Treeview',
                 background=[('selected', '#2980b9')],
                 foreground=[('selected', '#ffffff')])
    
    def center_window(self):
        """Centrar ventana"""
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (1200 // 2)
        y = (self.window.winfo_screenheight() // 2) - (800 // 2)
        self.window.geometry(f"1200x800+{x}+{y}")
    
    def setup_database(self):
        """Configurar base de datos espec√≠fica para almac√©n"""
        try:
            db_dir = 'database'
            os.makedirs(db_dir, exist_ok=True)
            
            db_path = os.path.join(db_dir, 'inventario_almacen.db')
            self.conn = sqlite3.connect(db_path)
            self.crear_tablas()
            self.poblar_datos_almacen()
        except Exception as e:
            print(f"Error BD almac√©n: {e}")
            self.conn = sqlite3.connect('inventario_almacen.db')
            self.crear_tablas()
            self.poblar_datos_almacen()
    
    def crear_tablas(self):
        """Crear tablas espec√≠ficas para almac√©n"""
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos_almacen (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE NOT NULL,
                nombre TEXT NOT NULL,
                saldo INTEGER DEFAULT 0,
                unidad TEXT NOT NULL,
                valor_unitario REAL NOT NULL,
                stock_minimo INTEGER DEFAULT 0,
                ubicacion TEXT,
                proveedor TEXT,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS movimientos_almacen (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                producto_id INTEGER,
                tipo TEXT NOT NULL,
                cantidad INTEGER NOT NULL,
                fecha DATE NOT NULL,
                factura TEXT,
                proveedor TEXT,
                destino TEXT,
                valor_total REAL,
                observaciones TEXT,
                fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (producto_id) REFERENCES productos_almacen (id)
            )
        ''')
        
        self.conn.commit()
    
    def poblar_datos_almacen(self):
        """Cargar datos espec√≠ficos de almac√©n"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM productos_almacen")
        
        if cursor.fetchone()[0] == 0:
            print("üîÑ Cargando datos de ALMAC√âN desde Excel...")
            
            # Intentar cargar desde Excel
            if self.cargar_desde_excel_almacen():
                print("‚úÖ Datos de ALMAC√âN cargados desde Excel")
                return
            
            # Datos de ejemplo espec√≠ficos para almac√©n
            productos_almacen = [
                ('ALM001', 'ACEITE 15W 40 AL', 3, 'GALONES', 45000, 2, 'A-01', 'LUBRICANTES SA'),
                ('ALM002', 'ACOPLE ESPIGO 3/4', 13, 'UND', 8500, 5, 'A-02', 'FERRETERIA CENTRAL'),
                ('ALM003', 'ACOPLES FOSTER 17$ BRONCE', 2, 'UND', 17000, 3, 'A-03', 'FOSTER LTDA'),
                ('ALM004', 'ARANDELA CUADRADA 4*4', 0, 'UND', 500, 10, 'A-04', 'TORNILLERIA BOGOTA'),
                ('ALM005', 'TORNILLO HEXAGONAL 1/4 x 2', 100, 'UND', 750, 20, 'A-05', 'FERRETERIA CENTRAL'),
                ('ALM006', 'FILTRO DE AIRE MOTOR', 5, 'UND', 25000, 3, 'A-06', 'REPUESTOS DIESEL'),
                ('ALM007', 'CABLE EL√âCTRICO 12 AWG', 200, 'METROS', 2500, 50, 'A-07', 'ELECTRICOS DEL VALLE'),
                ('ALM008', 'SOLDADURA E6013 3.2MM', 15, 'KG', 8500, 5, 'A-08', 'SOLDADURAS T√âCNICAS')
            ]
            
            cursor.executemany('''
                INSERT INTO productos_almacen (codigo, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', productos_almacen)
            
            self.conn.commit()
            print("üìã Productos de ALMAC√âN inicializados")
    
    def cargar_desde_excel_almacen(self):
        """Cargar espec√≠ficamente desde archivo de almac√©n"""
        archivo_almacen = 'INVENTARIO ALMACEN JUNIO  2025 xlsx.xlsx'
        
        if not os.path.exists(archivo_almacen):
            print(f"‚ùå {archivo_almacen} no encontrado")
            return False
        
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(archivo_almacen, data_only=True)
            sheet = workbook['Almacen'] if 'Almacen' in workbook.sheetnames else workbook.active
            
            productos = self.procesar_hoja_almacen(sheet)
            
            if productos:
                cursor = self.conn.cursor()
                cursor.executemany('''
                    INSERT OR REPLACE INTO productos_almacen 
                    (codigo, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', productos)
                
                self.conn.commit()
                print(f"‚úÖ {len(productos)} productos de ALMAC√âN cargados desde Excel")
                return True
            
        except Exception as e:
            print(f"Error cargando Excel de almac√©n: {e}")
            return False
        
        return False
    
    def procesar_hoja_almacen(self, sheet):
        """Procesar hoja espec√≠fica de almac√©n"""
        productos = []
        contador = 1
        
        print("   üìã Procesando productos de ALMAC√âN...")
        
        for row_num in range(1, min(sheet.max_row + 1, 200)):
            try:
                # Buscar productos v√°lidos en diferentes columnas
                for col in range(1, 6):
                    cell_value = sheet.cell(row=row_num, column=col).value
                    
                    if (cell_value and isinstance(cell_value, str) and 
                        len(str(cell_value).strip()) > 3):
                        
                        nombre = str(cell_value).strip().upper()
                        
                        # Filtrar headers y texto no deseado
                        if (not any(skip in nombre for skip in ['FECHA', 'FACTURA', 'PROVE', 'CANT', 'VALOR', 'TOTAL', 'DIA', 'SALDO MES']) and
                            nombre not in ['FECHA', 'N. FACTURA', 'PROVE'] and
                            len(nombre) > 2):
                            
                            # Buscar saldo en columnas adyacentes
                            saldo = 0
                            unidad = 'UND'
                            
                            for saldo_col in range(col + 1, min(col + 4, 7)):
                                saldo_val = sheet.cell(row=row_num, column=saldo_col).value
                                if isinstance(saldo_val, (int, float)) and saldo_val >= 0:
                                    saldo = int(saldo_val)
                                    
                                    # Buscar unidad
                                    unidad_val = sheet.cell(row=row_num, column=saldo_col + 1).value
                                    if isinstance(unidad_val, str) and len(str(unidad_val).strip()) <= 10:
                                        unidad = str(unidad_val).strip().upper()
                                    break
                            
                            # Crear producto
                            codigo = f'ALM{contador:03d}'
                            valor_unitario = self.estimar_valor_almacen(nombre)
                            stock_minimo = max(1, int(saldo * 0.2)) if saldo > 0 else 5
                            
                            productos.append((
                                codigo, nombre, saldo, unidad,
                                valor_unitario, stock_minimo, f'ALM-{contador:02d}', 'PROVEEDOR ALMACEN'
                            ))
                            
                            contador += 1
                            
                            if contador <= 10:
                                print(f"     ‚Ä¢ {nombre} - {saldo} {unidad}")
                            
                            break
                    
            except Exception as e:
                continue
        
        print(f"   ‚úÖ {len(productos)} productos de ALMAC√âN procesados")
        return productos
    
    def estimar_valor_almacen(self, nombre):
        """Estimar valor para productos de almac√©n"""
        nombre_lower = nombre.lower()
        
        if 'aceite' in nombre_lower:
            return 45000
        elif 'acople' in nombre_lower or 'foster' in nombre_lower:
            return 15000
        elif 'arandela' in nombre_lower or 'tornillo' in nombre_lower:
            return 500
        elif 'filtro' in nombre_lower:
            return 25000
        elif 'cable' in nombre_lower:
            return 2500
        elif 'soldadura' in nombre_lower:
            return 8500
        else:
            return 5000
    
    def create_interface(self):
        """Crear interfaz espec√≠fica para almac√©n"""
        # Header espec√≠fico para almac√©n
        self.create_almacen_header()
        
        # Contenedor principal
        main_container = tk.Frame(self.window, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Layout: sidebar + contenido
        self.create_almacen_sidebar(main_container)
        self.create_almacen_content(main_container)
    
    def create_almacen_header(self):
        """Header espec√≠fico para almac√©n"""
        header = tk.Frame(self.window, bg='#3498db', height=70)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        header_content = tk.Frame(header, bg='#3498db')
        header_content.pack(fill=tk.BOTH, expand=True, padx=30, pady=15)
        
        # Logo y t√≠tulo espec√≠fico
        title_frame = tk.Frame(header_content, bg='#3498db')
        title_frame.pack(side=tk.LEFT)
        
        icon_label = tk.Label(title_frame, text="üè≠", font=('Arial', 24), 
                             bg='#3498db', fg='#ffffff')
        icon_label.pack(side=tk.LEFT, padx=(0, 15))
        
        title_container = tk.Frame(title_frame, bg='#3498db')
        title_container.pack(side=tk.LEFT)
        
        title_label = tk.Label(title_container, text="Inventario de ALMAC√âN", 
                              font=('Segoe UI', 18, 'bold'), bg='#3498db', fg='#ffffff')
        title_label.pack(anchor='w')
        
        subtitle = tk.Label(title_container, text="Gesti√≥n de productos de almac√©n general", 
                           font=('Segoe UI', 10), bg='#3498db', fg='#ecf0f1')
        subtitle.pack(anchor='w')
        
        # Botones espec√≠ficos para almac√©n
        self.create_almacen_actions(header_content)
    
    def create_almacen_actions(self, parent):
        """Botones espec√≠ficos para almac√©n"""
        actions_frame = tk.Frame(parent, bg='#3498db')
        actions_frame.pack(side=tk.RIGHT)
        
        buttons = [
            ("+ Nuevo", self.nuevo_producto_almacen, "#27ae60", "Nuevo producto de almac√©n"),
            ("‚Üì Entrada", lambda: self.movimiento('entrada'), "#16a085", "Entrada de almac√©n"),
            ("‚Üë Salida", lambda: self.movimiento('salida'), "#f39c12", "Salida de almac√©n"),
            ("üìä", self.reporte_almacen, "#8e44ad", "Reporte de almac√©n"),
            ("‚úï", self.cerrar_ventana, "#e74c3c", "Cerrar")
        ]
        
        for text, command, color, tooltip in buttons:
            btn = tk.Button(actions_frame, text=text, command=command,
                           bg=color, fg='white', font=('Segoe UI', 9, 'bold'),
                           relief='flat', bd=0, padx=12, pady=6, cursor='hand2')
            btn.pack(side=tk.LEFT, padx=3)
            self.add_hover_effect(btn, color)
    
    def create_almacen_sidebar(self, parent):
        """Sidebar espec√≠fico para almac√©n"""
        sidebar = tk.Frame(parent, bg='#ffffff', width=280)
        sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=(20, 0), pady=20)
        sidebar.pack_propagate(False)
        
        # T√≠tulo
        sidebar_title = tk.Label(sidebar, text="üìä Control de Almac√©n", 
                                font=('Segoe UI', 16, 'bold'), 
                                bg='#ffffff', fg='#2c3e50')
        sidebar_title.pack(anchor='w', padx=20, pady=(20, 15))
        
        # Estad√≠sticas espec√≠ficas
        self.create_almacen_stats(sidebar)
        
        # Filtros espec√≠ficos
        self.create_almacen_filters(sidebar)
        
        # Acciones r√°pidas espec√≠ficas
        self.create_almacen_quick_actions(sidebar)
    
    def create_almacen_stats(self, parent):
        """Estad√≠sticas espec√≠ficas de almac√©n"""
        stats_container = tk.Frame(parent, bg='#ffffff')
        stats_container.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Obtener estad√≠sticas de almac√©n
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM productos_almacen")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM productos_almacen WHERE saldo <= stock_minimo AND saldo > 0")
        bajo = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM productos_almacen WHERE saldo = 0")
        agotado = cursor.fetchone()[0]
        
        cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos_almacen")
        valor = cursor.fetchone()[0] or 0
        
        # Cards espec√≠ficas para almac√©n
        cards_data = [
            ("Productos Almac√©n", str(total), "#3498db", "#ebf3fd"),
            ("Stock Bajo", str(bajo), "#f39c12", "#fef9e7"),
            ("Sin Stock", str(agotado), "#e74c3c", "#fdedec"),
            ("Valor Total", f"${valor:,.0f}", "#27ae60", "#eafaf1")
        ]
        
        for title, value, color, bg_color in cards_data:
            card = tk.Frame(stats_container, bg=bg_color, relief='solid', bd=1)
            card.pack(fill=tk.X, pady=8)
            
            card_content = tk.Frame(card, bg=bg_color)
            card_content.pack(fill=tk.BOTH, padx=15, pady=12)
            
            value_label = tk.Label(card_content, text=value, 
                                  font=('Segoe UI', 18, 'bold'),
                                  bg=bg_color, fg=color)
            value_label.pack(anchor='w')
            
            title_label = tk.Label(card_content, text=title,
                                  font=('Segoe UI', 10),
                                  bg=bg_color, fg='#7f8c8d')
            title_label.pack(anchor='w')
    
    def create_almacen_filters(self, parent):
        """Filtros espec√≠ficos para almac√©n"""
        filters_frame = tk.Frame(parent, bg='#ffffff')
        filters_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        tk.Label(filters_frame, text="üîç Filtros de Almac√©n", 
                font=('Segoe UI', 14, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(anchor='w', pady=(0, 10))
        
        # B√∫squeda espec√≠fica
        search_frame = tk.Frame(filters_frame, bg='#ffffff')
        search_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(search_frame, text="Buscar en almac√©n:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var,
                               font=('Segoe UI', 10), relief='solid', bd=1)
        search_entry.pack(fill=tk.X, pady=(3, 0))
        
        # Filtro por ubicaci√≥n
        ubicacion_frame = tk.Frame(filters_frame, bg='#ffffff')
        ubicacion_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(ubicacion_frame, text="Ubicaci√≥n:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.filter_ubicacion = tk.StringVar(value="Todas")
        self.filter_ubicacion.trace('w', self.on_search_change)
        ubicacion_combo = ttk.Combobox(ubicacion_frame, textvariable=self.filter_ubicacion,
                                      values=["Todas", "A-01", "A-02", "A-03", "A-04", "A-05"],
                                      state="readonly", font=('Segoe UI', 10))
        ubicacion_combo.pack(fill=tk.X, pady=(3, 0))
        
        # Filtro por estado de stock
        stock_frame = tk.Frame(filters_frame, bg='#ffffff')
        stock_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(stock_frame, text="Estado del stock:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.filter_stock = tk.StringVar(value="Todos")
        self.filter_stock.trace('w', self.on_search_change)
        stock_combo = ttk.Combobox(stock_frame, textvariable=self.filter_stock,
                                  values=["Todos", "Normal", "Bajo", "Agotado"],
                                  state="readonly", font=('Segoe UI', 10))
        stock_combo.pack(fill=tk.X, pady=(3, 0))
        
        # Bot√≥n limpiar filtros
        clear_btn = tk.Button(filters_frame, text="üóëÔ∏è Limpiar", 
                             command=self.limpiar_filtros,
                             bg='#95a5a6', fg='white', font=('Segoe UI', 9),
                             relief='flat', bd=0, pady=8, cursor='hand2')
        clear_btn.pack(fill=tk.X, pady=(10, 0))
    
    def create_almacen_quick_actions(self, parent):
        """Acciones r√°pidas espec√≠ficas para almac√©n"""
        actions_frame = tk.Frame(parent, bg='#ffffff')
        actions_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        tk.Label(actions_frame, text="‚ö° Acciones de Almac√©n", 
                font=('Segoe UI', 14, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(anchor='w', pady=(0, 15))
        
        actions = [
            ("‚úèÔ∏è Editar Producto", self.editar_producto, "#3498db"),        # Azul cielo
            ("üì• Entrada de Stock", lambda: self.movimiento_entrada(), "#27ae60"),  # Verde esmeralda
            ("üì§ Salida de Stock", lambda: self.movimiento_salida(), "#e67e22"),    # Naranja suave
            ("üìã Historial Movimientos", lambda: self.ver_historial_movimientos(), "#8e44ad"),  # P√∫rpura
            ("üì¶ Control de Stock", self.control_stock, "#7f8c8d"),         # Gris azulado
            ("üìã Lista de Pedidos", self.lista_pedidos, "#16a085"),         # Verde azulado
            ("üîÑ Actualizar Datos", self.actualizar_datos, "#2c3e50")       # Azul oscuro
        ]
        
        for text, command, color in actions:
            btn = tk.Button(actions_frame, text=text, command=command,
                           bg=color, fg='white', font=('Segoe UI', 10),
                           relief='flat', bd=0, pady=8, cursor='hand2')
            btn.pack(fill=tk.X, pady=3)
            self.add_hover_effect(btn, color)
    
    def create_almacen_content(self, parent):
        """Contenido principal espec√≠fico para almac√©n"""
        content_area = tk.Frame(parent, bg='#f8f9fa')
        content_area.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Card para la tabla de almac√©n
        table_card = tk.Frame(content_area, bg='#ffffff', relief='solid', bd=1)
        table_card.pack(fill=tk.BOTH, expand=True)
        
        # Header espec√≠fico
        table_header = tk.Frame(table_card, bg='#ffffff', height=50)
        table_header.pack(fill=tk.X)
        table_header.pack_propagate(False)
        
        tk.Label(table_header, text="üìã Productos de Almac√©n", 
                font=('Segoe UI', 16, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(side=tk.LEFT, padx=20, pady=15)
        
        self.products_count_label = tk.Label(table_header, text="", 
                                           font=('Segoe UI', 10),
                                           bg='#ffffff', fg='#7f8c8d')
        self.products_count_label.pack(side=tk.RIGHT, padx=20, pady=15)
        
        # Tabla espec√≠fica para almac√©n
        self.create_almacen_table(table_card)
    
    def create_almacen_table(self, parent):
        """Tabla espec√≠fica para almac√©n"""
        table_container = tk.Frame(parent, bg='#ffffff')
        table_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # TreeView espec√≠fico para almac√©n
        columns = ('C√≥digo', 'Producto', 'Saldo', 'Unidad', 'Valor Unit.', 'Stock Min.', 'Ubicaci√≥n', 'Estado')
        self.tree = ttk.Treeview(table_container, columns=columns, show='headings',
                                height=20, style='Almacen.Treeview')
        
        # Configurar columnas espec√≠ficas
        column_config = {
            'C√≥digo': {'width': 80, 'anchor': 'center'},
            'Producto': {'width': 250, 'anchor': 'w'},
            'Saldo': {'width': 80, 'anchor': 'center'},
            'Unidad': {'width': 80, 'anchor': 'center'},
            'Valor Unit.': {'width': 100, 'anchor': 'e'},
            'Stock Min.': {'width': 80, 'anchor': 'center'},
            'Ubicaci√≥n': {'width': 80, 'anchor': 'center'},
            'Estado': {'width': 100, 'anchor': 'center'}
        }
        
        for col, config in column_config.items():
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c))
            self.tree.column(col, width=config['width'], anchor=config['anchor'])
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(table_container, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Layout
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        table_container.grid_rowconfigure(0, weight=1)
        table_container.grid_columnconfigure(0, weight=1)
        
        # Eventos espec√≠ficos
        self.tree.bind('<Double-1>', lambda e: self.editar_producto())
        self.tree.bind('<Button-3>', self.mostrar_menu_contextual)
    
    # =================== FUNCIONES ESPEC√çFICAS DE ALMAC√âN ===================
    
    def cargar_inventario(self):
        """Cargar inventario espec√≠fico de almac√©n"""
        try:
            # Limpiar datos existentes
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Query espec√≠fica para almac√©n
            query = "SELECT codigo, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion FROM productos_almacen WHERE 1=1"
            params = []
            
            # Aplicar filtros espec√≠ficos
            if hasattr(self, 'search_var'):
                texto = self.search_var.get().strip().lower()
                if texto:
                    query += " AND (LOWER(nombre) LIKE ? OR LOWER(codigo) LIKE ?)"
                    params.extend([f'%{texto}%', f'%{texto}%'])
            
            if hasattr(self, 'filter_ubicacion'):
                ubicacion = self.filter_ubicacion.get()
                if ubicacion and ubicacion != "Todas":
                    query += " AND ubicacion = ?"
                    params.append(ubicacion)
            
            if hasattr(self, 'filter_stock'):
                stock_filter = self.filter_stock.get()
                if stock_filter == "Bajo":
                    query += " AND saldo <= stock_minimo AND saldo > 0"
                elif stock_filter == "Agotado":
                    query += " AND saldo = 0"
                elif stock_filter == "Normal":
                    query += " AND saldo > stock_minimo"
            
            query += " ORDER BY codigo"
            
            # Ejecutar query
            cursor = self.conn.cursor()
            cursor.execute(query, params)
            productos_mostrados = 0
            
            for row in cursor.fetchall():
                codigo, nombre, saldo, unidad, valor_unit, stock_min, ubicacion = row
                
                # Estados espec√≠ficos para almac√©n
                if saldo == 0:
                    estado = "üî¥ AGOTADO"
                    tags = ('agotado',)
                elif saldo <= stock_min:
                    estado = "üü° STOCK BAJO"
                    tags = ('bajo',)
                else:
                    estado = "üü¢ NORMAL"
                    tags = ('normal',)
                
                valor_formateado = f"${valor_unit:,.0f}"
                
                self.tree.insert('', tk.END, values=(
                    codigo, nombre, saldo, unidad, valor_formateado,
                    stock_min, ubicacion or "N/A", estado
                ), tags=tags)
                productos_mostrados += 1
            
            # Configurar colores espec√≠ficos
            self.tree.tag_configure('agotado', background='#ffebee', foreground='#d32f2f')
            self.tree.tag_configure('bajo', background='#fff3e0', foreground='#f57c00')
            self.tree.tag_configure('normal', background='#ffffff', foreground='#2c3e50')
            
            # Actualizar contador
            if hasattr(self, 'products_count_label'):
                cursor.execute("SELECT COUNT(*) FROM productos_almacen")
                total = cursor.fetchone()[0]
                self.products_count_label.config(
                    text=f"Mostrando {productos_mostrados} de {total} productos"
                )
            
            return total
            
        except Exception as e:
            print(f"Error cargando inventario de almac√©n: {e}")
            return 0
    
    def on_search_change(self, *args):
        """Actualizar cuando cambien los filtros"""
        self.cargar_inventario()
    
    def limpiar_filtros(self):
        """Limpiar filtros espec√≠ficos"""
        if hasattr(self, 'search_var'):
            self.search_var.set("")
        if hasattr(self, 'filter_ubicacion'):
            self.filter_ubicacion.set("Todas")
        if hasattr(self, 'filter_stock'):
            self.filter_stock.set("Todos")
    
    def mostrar_info_carga(self, total):
        """Mostrar informaci√≥n de carga espec√≠fica"""
        messagebox.showinfo(
            "üè≠ Almac√©n Cargado", 
            f"¬°Sistema de Almac√©n listo!\n\n" +
            f"üì¶ Productos de almac√©n: {total}\n" +
            f"üè™ Ubicaciones disponibles\n" +
            f"üìä Control de stock activo\n\n" +
            "Sistema espec√≠fico para gesti√≥n\n" +
            "de productos de almac√©n general."
        )
    
    # =================== FUNCIONES DE ACCIONES ===================
    
    def nuevo_producto_almacen(self):
        """Crear nuevo producto espec√≠fico de almac√©n"""
        ProductoAlmacenWindow(self.window, self, modo="nuevo")
    
    def editar_producto(self):
        """Editar producto de almac√©n"""
        producto = self.get_selected_producto()
        if producto:
            ProductoAlmacenWindow(self.window, self, modo="editar", producto=producto)
    
    def get_selected_producto(self):
        """Obtener producto seleccionado"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Selecciona un producto de almac√©n")
            return None
        
        item = self.tree.item(selection[0])
        codigo = item['values'][0]
        
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM productos_almacen WHERE codigo = ?", (codigo,))
        return cursor.fetchone()
    
    def movimiento_entrada(self):
        """Abrir di√°logo de entrada de stock"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import MovimientoInventarioDialog
            dialog = MovimientoInventarioDialog(self.window, 'almacen', 'entrada')
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir entrada de stock: {e}")
    
    def movimiento_salida(self):
        """Abrir di√°logo de salida de stock"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import MovimientoInventarioDialog
            dialog = MovimientoInventarioDialog(self.window, 'almacen', 'salida')
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir salida de stock: {e}")
    
    def ver_historial_movimientos(self):
        """Ver historial de movimientos"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import HistorialMovimientosWindow
            dialog = HistorialMovimientosWindow(self.window, 'almacen')
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir historial: {e}")
    
    def movimiento(self, tipo):
        """Registrar movimiento de almac√©n (m√©todo legacy)"""
        if tipo == 'entrada':
            self.movimiento_entrada()
        else:
            self.movimiento_salida()
    
    def control_stock(self):
        """Control espec√≠fico de stock de almac√©n"""
        messagebox.showinfo("Control de Stock", "Funci√≥n espec√≠fica de control de stock de almac√©n")
    
    def lista_pedidos(self):
        """Lista de pedidos espec√≠fica"""
        messagebox.showinfo("Lista de Pedidos", "Gesti√≥n de pedidos de almac√©n")
    
    def actualizar_datos(self):
        """Actualizar datos espec√≠ficos"""
        self.cargar_inventario()
        messagebox.showinfo("Actualizado", "Datos de almac√©n actualizados")
    
    def reporte_almacen(self):
        """Generar reporte espec√≠fico de almac√©n"""
        try:
            archivo = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialvalue=f"reporte_almacen_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if archivo:
                cursor = self.conn.cursor()
                cursor.execute("""
                    SELECT codigo, nombre, saldo, unidad, valor_unitario, 
                           stock_minimo, ubicacion, proveedor,
                           (saldo * valor_unitario) as valor_total
                    FROM productos_almacen 
                    ORDER BY codigo
                """)
                
                productos = cursor.fetchall()
                
                with open(archivo, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['=== REPORTE DE ALMAC√âN ==='])
                    writer.writerow([f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
                    writer.writerow([])
                    writer.writerow(['C√≥digo', 'Producto', 'Saldo', 'Unidad', 'Valor Unit.',
                                   'Stock M√≠n.', 'Ubicaci√≥n', 'Proveedor', 'Valor Total'])
                    
                    for producto in productos:
                        writer.writerow(producto)
                
                messagebox.showinfo("√âxito", f"Reporte de almac√©n generado: {archivo}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error generando reporte: {e}")
    
    def mostrar_menu_contextual(self, event):
        """Men√∫ contextual espec√≠fico"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            
            menu = tk.Menu(self.window, tearoff=0)
            menu.add_command(label="‚úèÔ∏è Editar producto", command=self.editar_producto)
            menu.add_command(label="üì• Entrada", command=lambda: self.movimiento('entrada'))
            menu.add_command(label="üì§ Salida", command=lambda: self.movimiento('salida'))
            menu.add_separator()
            menu.add_command(label="üìä Ver detalles", command=self.ver_detalles)
            
            menu.post(event.x_root, event.y_root)
    
    def ver_detalles(self):
        """Ver detalles espec√≠ficos del producto"""
        producto = self.get_selected_producto()
        if producto:
            messagebox.showinfo("Detalles del Producto", 
                               f"C√≥digo: {producto[1]}\n" +
                               f"Nombre: {producto[2]}\n" +
                               f"Saldo: {producto[3]} {producto[4]}\n" +
                               f"Valor: ${producto[5]:,}\n" +
                               f"Ubicaci√≥n: {producto[7]}")
    
    def sort_column(self, col):
        """Ordenar columna espec√≠fica"""
        try:
            data = [(self.tree.set(child, col), child) for child in self.tree.get_children('')]
            data.sort()
            
            for index, (val, child) in enumerate(data):
                self.tree.move(child, '', index)
        except:
            pass
    
    def add_hover_effect(self, button, color):
        """Efecto hover espec√≠fico"""
        def on_enter(e):
            button.config(bg=self.darken_color(color))
        def on_leave(e):
            button.config(bg=color)
        
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
    
    def darken_color(self, color):
        """Oscurecer color para hover"""
        colors = {
            "#27ae60": "#229954",
            "#3498db": "#2980b9",
            "#f39c12": "#e67e22",
            "#8e44ad": "#7d3c98",
            "#e74c3c": "#c0392b",
            "#16a085": "#138d75",
            "#95a5a6": "#7f8c8d",
            "#9b59b6": "#8e44ad"
        }
        return colors.get(color, color)
    
    def cerrar_ventana(self):
        """Cerrar ventana espec√≠fica"""
        if self.conn:
            self.conn.close()
        self.window.destroy()


# =================== VENTANAS AUXILIARES ESPEC√çFICAS ===================

class ProductoAlmacenWindow:
    """Ventana espec√≠fica para productos de almac√©n"""
    
    def __init__(self, parent, main_window, modo="nuevo", producto=None):
        self.parent = parent
        self.main_window = main_window
        self.modo = modo
        self.producto = producto
        
        self.window = tk.Toplevel(parent)
        titulo = "‚úèÔ∏è Editar Producto de Almac√©n" if modo == "editar" else "‚ûï Nuevo Producto de Almac√©n"
        self.window.title(titulo)
        self.window.geometry("500x400")
        self.window.configure(bg='#f8f9fa')
        self.window.resizable(False, False)
        
        self.center_window()
        self.window.transient(parent)
        self.window.grab_set()
        
        self.create_widgets()
        
        if modo == "editar" and producto:
            self.cargar_datos()
    
    def center_window(self):
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (500 // 2)
        y = (self.window.winfo_screenheight() // 2) - (400 // 2)
        self.window.geometry(f"500x400+{x}+{y}")
    
    def create_widgets(self):
        # Header espec√≠fico
        header = tk.Frame(self.window, bg='#3498db', height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        titulo = "‚úèÔ∏è Editar Producto de Almac√©n" if self.modo == "editar" else "‚ûï Nuevo Producto de Almac√©n"
        title_label = tk.Label(header, text=titulo,
                             font=('Segoe UI', 16, 'bold'),
                             bg='#3498db', fg='white')
        title_label.pack(expand=True)
        
        # Formulario espec√≠fico
        content = tk.Frame(self.window, bg='#ffffff')
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Campos espec√≠ficos para almac√©n
        campos = [
            ("C√≥digo:", "entry_codigo"),
            ("Nombre del Producto:", "entry_nombre"),
            ("Saldo Actual:", "entry_saldo"),
            ("Unidad:", "entry_unidad"),
            ("Valor Unitario:", "entry_valor"),
            ("Stock M√≠nimo:", "entry_stock_min"),
            ("Ubicaci√≥n:", "combo_ubicacion"),
            ("Proveedor:", "entry_proveedor")
        ]
        
        for i, (label_text, widget_name) in enumerate(campos):
            label = tk.Label(content, text=label_text,
                           font=('Segoe UI', 10, 'bold'),
                           bg='#ffffff', fg='#2c3e50')
            label.grid(row=i, column=0, sticky='w', pady=8, padx=(0, 15))
            
            if widget_name == "combo_ubicacion":
                widget = ttk.Combobox(content, values=["A-01", "A-02", "A-03", "A-04", "A-05", "A-06"],
                                    width=35, font=('Segoe UI', 10))
            else:
                widget = tk.Entry(content, width=40, font=('Segoe UI', 10),
                                relief='solid', bd=1)
            
            widget.grid(row=i, column=1, pady=8, sticky='ew')
            setattr(self, widget_name, widget)
        
        content.grid_columnconfigure(1, weight=1)
        
        # Botones espec√≠ficos
        self.create_buttons(content)
    
    def create_buttons(self, parent):
        btn_frame = tk.Frame(parent, bg='#ffffff')
        btn_frame.grid(row=8, column=0, columnspan=2, pady=20)
        
        texto = "üíæ Actualizar" if self.modo == "editar" else "üíæ Guardar"
        save_btn = tk.Button(btn_frame, text=texto, command=self.guardar,
                           bg="#27ae60", fg="white", font=('Segoe UI', 11, 'bold'),
                           relief='flat', bd=0, padx=20, pady=10, cursor='hand2')
        save_btn.pack(side=tk.LEFT, padx=5)
        
        cancel_btn = tk.Button(btn_frame, text="‚ùå Cancelar", command=self.window.destroy,
                             bg="#e74c3c", fg="white", font=('Segoe UI', 11, 'bold'),
                             relief='flat', bd=0, padx=20, pady=10, cursor='hand2')
        cancel_btn.pack(side=tk.LEFT, padx=5)
    
    def cargar_datos(self):
        if not self.producto:
            return
        
        self.entry_codigo.insert(0, self.producto[1] or "")
        self.entry_nombre.insert(0, self.producto[2] or "")
        self.entry_saldo.insert(0, str(self.producto[3] or ""))
        self.entry_unidad.insert(0, self.producto[4] or "")
        self.entry_valor.insert(0, str(self.producto[5] or ""))
        self.entry_stock_min.insert(0, str(self.producto[6] or ""))
        self.combo_ubicacion.set(self.producto[7] or "")
        self.entry_proveedor.insert(0, self.producto[8] or "")
        
        if self.modo == "editar":
            self.entry_codigo.config(state='disabled')
    
    def guardar(self):
        try:
            # Validaciones espec√≠ficas
            if not all([self.entry_codigo.get().strip(), self.entry_nombre.get().strip(),
                       self.entry_unidad.get().strip()]):
                messagebox.showerror("Error", "Completa los campos obligatorios")
                return
            
            # Validar n√∫meros
            try:
                saldo = int(self.entry_saldo.get() or 0)
                valor = float(self.entry_valor.get() or 0)
                stock_min = int(self.entry_stock_min.get() or 0)
            except ValueError:
                messagebox.showerror("Error", "Los valores num√©ricos no son v√°lidos")
                return
            
            cursor = self.main_window.conn.cursor()
            
            if self.modo == "nuevo":
                # Verificar c√≥digo √∫nico
                cursor.execute("SELECT id FROM productos_almacen WHERE codigo = ?", 
                             (self.entry_codigo.get().strip(),))
                if cursor.fetchone():
                    messagebox.showerror("Error", "El c√≥digo ya existe")
                    return
                
                # Insertar nuevo
                cursor.execute('''
                    INSERT INTO productos_almacen 
                    (codigo, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    self.entry_codigo.get().strip(),
                    self.entry_nombre.get().strip(),
                    saldo,
                    self.entry_unidad.get().strip(),
                    valor,
                    stock_min,
                    self.combo_ubicacion.get(),
                    self.entry_proveedor.get().strip()
                ))
                mensaje = "Producto de almac√©n creado exitosamente"
            else:
                # Actualizar existente
                cursor.execute('''
                    UPDATE productos_almacen SET 
                    nombre=?, saldo=?, unidad=?, valor_unitario=?, 
                    stock_minimo=?, ubicacion=?, proveedor=?
                    WHERE codigo=?
                ''', (
                    self.entry_nombre.get().strip(),
                    saldo,
                    self.entry_unidad.get().strip(),
                    valor,
                    stock_min,
                    self.combo_ubicacion.get(),
                    self.entry_proveedor.get().strip(),
                    self.entry_codigo.get().strip()
                ))
                mensaje = "Producto de almac√©n actualizado exitosamente"
            
            self.main_window.conn.commit()
            messagebox.showinfo("√âxito", mensaje)
            
            # Actualizar tabla principal
            self.main_window.cargar_inventario()
            
            if self.modo == "nuevo":
                self.limpiar_campos()
            else:
                self.window.destroy()
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar: {e}")
    
    def limpiar_campos(self):
        for widget_name in ['entry_codigo', 'entry_nombre', 'entry_saldo', 'entry_unidad',
                           'entry_valor', 'entry_stock_min', 'entry_proveedor']:
            widget = getattr(self, widget_name)
            widget.delete(0, tk.END)
        self.combo_ubicacion.set('')


class MovimientoAlmacenWindow:
    """Ventana espec√≠fica para movimientos de almac√©n"""
    
    def __init__(self, parent, main_window, tipo):
        self.parent = parent
        self.main_window = main_window
        self.tipo = tipo
        
        self.window = tk.Toplevel(parent)
        titulo = f"üì• Entrada de Almac√©n" if tipo == 'entrada' else f"üì§ Salida de Almac√©n"
        self.window.title(titulo)
        self.window.geometry("450x350")
        self.window.configure(bg='#f8f9fa')
        self.window.resizable(False, False)
        
        self.center_window()
        self.window.transient(parent)
        self.window.grab_set()
        
        self.create_widgets()
    
    def center_window(self):
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (450 // 2)
        y = (self.window.winfo_screenheight() // 2) - (350 // 2)
        self.window.geometry(f"450x350+{x}+{y}")
    
    def create_widgets(self):
        # Header espec√≠fico
        color = "#27ae60" if self.tipo == 'entrada' else "#f39c12"
        header = tk.Frame(self.window, bg=color, height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        icon = "üì•" if self.tipo == 'entrada' else "üì§"
        titulo = f"{icon} {'Entrada' if self.tipo == 'entrada' else 'Salida'} de Almac√©n"
        title_label = tk.Label(header, text=titulo,
                             font=('Segoe UI', 16, 'bold'),
                             bg=color, fg='white')
        title_label.pack(expand=True)
        
        # Formulario
        content = tk.Frame(self.window, bg='#ffffff')
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Producto
        tk.Label(content, text="Producto de Almac√©n:", 
                font=('Segoe UI', 10, 'bold'),
                bg='#ffffff', fg='#2c3e50').grid(row=0, column=0, sticky='w', pady=8, padx=(0, 10))
        
        self.combo_producto = ttk.Combobox(content, width=35, font=('Segoe UI', 10))
        self.combo_producto.grid(row=0, column=1, pady=8)
        self.cargar_productos()
        
        # Cantidad
        tk.Label(content, text="Cantidad:", 
                font=('Segoe UI', 10, 'bold'),
                bg='#ffffff', fg='#2c3e50').grid(row=1, column=0, sticky='w', pady=8, padx=(0, 10))
        
        self.entry_cantidad = tk.Entry(content, width=38, font=('Segoe UI', 10))
        self.entry_cantidad.grid(row=1, column=1, pady=8)
        
        # Campos adicionales seg√∫n tipo
        if self.tipo == 'entrada':
            tk.Label(content, text="Factura:", 
                    font=('Segoe UI', 10, 'bold'),
                    bg='#ffffff', fg='#2c3e50').grid(row=2, column=0, sticky='w', pady=8, padx=(0, 10))
            self.entry_factura = tk.Entry(content, width=38, font=('Segoe UI', 10))
            self.entry_factura.grid(row=2, column=1, pady=8)
            
            tk.Label(content, text="Proveedor:", 
                    font=('Segoe UI', 10, 'bold'),
                    bg='#ffffff', fg='#2c3e50').grid(row=3, column=0, sticky='w', pady=8, padx=(0, 10))
            self.entry_proveedor = tk.Entry(content, width=38, font=('Segoe UI', 10))
            self.entry_proveedor.grid(row=3, column=1, pady=8)
        else:
            tk.Label(content, text="Destino:", 
                    font=('Segoe UI', 10, 'bold'),
                    bg='#ffffff', fg='#2c3e50').grid(row=2, column=0, sticky='w', pady=8, padx=(0, 10))
            self.entry_destino = tk.Entry(content, width=38, font=('Segoe UI', 10))
            self.entry_destino.grid(row=2, column=1, pady=8)
        
        # Observaciones
        tk.Label(content, text="Observaciones:", 
                font=('Segoe UI', 10, 'bold'),
                bg='#ffffff', fg='#2c3e50').grid(row=4, column=0, sticky='w', pady=8, padx=(0, 10))
        self.entry_observaciones = tk.Entry(content, width=38, font=('Segoe UI', 10))
        self.entry_observaciones.grid(row=4, column=1, pady=8)
        
        # Botones
        btn_frame = tk.Frame(content, bg='#ffffff')
        btn_frame.grid(row=5, column=0, columnspan=2, pady=20)
        
        tk.Button(btn_frame, text=f"üíæ Registrar {self.tipo.title()}", 
                 command=self.registrar, bg=color, fg="white",
                 font=('Segoe UI', 11, 'bold'), relief='flat', bd=0, 
                 padx=20, pady=10, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="‚ùå Cancelar", command=self.window.destroy,
                 bg="#e74c3c", fg="white", font=('Segoe UI', 11, 'bold'),
                 relief='flat', bd=0, padx=20, pady=10, cursor='hand2').pack(side=tk.LEFT, padx=5)
    
    def cargar_productos(self):
        cursor = self.main_window.conn.cursor()
        cursor.execute("SELECT codigo, nombre FROM productos_almacen ORDER BY codigo")
        productos = [f"{row[0]} - {row[1]}" for row in cursor.fetchall()]
        self.combo_producto['values'] = productos
    
    def registrar(self):
        try:
            if not all([self.combo_producto.get(), self.entry_cantidad.get()]):
                messagebox.showerror("Error", "Completa los campos obligatorios")
                return
            
            codigo = self.combo_producto.get().split(' - ')[0]
            cantidad = int(self.entry_cantidad.get())
            
            if cantidad <= 0:
                raise ValueError("Cantidad debe ser mayor a 0")
            
            cursor = self.main_window.conn.cursor()
            
            # Obtener producto
            cursor.execute("SELECT id, saldo FROM productos_almacen WHERE codigo = ?", (codigo,))
            resultado = cursor.fetchone()
            if not resultado:
                messagebox.showerror("Error", "Producto no encontrado")
                return
            
            producto_id, saldo_actual = resultado
            
            # Validar stock para salidas
            if self.tipo == 'salida' and cantidad > saldo_actual:
                messagebox.showerror("Error", f"Stock insuficiente. Disponible: {saldo_actual}")
                return
            
            # Nuevo saldo
            nuevo_saldo = saldo_actual + cantidad if self.tipo == 'entrada' else saldo_actual - cantidad
            
            # Registrar movimiento
            factura = getattr(self, 'entry_factura', None)
            proveedor = getattr(self, 'entry_proveedor', None)
            destino = getattr(self, 'entry_destino', None)
            
            cursor.execute('''
                INSERT INTO movimientos_almacen 
                (producto_id, tipo, cantidad, fecha, factura, proveedor, destino, observaciones)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                producto_id, self.tipo, cantidad, date.today(),
                factura.get() if factura else None,
                proveedor.get() if proveedor else None,
                destino.get() if destino else None,
                self.entry_observaciones.get()
            ))
            
            # Actualizar saldo
            cursor.execute("UPDATE productos_almacen SET saldo = ? WHERE id = ?", 
                         (nuevo_saldo, producto_id))
            
            self.main_window.conn.commit()
            messagebox.showinfo("√âxito", f"{self.tipo.title()} de almac√©n registrada correctamente")
            
            # Actualizar tabla
            self.main_window.cargar_inventario()
            self.window.destroy()
            
        except ValueError as e:
            messagebox.showerror("Error", f"Cantidad inv√°lida: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Error registrando movimiento: {e}")