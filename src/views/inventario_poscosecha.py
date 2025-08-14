#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Inventario de Poscosecha - Módulo Independiente
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import os
from datetime import datetime, date
import sqlite3
import csv

class InventarioPoscosechaWindow:
    def __init__(self, parent, main_window=None):
        self.parent = parent
        self.main_window = main_window
        self.setup_database()
        
        # Crear ventana principal
        self.window = tk.Toplevel(parent)
        self.window.title("Sistema de Inventario - POSCOSECHA")
        self.window.geometry("1200x800")
        self.window.configure(bg='#f0f0f0')
        
        # Configurar ventana
        self.window.resizable(True, True)
        self.window.minsize(1000, 600)
        
        # Centrar ventana
        self.center_window()
        
        # Hacer modal
        self.window.transient(parent)
        self.window.grab_set()
        
        # Configurar estilos
        self.configurar_estilos()
        
        # Crear interfaz
        self.create_interface()
        
        # Cargar datos automáticamente
        total_cargados = self.cargar_inventario()
        
        # Mostrar información de carga si es exitoso
        if total_cargados and total_cargados > 5:
            self.window.after(2000, lambda: self.mostrar_info_carga(total_cargados))
    
    def configurar_estilos(self):
        """Configurar estilos para poscosecha"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Estilo específico para poscosecha
        style.configure('Poscosecha.Treeview',
                       background='#ffffff',
                       foreground='#2c3e50',
                       fieldbackground='#ffffff',
                       font=('Arial', 9),
                       rowheight=25)
        
        # Headers con color específico para poscosecha - verde oscuro
        style.configure('Poscosecha.Treeview.Heading',
                       background='#16a085',  # Verde oscuro para poscosecha
                       foreground='white',
                       font=('Arial', 9, 'bold'),
                       relief='flat')
        
        style.map('Poscosecha.Treeview',
                 background=[('selected', '#138d75')],
                 foreground=[('selected', '#ffffff')])
    
    def center_window(self):
        """Centrar ventana"""
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (1200 // 2)
        y = (self.window.winfo_screenheight() // 2) - (800 // 2)
        self.window.geometry(f"1200x800+{x}+{y}")
    
    def setup_database(self):
        """Configurar base de datos específica para poscosecha"""
        try:
            db_dir = 'database'
            os.makedirs(db_dir, exist_ok=True)
            
            db_path = os.path.join(db_dir, 'inventario_poscosecha.db')
            self.conn = sqlite3.connect(db_path)
            self.crear_tablas()
            self.poblar_datos_poscosecha()
        except Exception as e:
            print(f"Error BD poscosecha: {e}")
            self.conn = sqlite3.connect('inventario_poscosecha.db')
            self.crear_tablas()
            self.poblar_datos_poscosecha()
    
    def crear_tablas(self):
        """Crear tablas específicas para poscosecha"""
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos_poscosecha (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE NOT NULL,
                categoria TEXT NOT NULL,
                nombre TEXT NOT NULL,
                saldo INTEGER DEFAULT 0,
                unidad TEXT NOT NULL,
                valor_unitario REAL NOT NULL,
                stock_minimo INTEGER DEFAULT 0,
                ubicacion TEXT,
                proveedor TEXT,
                tipo_producto TEXT,
                fecha_vencimiento DATE,
                lote TEXT,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS movimientos_poscosecha (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                producto_id INTEGER,
                tipo TEXT NOT NULL,
                cantidad INTEGER NOT NULL,
                fecha DATE NOT NULL,
                factura TEXT,
                proveedor TEXT,
                destino TEXT,
                valor_total REAL,
                observaciones TEXT,
                responsable TEXT,
                fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (producto_id) REFERENCES productos_poscosecha (id)
            )
        ''')
        
        self.conn.commit()
    
    def poblar_datos_poscosecha(self):
        """Cargar datos específicos de poscosecha"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM productos_poscosecha")
        
        if cursor.fetchone()[0] == 0:
            print("🔄 Cargando datos de POSCOSECHA desde Excel...")
            
            # Intentar cargar desde Excel
            if self.cargar_desde_excel_poscosecha():
                print("✅ Datos de POSCOSECHA cargados desde Excel")
                return
            
            # Datos de ejemplo específicos para poscosecha
            productos_poscosecha = [
                ('PC001', 'EMBALAJE', 'CAJAS CARTON BANANO 18.14 KG', 500, 'UND', 2500, 50, 'PC-01', 'CARTONERIA DEL VALLE', 'EMPAQUE', None, 'LT2024001'),
                ('PC002', 'EMBALAJE', 'BOLSAS POLIETILENO CALIBRE 2', 2000, 'UND', 150, 200, 'PC-02', 'PLASTICOS SA', 'EMPAQUE', None, 'LT2024002'),
                ('PC003', 'QUIMICO', 'TIABENDAZOL FUNGICIDA', 25, 'KG', 45000, 5, 'PC-03', 'AGROQUIMICOS ANDINOS', 'TRATAMIENTO', '2025-12-31', 'LT2024003'),
                ('PC004', 'HERRAMIENTA', 'CUCHILLOS POSCOSECHA INOX', 15, 'UND', 35000, 5, 'PC-04', 'HERRAMIENTAS INDUSTRIALES', 'HERRAMIENTA', None, 'LT2024004'),
                ('PC005', 'QUIMICO', 'CERA PARA FRUTAS', 200, 'LT', 18500, 20, 'PC-05', 'PRODUCTOS POSCOSECHA', 'TRATAMIENTO', '2025-10-15', 'LT2024005'),
                ('PC006', 'ETIQUETA', 'ETIQUETAS PLU BANANO', 10000, 'UND', 25, 1000, 'PC-06', 'ETIQUETAS DEL PACIFICO', 'IDENTIFICACION', None, 'LT2024006'),
                ('PC007', 'EMBALAJE', 'PALLETS PLASTICO 1200x800', 80, 'UND', 125000, 10, 'PC-07', 'PALLETS COLOMBIA', 'EMPAQUE', None, 'LT2024007'),
                ('PC008', 'QUIMICO', 'DESINFECTANTE CLORO', 0, 'LT', 8500, 10, 'PC-08', 'QUIMICOS INDUSTRIALES', 'LIMPIEZA', '2025-08-20', 'LT2024008')
            ]
            
            cursor.executemany('''
                INSERT INTO productos_poscosecha (codigo, categoria, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor, tipo_producto, fecha_vencimiento, lote)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', productos_poscosecha)
            
            self.conn.commit()
            print("📋 Productos de POSCOSECHA inicializados")
    
    def cargar_desde_excel_poscosecha(self):
        """Cargar específicamente desde archivo de poscosecha"""
        archivo_poscosecha = 'SALDOS POSCOSECHA  JUNIO  2025.xlsx'
        
        if not os.path.exists(archivo_poscosecha):
            print(f"❌ {archivo_poscosecha} no encontrado")
            return False
        
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(archivo_poscosecha, data_only=True)
            sheet = workbook['Poscosecha'] if 'Poscosecha' in workbook.sheetnames else workbook.active
            
            productos = self.procesar_hoja_poscosecha(sheet)
            
            if productos:
                cursor = self.conn.cursor()
                cursor.executemany('''
                    INSERT OR REPLACE INTO productos_poscosecha 
                    (codigo, categoria, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor, tipo_producto, fecha_vencimiento, lote)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', productos)
                
                self.conn.commit()
                print(f"✅ {len(productos)} productos de POSCOSECHA cargados desde Excel")
                return True
            
        except Exception as e:
            print(f"Error cargando Excel de poscosecha: {e}")
            return False
        
        return False
    
    def procesar_hoja_poscosecha(self, sheet):
        """Procesar hoja específica de poscosecha"""
        productos = []
        contador = 1
        
        print("   📋 Procesando productos de POSCOSECHA...")
        
        for row_num in range(1, min(sheet.max_row + 1, 200)):
            try:
                # Buscar productos válidos en diferentes columnas
                for col in range(1, 6):
                    cell_value = sheet.cell(row=row_num, column=col).value
                    
                    if (cell_value and isinstance(cell_value, str) and 
                        len(str(cell_value).strip()) > 5):
                        
                        nombre = str(cell_value).strip().upper()
                        
                        # Filtrar headers y texto no deseado
                        if (not any(skip in nombre for skip in ['FECHA', 'FACTURA', 'CANT', 'VALOR', 'TOTAL', 'SALDO', 'ANTERIOR']) and
                            len(nombre) > 3):
                            
                            # Determinar categoría basada en el nombre
                            categoria = self.determinar_categoria_poscosecha(nombre)
                            tipo_producto = self.determinar_tipo_producto(nombre)
                            
                            # Buscar saldo en columnas adyacentes
                            saldo = 0
                            unidad = 'UND'
                            
                            for saldo_col in range(col + 1, min(col + 4, 8)):
                                saldo_val = sheet.cell(row=row_num, column=saldo_col).value
                                if isinstance(saldo_val, (int, float)) and saldo_val >= 0:
                                    saldo = int(saldo_val)
                                    
                                    # Buscar unidad
                                    unidad_val = sheet.cell(row=row_num, column=saldo_col + 1).value
                                    if isinstance(unidad_val, str) and len(str(unidad_val).strip()) <= 10:
                                        unidad = str(unidad_val).strip().upper()
                                    break
                            
                            # Crear producto
                            codigo = f'PC{contador:03d}'
                            valor_unitario = self.estimar_valor_poscosecha(nombre, categoria)
                            stock_minimo = max(5, int(saldo * 0.1)) if saldo > 0 else 10
                            
                            productos.append((
                                codigo, categoria, nombre, saldo, unidad,
                                valor_unitario, stock_minimo, f'PC-{contador:02d}', 
                                'PROVEEDOR POSCOSECHA', tipo_producto, None, f'LT2024{contador:03d}'
                            ))
                            
                            contador += 1
                            
                            if contador <= 10:
                                print(f"     • {nombre} - {saldo} {unidad}")
                            
                            break
                    
            except Exception as e:
                continue
        
        print(f"   ✅ {len(productos)} productos de POSCOSECHA procesados")
        return productos
    
    def determinar_categoria_poscosecha(self, nombre):
        """Determinar categoría basada en el nombre del producto"""
        nombre_lower = nombre.lower()
        
        if any(palabra in nombre_lower for palabra in ['caja', 'bolsa', 'empaque', 'carton', 'pallet']):
            return 'EMBALAJE'
        elif any(palabra in nombre_lower for palabra in ['tiabendazol', 'fungicida', 'cera', 'cloro', 'desinfect']):
            return 'QUIMICO'
        elif any(palabra in nombre_lower for palabra in ['etiqueta', 'label', 'sticker']):
            return 'ETIQUETA'
        elif any(palabra in nombre_lower for palabra in ['cuchillo', 'herramienta', 'equipo']):
            return 'HERRAMIENTA'
        else:
            return 'GENERAL'
    
    def determinar_tipo_producto(self, nombre):
        """Determinar tipo de producto"""
        nombre_lower = nombre.lower()
        
        if any(palabra in nombre_lower for palabra in ['caja', 'bolsa', 'empaque', 'pallet']):
            return 'EMPAQUE'
        elif any(palabra in nombre_lower for palabra in ['fungicida', 'cera', 'tiabendazol']):
            return 'TRATAMIENTO'
        elif any(palabra in nombre_lower for palabra in ['etiqueta']):
            return 'IDENTIFICACION'
        elif any(palabra in nombre_lower for palabra in ['cloro', 'desinfect']):
            return 'LIMPIEZA'
        else:
            return 'GENERAL'
    
    def estimar_valor_poscosecha(self, nombre, categoria):
        """Estimar valor para productos de poscosecha"""
        nombre_lower = nombre.lower()
        
        if categoria == 'EMBALAJE':
            if 'caja' in nombre_lower:
                return 2500
            elif 'bolsa' in nombre_lower:
                return 150
            elif 'pallet' in nombre_lower:
                return 125000
            else:
                return 5000
        elif categoria == 'QUIMICO':
            if 'tiabendazol' in nombre_lower or 'fungicida' in nombre_lower:
                return 45000
            elif 'cera' in nombre_lower:
                return 18500
            elif 'cloro' in nombre_lower:
                return 8500
            else:
                return 25000
        elif categoria == 'ETIQUETA':
            return 25
        elif categoria == 'HERRAMIENTA':
            return 35000
        else:
            return 10000
    
    def create_interface(self):
        """Crear interfaz específica para poscosecha"""
        # Header específico para poscosecha
        self.create_poscosecha_header()
        
        # Contenedor principal
        main_container = tk.Frame(self.window, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Layout: sidebar + contenido
        self.create_poscosecha_sidebar(main_container)
        self.create_poscosecha_content(main_container)
    
    def create_poscosecha_header(self):
        """Header específico para poscosecha"""
        header = tk.Frame(self.window, bg='#16a085', height=70)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        header_content = tk.Frame(header, bg='#16a085')
        header_content.pack(fill=tk.BOTH, expand=True, padx=30, pady=15)
        
        # Logo y título específico
        title_frame = tk.Frame(header_content, bg='#16a085')
        title_frame.pack(side=tk.LEFT)
        
        icon_label = tk.Label(title_frame, text="🥬", font=('Arial', 24), 
                             bg='#16a085', fg='#ffffff')
        icon_label.pack(side=tk.LEFT, padx=(0, 15))
        
        title_container = tk.Frame(title_frame, bg='#16a085')
        title_container.pack(side=tk.LEFT)
        
        title_label = tk.Label(title_container, text="Inventario de POSCOSECHA", 
                              font=('Segoe UI', 18, 'bold'), bg='#16a085', fg='#ffffff')
        title_label.pack(anchor='w')
        
        subtitle = tk.Label(title_container, text="Gestión de productos para procesamiento post-cosecha", 
                           font=('Segoe UI', 10), bg='#16a085', fg='#ecf0f1')
        subtitle.pack(anchor='w')
        
        # Botones específicos para poscosecha
        self.create_poscosecha_actions(header_content)
    
    def create_poscosecha_actions(self, parent):
        """Botones específicos para poscosecha"""
        actions_frame = tk.Frame(parent, bg='#16a085')
        actions_frame.pack(side=tk.RIGHT)
        
        buttons = [
            ("+ Nuevo", self.nuevo_producto_poscosecha, "#27ae60", "Nuevo producto de poscosecha"),
            ("↓ Entrada", lambda: self.movimiento('entrada'), "#3498db", "Entrada de poscosecha"),
            ("↑ Salida", lambda: self.movimiento('salida'), "#f39c12", "Salida de poscosecha"),
            ("📊", self.reporte_poscosecha, "#8e44ad", "Reporte de poscosecha"),
            ("✕", self.cerrar_ventana, "#e74c3c", "Cerrar")
        ]
        
        for text, command, color, tooltip in buttons:
            btn = tk.Button(actions_frame, text=text, command=command,
                           bg=color, fg='white', font=('Segoe UI', 9, 'bold'),
                           relief='flat', bd=0, padx=12, pady=6, cursor='hand2')
            btn.pack(side=tk.LEFT, padx=3)
            self.add_hover_effect(btn, color)
    
    def create_poscosecha_sidebar(self, parent):
        """Sidebar específico para poscosecha"""
        sidebar = tk.Frame(parent, bg='#ffffff', width=300)
        sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=(20, 0), pady=20)
        sidebar.pack_propagate(False)
        
        # Título
        sidebar_title = tk.Label(sidebar, text="🥬 Control Poscosecha", 
                                font=('Segoe UI', 16, 'bold'), 
                                bg='#ffffff', fg='#2c3e50')
        sidebar_title.pack(anchor='w', padx=20, pady=(20, 15))
        
        # Estadísticas específicas
        self.create_poscosecha_stats(sidebar)
        
        # Filtros específicos
        self.create_poscosecha_filters(sidebar)
        
        # Acciones rápidas específicas
        self.create_poscosecha_quick_actions(sidebar)
    
    def create_poscosecha_stats(self, parent):
        """Estadísticas específicas de poscosecha"""
        stats_container = tk.Frame(parent, bg='#ffffff')
        stats_container.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Obtener estadísticas de poscosecha
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM productos_poscosecha")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM productos_poscosecha WHERE saldo <= stock_minimo AND saldo > 0")
        bajo = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM productos_poscosecha WHERE saldo = 0")
        agotado = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(DISTINCT categoria) FROM productos_poscosecha")
        categorias = cursor.fetchone()[0]
        
        cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos_poscosecha")
        valor = cursor.fetchone()[0] or 0
        
        # Cards específicas para poscosecha
        cards_data = [
            ("Productos Total", str(total), "#16a085", "#e8f6f3"),
            ("Stock Bajo", str(bajo), "#f39c12", "#fef9e7"),
            ("Sin Stock", str(agotado), "#e74c3c", "#fdedec"),
            ("Categorías", str(categorias), "#9b59b6", "#f4ecf7"),
            ("Valor Total", f"${valor:,.0f}", "#27ae60", "#eafaf1")
        ]
        
        for title, value, color, bg_color in cards_data:
            card = tk.Frame(stats_container, bg=bg_color, relief='solid', bd=1)
            card.pack(fill=tk.X, pady=6)
            
            card_content = tk.Frame(card, bg=bg_color)
            card_content.pack(fill=tk.BOTH, padx=12, pady=10)
            
            value_label = tk.Label(card_content, text=value, 
                                  font=('Segoe UI', 16, 'bold'),
                                  bg=bg_color, fg=color)
            value_label.pack(anchor='w')
            
            title_label = tk.Label(card_content, text=title,
                                  font=('Segoe UI', 9),
                                  bg=bg_color, fg='#7f8c8d')
            title_label.pack(anchor='w')
    
    def create_poscosecha_filters(self, parent):
        """Filtros específicos para poscosecha"""
        filters_frame = tk.Frame(parent, bg='#ffffff')
        filters_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        tk.Label(filters_frame, text="🔍 Filtros Poscosecha", 
                font=('Segoe UI', 14, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(anchor='w', pady=(0, 10))
        
        # Búsqueda específica
        search_frame = tk.Frame(filters_frame, bg='#ffffff')
        search_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(search_frame, text="Buscar producto:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var,
                               font=('Segoe UI', 10), relief='solid', bd=1)
        search_entry.pack(fill=tk.X, pady=(3, 0))
        
        # Filtro por categoría
        categoria_frame = tk.Frame(filters_frame, bg='#ffffff')
        categoria_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(categoria_frame, text="Categoría:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.filter_categoria = tk.StringVar(value="Todas")
        self.filter_categoria.trace('w', self.on_search_change)
        categoria_combo = ttk.Combobox(categoria_frame, textvariable=self.filter_categoria,
                                      values=["Todas", "EMBALAJE", "QUIMICO", "ETIQUETA", "HERRAMIENTA", "GENERAL"],
                                      state="readonly", font=('Segoe UI', 10))
        categoria_combo.pack(fill=tk.X, pady=(3, 0))
        
        # Filtro por tipo de producto
        tipo_frame = tk.Frame(filters_frame, bg='#ffffff')
        tipo_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(tipo_frame, text="Tipo de producto:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.filter_tipo = tk.StringVar(value="Todos")
        self.filter_tipo.trace('w', self.on_search_change)
        tipo_combo = ttk.Combobox(tipo_frame, textvariable=self.filter_tipo,
                                 values=["Todos", "EMPAQUE", "TRATAMIENTO", "IDENTIFICACION", "LIMPIEZA", "GENERAL"],
                                 state="readonly", font=('Segoe UI', 10))
        tipo_combo.pack(fill=tk.X, pady=(3, 0))
        
        # Filtro por estado de stock
        stock_frame = tk.Frame(filters_frame, bg='#ffffff')
        stock_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(stock_frame, text="Estado del stock:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.filter_stock = tk.StringVar(value="Todos")
        self.filter_stock.trace('w', self.on_search_change)
        stock_combo = ttk.Combobox(stock_frame, textvariable=self.filter_stock,
                                  values=["Todos", "Normal", "Bajo", "Agotado"],
                                  state="readonly", font=('Segoe UI', 10))
        stock_combo.pack(fill=tk.X, pady=(3, 0))
        
        # Botón limpiar filtros
        clear_btn = tk.Button(filters_frame, text="🗑️ Limpiar", 
                             command=self.limpiar_filtros,
                             bg='#95a5a6', fg='white', font=('Segoe UI', 9),
                             relief='flat', bd=0, pady=8, cursor='hand2')
        clear_btn.pack(fill=tk.X, pady=(10, 0))
    
    def create_poscosecha_quick_actions(self, parent):
        """Acciones rápidas específicas para poscosecha"""
        actions_frame = tk.Frame(parent, bg='#ffffff')
        actions_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        tk.Label(actions_frame, text="⚡ Acciones Poscosecha", 
                font=('Segoe UI', 14, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(anchor='w', pady=(0, 15))
        
        actions = [
            ("✏️ Editar Producto", self.editar_producto, "#16a085"),        # Verde azulado
            ("📥 Entrada de Stock", lambda: self.movimiento_entrada(), "#27ae60"),  # Verde esmeralda
            ("📤 Salida de Stock", lambda: self.movimiento_salida(), "#e67e22"),    # Naranja suave
            ("📋 Historial Movimientos", lambda: self.ver_historial_movimientos(), "#8e44ad"),  # Púrpura
            ("📦 Control Empaque", self.control_empaque, "#3498db"),         # Azul cielo
            ("🧪 Tratamientos", self.control_tratamientos, "#7f8c8d"),       # Gris azulado
            ("🔄 Actualizar Stock", self.actualizar_stock, "#2c3e50")       # Azul oscuro
        ]
        
        for text, command, color in actions:
            btn = tk.Button(actions_frame, text=text, command=command,
                           bg=color, fg='white', font=('Segoe UI', 10),
                           relief='flat', bd=0, pady=8, cursor='hand2')
            btn.pack(fill=tk.X, pady=3)
            self.add_hover_effect(btn, color)
    
    def create_poscosecha_content(self, parent):
        """Contenido principal específico para poscosecha"""
        content_area = tk.Frame(parent, bg='#f8f9fa')
        content_area.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Card para la tabla de poscosecha
        table_card = tk.Frame(content_area, bg='#ffffff', relief='solid', bd=1)
        table_card.pack(fill=tk.BOTH, expand=True)
        
        # Header específico
        table_header = tk.Frame(table_card, bg='#ffffff', height=50)
        table_header.pack(fill=tk.X)
        table_header.pack_propagate(False)
        
        tk.Label(table_header, text="🥬 Productos de Poscosecha", 
                font=('Segoe UI', 16, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(side=tk.LEFT, padx=20, pady=15)
        
        self.products_count_label = tk.Label(table_header, text="", 
                                           font=('Segoe UI', 10),
                                           bg='#ffffff', fg='#7f8c8d')
        self.products_count_label.pack(side=tk.RIGHT, padx=20, pady=15)
        
        # Tabla específica para poscosecha
        self.create_poscosecha_table(table_card)
    
    def create_poscosecha_table(self, parent):
        """Tabla específica para poscosecha"""
        table_container = tk.Frame(parent, bg='#ffffff')
        table_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # TreeView específico para poscosecha
        columns = ('Código', 'Categoría', 'Producto', 'Saldo', 'Unidad', 'Valor Unit.', 'Stock Min.', 'Tipo', 'Estado')
        self.tree = ttk.Treeview(table_container, columns=columns, show='headings',
                                height=20, style='Poscosecha.Treeview')
        
        # Configurar columnas específicas
        column_config = {
            'Código': {'width': 80, 'anchor': 'center'},
            'Categoría': {'width': 100, 'anchor': 'center'},
            'Producto': {'width': 220, 'anchor': 'w'},
            'Saldo': {'width': 80, 'anchor': 'center'},
            'Unidad': {'width': 60, 'anchor': 'center'},
            'Valor Unit.': {'width': 100, 'anchor': 'e'},
            'Stock Min.': {'width': 80, 'anchor': 'center'},
            'Tipo': {'width': 120, 'anchor': 'center'},
            'Estado': {'width': 100, 'anchor': 'center'}
        }
        
        for col, config in column_config.items():
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c))
            self.tree.column(col, width=config['width'], anchor=config['anchor'])
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(table_container, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Layout
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        table_container.grid_rowconfigure(0, weight=1)
        table_container.grid_columnconfigure(0, weight=1)
        
        # Eventos específicos
        self.tree.bind('<Double-1>', lambda e: self.editar_producto())
        self.tree.bind('<Button-3>', self.mostrar_menu_contextual)
    
    # =================== FUNCIONES ESPECÍFICAS DE POSCOSECHA ===================
    
    def cargar_inventario(self):
        """Cargar inventario específico de poscosecha"""
        try:
            # Limpiar datos existentes
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Query específica para poscosecha
            query = "SELECT codigo, categoria, nombre, saldo, unidad, valor_unitario, stock_minimo, tipo_producto FROM productos_poscosecha WHERE 1=1"
            params = []
            
            # Aplicar filtros específicos
            if hasattr(self, 'search_var'):
                texto = self.search_var.get().strip().lower()
                if texto:
                    query += " AND (LOWER(nombre) LIKE ? OR LOWER(codigo) LIKE ? OR LOWER(categoria) LIKE ?)"
                    params.extend([f'%{texto}%', f'%{texto}%', f'%{texto}%'])
            
            if hasattr(self, 'filter_categoria'):
                categoria = self.filter_categoria.get()
                if categoria and categoria != "Todas":
                    query += " AND categoria = ?"
                    params.append(categoria)
            
            if hasattr(self, 'filter_tipo'):
                tipo = self.filter_tipo.get()
                if tipo and tipo != "Todos":
                    query += " AND tipo_producto = ?"
                    params.append(tipo)
            
            if hasattr(self, 'filter_stock'):
                stock_filter = self.filter_stock.get()
                if stock_filter == "Bajo":
                    query += " AND saldo <= stock_minimo AND saldo > 0"
                elif stock_filter == "Agotado":
                    query += " AND saldo = 0"
                elif stock_filter == "Normal":
                    query += " AND saldo > stock_minimo"
            
            query += " ORDER BY categoria, codigo"
            
            # Ejecutar query
            cursor = self.conn.cursor()
            cursor.execute(query, params)
            productos_mostrados = 0
            
            for row in cursor.fetchall():
                codigo, categoria, nombre, saldo, unidad, valor_unit, stock_min, tipo_producto = row
                
                # Estados específicos para poscosecha
                if saldo == 0:
                    estado = "🔴 AGOTADO"
                    tags = ('agotado',)
                elif saldo <= stock_min:
                    estado = "🟡 STOCK BAJO"
                    tags = ('bajo',)
                else:
                    estado = "🟢 NORMAL"
                    tags = ('normal',)
                
                valor_formateado = f"${valor_unit:,.0f}"
                
                self.tree.insert('', tk.END, values=(
                    codigo, categoria, nombre, saldo, unidad, valor_formateado,
                    stock_min, tipo_producto or "GENERAL", estado
                ), tags=tags)
                productos_mostrados += 1
            
            # Configurar colores específicos
            self.tree.tag_configure('agotado', background='#ffebee', foreground='#d32f2f')
            self.tree.tag_configure('bajo', background='#fff3e0', foreground='#f57c00')
            self.tree.tag_configure('normal', background='#ffffff', foreground='#2c3e50')
            
            # Actualizar contador
            if hasattr(self, 'products_count_label'):
                cursor.execute("SELECT COUNT(*) FROM productos_poscosecha")
                total = cursor.fetchone()[0]
                self.products_count_label.config(
                    text=f"Mostrando {productos_mostrados} de {total} productos"
                )
            
            return total
            
        except Exception as e:
            print(f"Error cargando inventario de poscosecha: {e}")
            return 0
    
    def on_search_change(self, *args):
        """Actualizar cuando cambien los filtros"""
        self.cargar_inventario()
    
    def limpiar_filtros(self):
        """Limpiar filtros específicos"""
        if hasattr(self, 'search_var'):
            self.search_var.set("")
        if hasattr(self, 'filter_categoria'):
            self.filter_categoria.set("Todas")
        if hasattr(self, 'filter_tipo'):
            self.filter_tipo.set("Todos")
        if hasattr(self, 'filter_stock'):
            self.filter_stock.set("Todos")
    
    def mostrar_info_carga(self, total):
        """Mostrar información de carga específica"""
        messagebox.showinfo(
            "🥬 Poscosecha Cargada", 
            f"¡Sistema de Poscosecha listo!\n\n" +
            f"📦 Productos poscosecha: {total}\n" +
            f"🏷️ Múltiples categorías\n" +
            f"📊 Control de empaque y tratamientos\n\n" +
            "Sistema especializado para gestión\n" +
            "de productos post-cosecha."
        )
    
    # =================== FUNCIONES DE ACCIONES ===================
    
    def nuevo_producto_poscosecha(self):
        """Crear nuevo producto específico de poscosecha"""
        ProductoPoscosechaWindow(self.window, self, modo="nuevo")
    
    def editar_producto(self):
        """Editar producto de poscosecha"""
        producto = self.get_selected_producto()
        if producto:
            ProductoPoscosechaWindow(self.window, self, modo="editar", producto=producto)
    
    def get_selected_producto(self):
        """Obtener producto seleccionado"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Selecciona un producto de poscosecha")
            return None
        
        item = self.tree.item(selection[0])
        codigo = item['values'][0]
        
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM productos_poscosecha WHERE codigo = ?", (codigo,))
        return cursor.fetchone()
    
    def movimiento_entrada(self):
        """Abrir diálogo de entrada de stock"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import MovimientoInventarioDialog
            dialog = MovimientoInventarioDialog(self.window, 'poscosecha', 'entrada')
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir entrada de stock: {e}")
    
    def movimiento_salida(self):
        """Abrir diálogo de salida de stock"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import MovimientoInventarioDialog
            dialog = MovimientoInventarioDialog(self.window, 'poscosecha', 'salida')
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir salida de stock: {e}")
    
    def ver_historial_movimientos(self):
        """Ver historial de movimientos"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import HistorialMovimientosWindow
            dialog = HistorialMovimientosWindow(self.window, 'poscosecha')
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir historial: {e}")
    
    def movimiento(self, tipo):
        """Registrar movimiento de poscosecha (método legacy)"""
        if tipo == 'entrada':
            self.movimiento_entrada()
        else:
            self.movimiento_salida()
    
    def control_empaque(self):
        """Control específico de empaque"""
        messagebox.showinfo("Control de Empaque", "Función específica de control de empaque y embalaje")
    
    def control_tratamientos(self):
        """Control específico de tratamientos"""
        messagebox.showinfo("Control de Tratamientos", "Gestión de tratamientos químicos post-cosecha")
    
    def actualizar_stock(self):
        """Actualizar stock específico"""
        self.cargar_inventario()
        messagebox.showinfo("Actualizado", "Stock de poscosecha actualizado")
    
    def reporte_poscosecha(self):
        """Generar reporte específico de poscosecha"""
        try:
            archivo = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialvalue=f"reporte_poscosecha_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if archivo:
                cursor = self.conn.cursor()
                cursor.execute("""
                    SELECT codigo, categoria, nombre, saldo, unidad, valor_unitario, 
                           stock_minimo, ubicacion, proveedor, tipo_producto,
                           (saldo * valor_unitario) as valor_total
                    FROM productos_poscosecha 
                    ORDER BY categoria, codigo
                """)
                
                productos = cursor.fetchall()
                
                with open(archivo, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['=== REPORTE DE POSCOSECHA ==='])
                    writer.writerow([f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
                    writer.writerow([])
                    writer.writerow(['Código', 'Categoría', 'Producto', 'Saldo', 'Unidad', 'Valor Unit.',
                                   'Stock Mín.', 'Ubicación', 'Proveedor', 'Tipo', 'Valor Total'])
                    
                    for producto in productos:
                        writer.writerow(producto)
                
                messagebox.showinfo("Éxito", f"Reporte de poscosecha generado: {archivo}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error generando reporte: {e}")
    
    def mostrar_menu_contextual(self, event):
        """Menú contextual específico"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            
            menu = tk.Menu(self.window, tearoff=0)
            menu.add_command(label="✏️ Editar producto", command=self.editar_producto)
            menu.add_command(label="📥 Entrada", command=lambda: self.movimiento('entrada'))
            menu.add_command(label="📤 Salida", command=lambda: self.movimiento('salida'))
            menu.add_separator()
            menu.add_command(label="📦 Control empaque", command=self.control_empaque)
            menu.add_command(label="🧪 Tratamientos", command=self.control_tratamientos)
            
            menu.post(event.x_root, event.y_root)
    
    def sort_column(self, col):
        """Ordenar columna específica"""
        try:
            data = [(self.tree.set(child, col), child) for child in self.tree.get_children('')]
            data.sort()
            
            for index, (val, child) in enumerate(data):
                self.tree.move(child, '', index)
        except:
            pass
    
    def add_hover_effect(self, button, color):
        """Efecto hover específico"""
        def on_enter(e):
            button.config(bg=self.darken_color(color))
        def on_leave(e):
            button.config(bg=color)
        
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
    
    def darken_color(self, color):
        """Oscurecer color para hover"""
        colors = {
            "#27ae60": "#229954",
            "#16a085": "#138d75",
            "#3498db": "#2980b9",
            "#f39c12": "#e67e22",
            "#8e44ad": "#7d3c98",
            "#e74c3c": "#c0392b",
            "#95a5a6": "#7f8c8d",
            "#e67e22": "#d35400"
        }
        return colors.get(color, color)
    
    def cerrar_ventana(self):
        """Cerrar ventana específica"""
        if self.conn:
            self.conn.close()
        self.window.destroy()


# =================== VENTANAS AUXILIARES ESPECÍFICAS ===================

class ProductoPoscosechaWindow:
    """Ventana específica para productos de poscosecha"""
    
    def __init__(self, parent, main_window, modo="nuevo", producto=None):
        self.parent = parent
        self.main_window = main_window
        self.modo = modo
        self.producto = producto
        
        self.window = tk.Toplevel(parent)
        titulo = "✏️ Editar Producto Poscosecha" if modo == "editar" else "➕ Nuevo Producto Poscosecha"
        self.window.title(titulo)
        self.window.geometry("550x500")
        self.window.configure(bg='#f8f9fa')
        self.window.resizable(False, False)
        
        self.center_window()
        self.window.transient(parent)
        self.window.grab_set()
        
        self.create_widgets()
        
        if modo == "editar" and producto:
            self.cargar_datos()
    
    def center_window(self):
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (550 // 2)
        y = (self.window.winfo_screenheight() // 2) - (500 // 2)
        self.window.geometry(f"550x500+{x}+{y}")
    
    def create_widgets(self):
        # Header específico
        header = tk.Frame(self.window, bg='#16a085', height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        titulo = "✏️ Editar Producto Poscosecha" if self.modo == "editar" else "➕ Nuevo Producto Poscosecha"
        title_label = tk.Label(header, text=titulo,
                             font=('Segoe UI', 16, 'bold'),
                             bg='#16a085', fg='white')
        title_label.pack(expand=True)
        
        # Formulario específico
        content = tk.Frame(self.window, bg='#ffffff')
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Campos específicos para poscosecha
        campos = [
            ("Código:", "entry_codigo"),
            ("Categoría:", "combo_categoria"),
            ("Nombre del Producto:", "entry_nombre"),
            ("Tipo de Producto:", "combo_tipo"),
            ("Saldo Actual:", "entry_saldo"),
            ("Unidad:", "combo_unidad"),
            ("Valor Unitario:", "entry_valor"),
            ("Stock Mínimo:", "entry_stock_min"),
            ("Ubicación:", "entry_ubicacion"),
            ("Proveedor:", "entry_proveedor"),
            ("Lote:", "entry_lote")
        ]
        
        for i, (label_text, widget_name) in enumerate(campos):
            label = tk.Label(content, text=label_text,
                           font=('Segoe UI', 10, 'bold'),
                           bg='#ffffff', fg='#2c3e50')
            label.grid(row=i, column=0, sticky='w', pady=8, padx=(0, 15))
            
            if widget_name == "combo_categoria":
                widget = ttk.Combobox(content, values=["EMBALAJE", "QUIMICO", "ETIQUETA", "HERRAMIENTA", "GENERAL"],
                                    width=35, font=('Segoe UI', 10))
            elif widget_name == "combo_tipo":
                widget = ttk.Combobox(content, values=["EMPAQUE", "TRATAMIENTO", "IDENTIFICACION", "LIMPIEZA", "GENERAL"],
                                    width=35, font=('Segoe UI', 10))
            elif widget_name == "combo_unidad":
                widget = ttk.Combobox(content, values=["UND", "KG", "LT", "ML", "GR", "MT", "M2"],
                                    width=35, font=('Segoe UI', 10))
            else:
                widget = tk.Entry(content, width=40, font=('Segoe UI', 10),
                                relief='solid', bd=1)
            
            widget.grid(row=i, column=1, pady=8, sticky='ew')
            setattr(self, widget_name, widget)
        
        content.grid_columnconfigure(1, weight=1)
        
        # Botones específicos
        self.create_buttons(content)
    
    def create_buttons(self, parent):
        btn_frame = tk.Frame(parent, bg='#ffffff')
        btn_frame.grid(row=11, column=0, columnspan=2, pady=20)
        
        texto = "💾 Actualizar" if self.modo == "editar" else "💾 Guardar"
        save_btn = tk.Button(btn_frame, text=texto, command=self.guardar,
                           bg="#27ae60", fg="white", font=('Segoe UI', 11, 'bold'),
                           relief='flat', bd=0, padx=20, pady=10, cursor='hand2')
        save_btn.pack(side=tk.LEFT, padx=5)
        
        cancel_btn = tk.Button(btn_frame, text="❌ Cancelar", command=self.window.destroy,
                             bg="#e74c3c", fg="white", font=('Segoe UI', 11, 'bold'),
                             relief='flat', bd=0, padx=20, pady=10, cursor='hand2')
        cancel_btn.pack(side=tk.LEFT, padx=5)
    
    def cargar_datos(self):
        if not self.producto:
            return
        
        self.entry_codigo.insert(0, self.producto[1] or "")
        self.combo_categoria.set(self.producto[2] or "")
        self.entry_nombre.insert(0, self.producto[3] or "")
        self.entry_saldo.insert(0, str(self.producto[4] or ""))
        self.combo_unidad.set(self.producto[5] or "")
        self.entry_valor.insert(0, str(self.producto[6] or ""))
        self.entry_stock_min.insert(0, str(self.producto[7] or ""))
        self.entry_ubicacion.insert(0, self.producto[8] or "")
        self.entry_proveedor.insert(0, self.producto[9] or "")
        self.combo_tipo.set(self.producto[10] or "")
        self.entry_lote.insert(0, self.producto[12] or "")
        
        if self.modo == "editar":
            self.entry_codigo.config(state='disabled')
    
    def guardar(self):
        try:
            # Validaciones específicas
            if not all([self.entry_codigo.get().strip(), self.combo_categoria.get(),
                       self.entry_nombre.get().strip(), self.combo_unidad.get()]):
                messagebox.showerror("Error", "Completa los campos obligatorios")
                return
            
            # Validar números
            try:
                saldo = int(self.entry_saldo.get() or 0)
                valor = float(self.entry_valor.get() or 0)
                stock_min = int(self.entry_stock_min.get() or 0)
            except ValueError:
                messagebox.showerror("Error", "Los valores numéricos no son válidos")
                return
            
            cursor = self.main_window.conn.cursor()
            
            if self.modo == "nuevo":
                # Verificar código único
                cursor.execute("SELECT id FROM productos_poscosecha WHERE codigo = ?", 
                             (self.entry_codigo.get().strip(),))
                if cursor.fetchone():
                    messagebox.showerror("Error", "El código ya existe")
                    return
                
                # Insertar nuevo
                cursor.execute('''
                    INSERT INTO productos_poscosecha 
                    (codigo, categoria, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor, tipo_producto, lote)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    self.entry_codigo.get().strip(),
                    self.combo_categoria.get(),
                    self.entry_nombre.get().strip(),
                    saldo,
                    self.combo_unidad.get(),
                    valor,
                    stock_min,
                    self.entry_ubicacion.get().strip(),
                    self.entry_proveedor.get().strip(),
                    self.combo_tipo.get(),
                    self.entry_lote.get().strip()
                ))
                mensaje = "Producto de poscosecha creado exitosamente"
            else:
                # Actualizar existente
                cursor.execute('''
                    UPDATE productos_poscosecha SET 
                    categoria=?, nombre=?, saldo=?, unidad=?, valor_unitario=?, 
                    stock_minimo=?, ubicacion=?, proveedor=?, tipo_producto=?, lote=?
                    WHERE codigo=?
                ''', (
                    self.combo_categoria.get(),
                    self.entry_nombre.get().strip(),
                    saldo,
                    self.combo_unidad.get(),
                    valor,
                    stock_min,
                    self.entry_ubicacion.get().strip(),
                    self.entry_proveedor.get().strip(),
                    self.combo_tipo.get(),
                    self.entry_lote.get().strip(),
                    self.entry_codigo.get().strip()
                ))
                mensaje = "Producto de poscosecha actualizado exitosamente"
            
            self.main_window.conn.commit()
            messagebox.showinfo("Éxito", mensaje)
            
            # Actualizar tabla principal
            self.main_window.cargar_inventario()
            
            if self.modo == "nuevo":
                self.limpiar_campos()
            else:
                self.window.destroy()
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar: {e}")
    
    def limpiar_campos(self):
        for widget_name in ['entry_codigo', 'entry_nombre', 'entry_saldo',
                           'entry_valor', 'entry_stock_min', 'entry_ubicacion', 
                           'entry_proveedor', 'entry_lote']:
            widget = getattr(self, widget_name)
            widget.delete(0, tk.END)
        
        for combo_name in ['combo_categoria', 'combo_tipo', 'combo_unidad']:
            combo = getattr(self, combo_name)
            combo.set('')


class MovimientoPoscosechaWindow:
    """Ventana específica para movimientos de poscosecha"""
    
    def __init__(self, parent, main_window, tipo):
        self.parent = parent
        self.main_window = main_window
        self.tipo = tipo
        
        self.window = tk.Toplevel(parent)
        titulo = f"📥 Entrada Poscosecha" if tipo == 'entrada' else f"📤 Salida Poscosecha"
        self.window.title(titulo)
        self.window.geometry("500x400")
        self.window.configure(bg='#f8f9fa')
        self.window.resizable(False, False)
        
        self.center_window()
        self.window.transient(parent)
        self.window.grab_set()
        
        self.create_widgets()
    
    def center_window(self):
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (500 // 2)
        y = (self.window.winfo_screenheight() // 2) - (400 // 2)
        self.window.geometry(f"500x400+{x}+{y}")
    
    def create_widgets(self):
        # Header específico
        color = "#27ae60" if self.tipo == 'entrada' else "#f39c12"
        header = tk.Frame(self.window, bg=color, height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        icon = "📥" if self.tipo == 'entrada' else "📤"
        titulo = f"{icon} {'Entrada' if self.tipo == 'entrada' else 'Salida'} Poscosecha"
        title_label = tk.Label(header, text=titulo,
                             font=('Segoe UI', 16, 'bold'),
                             bg=color, fg='white')
        title_label.pack(expand=True)
        
        # Formulario
        content = tk.Frame(self.window, bg='#ffffff')
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Producto
        tk.Label(content, text="Producto Poscosecha:", 
                font=('Segoe UI', 10, 'bold'),
                bg='#ffffff', fg='#2c3e50').grid(row=0, column=0, sticky='w', pady=8, padx=(0, 10))
        
        self.combo_producto = ttk.Combobox(content, width=35, font=('Segoe UI', 10))
        self.combo_producto.grid(row=0, column=1, pady=8)
        self.cargar_productos()
        
        # Cantidad
        tk.Label(content, text="Cantidad:", 
                font=('Segoe UI', 10, 'bold'),
                bg='#ffffff', fg='#2c3e50').grid(row=1, column=0, sticky='w', pady=8, padx=(0, 10))
        
        self.entry_cantidad = tk.Entry(content, width=38, font=('Segoe UI', 10))
        self.entry_cantidad.grid(row=1, column=1, pady=8)
        
        # Responsable
        tk.Label(content, text="Responsable:", 
                font=('Segoe UI', 10, 'bold'),
                bg='#ffffff', fg='#2c3e50').grid(row=2, column=0, sticky='w', pady=8, padx=(0, 10))
        
        self.entry_responsable = tk.Entry(content, width=38, font=('Segoe UI', 10))
        self.entry_responsable.grid(row=2, column=1, pady=8)
        
        # Campos adicionales según tipo
        if self.tipo == 'entrada':
            tk.Label(content, text="Factura:", 
                    font=('Segoe UI', 10, 'bold'),
                    bg='#ffffff', fg='#2c3e50').grid(row=3, column=0, sticky='w', pady=8, padx=(0, 10))
            self.entry_factura = tk.Entry(content, width=38, font=('Segoe UI', 10))
            self.entry_factura.grid(row=3, column=1, pady=8)
            
            tk.Label(content, text="Proveedor:", 
                    font=('Segoe UI', 10, 'bold'),
                    bg='#ffffff', fg='#2c3e50').grid(row=4, column=0, sticky='w', pady=8, padx=(0, 10))
            self.entry_proveedor = tk.Entry(content, width=38, font=('Segoe UI', 10))
            self.entry_proveedor.grid(row=4, column=1, pady=8)
        else:
            tk.Label(content, text="Destino/Área:", 
                    font=('Segoe UI', 10, 'bold'),
                    bg='#ffffff', fg='#2c3e50').grid(row=3, column=0, sticky='w', pady=8, padx=(0, 10))
            self.entry_destino = tk.Entry(content, width=38, font=('Segoe UI', 10))
            self.entry_destino.grid(row=3, column=1, pady=8)
        
        # Observaciones
        tk.Label(content, text="Observaciones:", 
                font=('Segoe UI', 10, 'bold'),
                bg='#ffffff', fg='#2c3e50').grid(row=5, column=0, sticky='w', pady=8, padx=(0, 10))
        self.entry_observaciones = tk.Entry(content, width=38, font=('Segoe UI', 10))
        self.entry_observaciones.grid(row=5, column=1, pady=8)
        
        # Botones
        btn_frame = tk.Frame(content, bg='#ffffff')
        btn_frame.grid(row=6, column=0, columnspan=2, pady=20)
        
        tk.Button(btn_frame, text=f"💾 Registrar {self.tipo.title()}", 
                 command=self.registrar, bg=color, fg="white",
                 font=('Segoe UI', 11, 'bold'), relief='flat', bd=0, 
                 padx=20, pady=10, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="❌ Cancelar", command=self.window.destroy,
                 bg="#e74c3c", fg="white", font=('Segoe UI', 11, 'bold'),
                 relief='flat', bd=0, padx=20, pady=10, cursor='hand2').pack(side=tk.LEFT, padx=5)
    
    def cargar_productos(self):
        cursor = self.main_window.conn.cursor()
        cursor.execute("SELECT codigo, nombre FROM productos_poscosecha ORDER BY codigo")
        productos = [f"{row[0]} - {row[1]}" for row in cursor.fetchall()]
        self.combo_producto['values'] = productos
    
    def registrar(self):
        try:
            if not all([self.combo_producto.get(), self.entry_cantidad.get(), self.entry_responsable.get()]):
                messagebox.showerror("Error", "Completa los campos obligatorios")
                return
            
            codigo = self.combo_producto.get().split(' - ')[0]
            cantidad = int(self.entry_cantidad.get())
            
            if cantidad <= 0:
                raise ValueError("Cantidad debe ser mayor a 0")
            
            cursor = self.main_window.conn.cursor()
            
            # Obtener producto
            cursor.execute("SELECT id, saldo FROM productos_poscosecha WHERE codigo = ?", (codigo,))
            resultado = cursor.fetchone()
            if not resultado:
                messagebox.showerror("Error", "Producto no encontrado")
                return
            
            producto_id, saldo_actual = resultado
            
            # Validar stock para salidas
            if self.tipo == 'salida' and cantidad > saldo_actual:
                messagebox.showerror("Error", f"Stock insuficiente. Disponible: {saldo_actual}")
                return
            
            # Nuevo saldo
            nuevo_saldo = saldo_actual + cantidad if self.tipo == 'entrada' else saldo_actual - cantidad
            
            # Registrar movimiento
            factura = getattr(self, 'entry_factura', None)
            proveedor = getattr(self, 'entry_proveedor', None)
            destino = getattr(self, 'entry_destino', None)
            
            cursor.execute('''
                INSERT INTO movimientos_poscosecha 
                (producto_id, tipo, cantidad, fecha, factura, proveedor, destino, observaciones, responsable)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                producto_id, self.tipo, cantidad, date.today(),
                factura.get() if factura else None,
                proveedor.get() if proveedor else None,
                destino.get() if destino else None,
                self.entry_observaciones.get(),
                self.entry_responsable.get()
            ))
            
            # Actualizar saldo
            cursor.execute("UPDATE productos_poscosecha SET saldo = ? WHERE id = ?", 
                         (nuevo_saldo, producto_id))
            
            self.main_window.conn.commit()
            messagebox.showinfo("Éxito", f"{self.tipo.title()} de poscosecha registrada correctamente")
            
            # Actualizar tabla
            self.main_window.cargar_inventario()
            self.window.destroy()
            
        except ValueError as e:
            messagebox.showerror("Error", f"Cantidad inválida: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Error registrando movimiento: {e}")