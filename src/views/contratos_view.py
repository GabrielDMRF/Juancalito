import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import os
from datetime import datetime, date
from pathlib import Path
from tkcalendar import DateEntry

# Agregar path para importar modelos
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from models.database import get_db, Empleado, Contrato, TipoContrato

# ===================== IMPORTAR GENERADORES =====================
# Generador de contratos Excel (HABILITADO)
EXCEL_GENERATOR_AVAILABLE = True
print("Generador de contratos Excel habilitado")

def abrir_generador_contratos_excel(parent, contratos_window, contrato):
    """Generar Excel de contratos usando la plantilla real"""
    try:
        from tkinter import messagebox
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        import os
        
        # Obtener empleado
        empleado = contratos_window.db.query(Empleado).filter(Empleado.id == contrato.empleado_id).first()
        if not empleado:
            messagebox.showerror("Error", "Empleado no encontrado")
            return
        
        # Ruta de la plantilla
        template_path = "templates_contratos/CONTRATO EXCEL FLORE JUNCALITO.xlsx"
        if not os.path.exists(template_path):
            messagebox.showerror("Error", f"Plantilla no encontrada: {template_path}")
            return
        
        # Crear carpeta del empleado
        nombre_seguro = empleado.nombre_completo.replace(" ", "_").replace("/", "_")
        cedula_segura = empleado.cedula.replace(" ", "_")
        empleado_dir = os.path.join("empleados_data", f"{nombre_seguro}_{cedula_segura}")
        contratos_dir = os.path.join(empleado_dir, "contratos")
        
        # Crear carpetas con información de debug
        print(f"Creando carpeta empleado: {empleado_dir}")
        os.makedirs(empleado_dir, exist_ok=True)
        
        print(f"Creando carpeta contratos: {contratos_dir}")
        os.makedirs(contratos_dir, exist_ok=True)
        
        # Verificar que las carpetas existen
        if not os.path.exists(empleado_dir):
            messagebox.showerror("Error", f"No se pudo crear la carpeta del empleado: {empleado_dir}")
            return
            
        if not os.path.exists(contratos_dir):
            messagebox.showerror("Error", f"No se pudo crear la carpeta de contratos: {contratos_dir}")
            return
        
        # Cargar la plantilla
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Función para convertir fechas a español
        def fecha_a_espanol(fecha):
            if not fecha:
                return "NO DEFINIDA"
            meses_espanol = {
                1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
            }
            dia = fecha.day
            mes = meses_espanol[fecha.month]
            anio = fecha.year
            return f"{dia} DE {mes} DE {anio}"
        
        # Función para convertir números a letras
        def numero_a_letras(numero):
            if numero is None or numero == 0:
                return "CERO PESOS"
            numero = int(numero)
            return f"{numero:,} PESOS"
        
        # Variables disponibles
        variables = {
            # Variables del empleado
            '{NOMBRE_EMPLEADO}': empleado.nombre_completo or "NO DEFINIDO",
            '{CEDULA_EMPLEADO}': empleado.cedula or "NO DEFINIDO",
            '{DIRECCION_EMPLEADO}': empleado.direccion or "DIRECCION NO ESPECIFICADA",
            '{TELEFONO_EMPLEADO}': empleado.telefono or "NO DEFINIDO",
            '{EMAIL_EMPLEADO}': empleado.email or "NO DEFINIDO",
            '{CARGO_EMPLEADO}': empleado.cargo or "OPERARIO OFICIOS VARIOS",
            '{AREA_TRABAJO}': empleado.area_trabajo or "NO DEFINIDO",
            
            # Variables del contrato
            '{NUMERO_CONTRATO}': contrato.numero_contrato or "NO DEFINIDO",
            '{TIPO_CONTRATO}': contrato.tipo_contrato_rel.nombre if contrato.tipo_contrato_rel else "NO DEFINIDO",
            '{SALARIO_NUMEROS}': f"$ {contrato.salario_contrato:,}" if contrato.salario_contrato else "NO DEFINIDO",
            '{SALARIO_LETRAS}': numero_a_letras(contrato.salario_contrato) if contrato.salario_contrato else "NO DEFINIDO",
            '{FECHA_INICIO}': fecha_a_espanol(contrato.fecha_inicio) if contrato.fecha_inicio else "NO DEFINIDA",
            '{FECHA_FIN}': fecha_a_espanol(contrato.fecha_fin) if contrato.fecha_fin else "NO DEFINIDA",
            '{ESTADO_CONTRATO}': contrato.estado or "NO DEFINIDO",
            
            # Variables de fechas y tiempos
            '{FECHA_GENERACION}': datetime.now().strftime('%d/%m/%Y'),
            '{FECHA_ACTUAL}': datetime.now().strftime('%d/%m/%Y'),
            '{HORA_GENERACION}': datetime.now().strftime('%H:%M:%S'),
            
            # Variables de la empresa
            '{NOMBRE_EMPRESA}': "FLORES JUNCALITO S.A.S",
            '{DIRECCION_EMPRESA}': "CALLE 19* C N. 88-07",
            '{CIUDAD_EMPRESA}': "EL ROSAL CUNDINAMARCA",
            '{NOMBRE_EMPLEADOR}': "FLORES JUNCALITO S.A.S",
            '{DIRECCION_EMPLEADOR}': "CALLE 19* C N. 88-07",
            
            # Variables adicionales de la plantilla
            '{LUGAR_NACIMIENTO}': empleado.lugar_nacimiento or "EL ROSAL CUNDINAMARCA",
            '{FECHA_NACIMIENTO}': fecha_a_espanol(empleado.fecha_nacimiento) if empleado.fecha_nacimiento else "NO DEFINIDA",
            '{TIPO_SALARIO}': "ORDINARIO" if contrato.salario_contrato and contrato.salario_contrato < 10000000 else "INTEGRAL",
            '{CIUDAD_CONTRATACION}': "EL ROSAL CUNDINAMARCA",
            '{PERIODOS_PAGO}': "MENSUAL",
            '{FECHA_INICIO_LABORES}': fecha_a_espanol(contrato.fecha_inicio) if contrato.fecha_inicio else "NO DEFINIDA",
            '{LUGAR_LABORES}': "EL ROSAL CUNDINAMARCA",
            '{VENCE_EL_DIA}': fecha_a_espanol(contrato.fecha_fin) if contrato.fecha_fin else "NO DEFINIDA",
            
            # Variables adicionales comunes
            '{NOMBRE_TRABAJADOR}': empleado.nombre_completo or "NO DEFINIDO",
            '{DIRECCION_TRABAJADOR}': empleado.direccion or "DIRECCION NO ESPECIFICADA",
            '{CARGO_OFICIO}': empleado.cargo or "OPERARIO OFICIOS VARIOS",
            '{VALOR_SALARIO}': f"$ {contrato.salario_contrato:,}" if contrato.salario_contrato else "NO DEFINIDO",
            '{VALOR_LETRAS}': numero_a_letras(contrato.salario_contrato) if contrato.salario_contrato else "NO DEFINIDO",
            '{FECHA_INICIACION}': fecha_a_espanol(contrato.fecha_inicio) if contrato.fecha_inicio else "NO DEFINIDA",
            '{LUGAR_DESEMPEÑO}': "EL ROSAL CUNDINAMARCA",
            '{TERMINO_INICIAL}': fecha_a_espanol(contrato.fecha_inicio) if contrato.fecha_inicio else "NO DEFINIDA",
            '{FECHA_VENCIMIENTO}': fecha_a_espanol(contrato.fecha_fin) if contrato.fecha_fin else "NO DEFINIDA"
        }
        
        # Reemplazar variables en toda la plantilla
        variables_encontradas = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    valor_original = cell.value
                    valor_nuevo = valor_original
                    
                    for variable, valor in variables.items():
                        if variable in valor_original:
                            valor_nuevo = valor_nuevo.replace(variable, str(valor))
                            variables_encontradas += 1
                    
                    if valor_nuevo != valor_original:
                        cell.value = valor_nuevo
        
        # Guardar archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"contrato_{cedula_segura}_{timestamp}.xlsx"
        filepath = os.path.join(contratos_dir, filename)
        
        print(f"Guardando archivo en: {filepath}")
        wb.save(filepath)
        
        # Verificar que el archivo se guardó correctamente
        if os.path.exists(filepath):
            print(f"Archivo guardado exitosamente en: {filepath}")
            messagebox.showinfo("Éxito", f"Contrato Excel generado usando plantilla:\n{filepath}\n\nVariables reemplazadas: {variables_encontradas}")
        else:
            print(f"Error: El archivo no se guardó en: {filepath}")
            messagebox.showerror("Error", f"No se pudo guardar el archivo en: {filepath}")
            return
        
        # Abrir archivo
        try:
            os.startfile(filepath)
        except:
            pass
            
    except Exception as e:
        messagebox.showerror("Error", f"Error generando Excel: {e}")
        print(f"Error generador Excel: {e}")



class ContratosWindow:
    def __init__(self, parent, main_window):
        self.parent = parent
        self.main_window = main_window
        self.db = get_db()
        
        # Crear ventana de Contratos responsive
        self.window = tk.Toplevel(parent)
        self.window.title("Gestión de Contratos")
        self.window.configure(bg='#ecf0f1')
        self.window.resizable(True, True)
        self.window.minsize(900, 500)
        
        # Configurar tamaño inicial responsive
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Calcular tamaño inicial (80% de la pantalla)
        window_width = min(1200, int(screen_width * 0.8))
        window_height = min(750, int(screen_height * 0.8))
        
        # Centrar ventana
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Configurar expansión de grid
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        # No hacer modal para que tenga controles estándar de Windows
        # self.window.transient(parent)
        # self.window.grab_set()
        
        self.create_widgets()
        
        # Configurar evento de cierre
        self.window.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Cargar datos iniciales de contratos después de que todo esté configurado
        self.window.after(100, self.cargar_contratos)
    
    def on_closing(self):
        """Manejar el cierre de la ventana"""
        try:
            # Limpiar referencias en main_window
            if hasattr(self.main_window, 'contratos_window'):
                self.main_window.contratos_window = None
            
            # Limpiar referencias locales
            if hasattr(self, 'tree'):
                self.tree = None
            if hasattr(self, 'db'):
                self.db = None
                
            # Destruir ventana
            if hasattr(self, 'window') and self.window:
                self.window.destroy()
                
        except Exception as e:
            print(f"Error cerrando ventana de contratos: {e}")
            # Forzar destrucción si hay error
            try:
                if hasattr(self, 'window'):
                    self.window.quit()
            except:
                pass
    

    
    def create_widgets(self):
        # Frame principal responsive
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar expansión del frame principal
        main_frame.grid_rowconfigure(3, weight=1)  # La tabla se expandirá
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Título responsive
        title = tk.Label(main_frame, text="📋 Gestión de Contratos Laborales", 
                         font=('Segoe UI', 18, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        title.grid(row=0, column=0, columnspan=4, pady=(0, 25))
        
        # Frame para botones principales responsive
        btn_frame = tk.Frame(main_frame, bg='#ecf0f1')
        btn_frame.grid(row=1, column=0, columnspan=4, pady=(0, 20), sticky=(tk.W, tk.E))
        btn_frame.grid_columnconfigure(0, weight=1)
        
        # Botones principales con iconos y colores mejorados
        buttons_info = [
            ("➕ Nuevo Contrato", self.nuevo_contrato, "#27ae60", "Crear nuevo contrato"),
            ("✏️ Editar Contrato", self.editar_contrato, "#3498db", "Modificar contrato seleccionado"),
            ("👁️ Ver Detalles", self.ver_contrato, "#f39c12", "Ver información completa"),
            ("📊 Generar Excel", self.generar_contrato_excel, "#16a085", "Generar contrato en Excel"),
            ("📁 Abrir Carpeta", self.abrir_carpeta_empleado, "#9b59b6", "Abrir carpeta del empleado"),
            ("📄 Abrir Documento", self.abrir_documento_empleado, "#e67e22", "Abrir último documento generado"),
            ("🔄 Actualizar", self.cargar_contratos, "#34495e", "Actualizar lista"),
            ("❌ Cerrar", self.window.destroy, "#e74c3c", "Cerrar ventana")
        ]
        
        # Configurar grid para botones responsive
        for i in range(8):
            btn_frame.grid_columnconfigure(i, weight=1)
        
        for i, (text, command, color, tooltip) in enumerate(buttons_info):
            btn = tk.Button(btn_frame, text=text, command=command,
                           bg=color, fg='white', font=('Segoe UI', 10, 'bold'),
                           relief='flat', bd=0, padx=15, pady=8, cursor='hand2',
                           activebackground=self.darken_color(color))
            btn.grid(row=0, column=i, padx=3, pady=2, sticky=(tk.W, tk.E))
            
            # Agregar efecto hover
            self.add_hover_effect(btn, color)
            
            # Tooltip (simplificado)
            btn.bind("<Enter>", lambda e, t=tooltip: self.show_tooltip(e, t))
            btn.bind("<Leave>", lambda e: self.hide_tooltip())
        
        # Frame para información y estadísticas
        info_frame = tk.LabelFrame(main_frame, text="📊 Información General", 
                                  font=('Segoe UI', 12, 'bold'),
                                  bg='#ecf0f1', fg='#2c3e50', padx=15, pady=10)
        info_frame.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Labels de estadísticas
        self.stats_label = tk.Label(info_frame, text="Cargando estadísticas...", 
                                   font=('Segoe UI', 10),
                                   bg='#ecf0f1', fg='#7f8c8d')
        self.stats_label.pack()
        
        # Lista de contratos con diseño mejorado
        contratos_frame = ttk.LabelFrame(main_frame, text="📄 Lista de Contratos", padding="15")
        contratos_frame.grid(row=3, column=0, columnspan=4, pady=(15, 0), sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Treeview responsive para mostrar contratos con más columnas
        columns = ('ID', 'Número', 'Empleado', 'Cédula', 'Tipo', 'Inicio', 'Fin', 'Salario', 'Estado', 'Archivos')
        self.tree = ttk.Treeview(contratos_frame, columns=columns, show='headings')
        
        # Configurar columnas con anchos mínimos y expansión
        column_widths = {
            'ID': 50, 'Número': 120, 'Empleado': 200, 'Cédula': 100,
            'Tipo': 120, 'Inicio': 100, 'Fin': 100, 'Salario': 120, 
            'Estado': 100, 'Archivos': 80
        }
        
        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c))
            self.tree.column(col, width=column_widths.get(col, 100), minwidth=50)
        
        # Scrollbars mejoradas
        v_scrollbar = ttk.Scrollbar(contratos_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(contratos_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid para tree y scrollbars
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Eventos
        self.tree.bind('<Double-1>', lambda e: self.ver_contrato())
        self.tree.bind('<Button-3>', self.mostrar_menu_contextual)  # Click derecho
        
        # Configurar grid weights para responsividad
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        contratos_frame.columnconfigure(0, weight=1)
        contratos_frame.rowconfigure(0, weight=1)
        
        # Actualizar estadísticas iniciales
        self.actualizar_estadisticas()
    
    def add_hover_effect(self, button, color):
        """Agregar efecto hover a botones"""
        def on_enter(e):
            button.config(bg=self.darken_color(color))
        def on_leave(e):
            button.config(bg=color)
        
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
    
    def darken_color(self, hex_color):
        """Oscurecer color para efecto hover"""
        color_map = {
            "#27ae60": "#229954", "#3498db": "#2980b9", "#f39c12": "#e67e22",
            "#16a085": "#138d75", "#9b59b6": "#8e44ad", "#34495e": "#2c3e50",
            "#e74c3c": "#c0392b"
        }
        return color_map.get(hex_color, hex_color)
    
    def show_tooltip(self, event, text):
        """Mostrar tooltip simple"""
        # Implementación simplificada de tooltip
        pass
    
    def hide_tooltip(self):
        """Ocultar tooltip"""
        pass
    
    def sort_column(self, col):
        """Ordenar columnas"""
        try:
            data = [(self.tree.set(child, col), child) for child in self.tree.get_children('')]
            data.sort()
            
            for index, (val, child) in enumerate(data):
                self.tree.move(child, '', index)
        except Exception as e:
            print(f"Error ordenando columna {col}: {e}")
    
    def mostrar_menu_contextual(self, event):
        """Mostrar menú contextual con click derecho"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            
            menu = tk.Menu(self.window, tearoff=0)
            menu.add_command(label="👁️ Ver Detalles", command=self.ver_contrato)
            menu.add_command(label="✏️ Editar", command=self.editar_contrato)
            menu.add_separator()
            menu.add_command(label="📊 Generar Excel", command=self.generar_contrato_excel)
            
            menu.post(event.x_root, event.y_root)
    
    def cargar_contratos(self):
        """Cargar todos los contratos en el TreeView"""
        try:
            # Verificar que el tree existe y no ha sido destruido
            if not hasattr(self, 'tree'):
                print("TreeView no existe")
                return
                
            # Verificar si la ventana principal sigue existiendo
            try:
                self.window.winfo_exists()
                self.tree.winfo_exists()
            except:
                print("Ventana o TreeView fueron destruidos")
                return
                
            # Limpiar datos existentes de forma segura
            try:
                for item in self.tree.get_children():
                    self.tree.delete(item)
            except Exception as clear_error:
                print(f"Error limpiando TreeView: {clear_error}")
                return
            
            # Cargar contratos con información del empleado
            contratos = self.db.query(Contrato).join(Empleado).all()
            
            print(f"Procesando {len(contratos)} contratos...")
            
            for contrato in contratos:
                try:
                    # Formatear fechas
                    fecha_inicio = contrato.fecha_inicio.strftime("%d/%m/%Y") if contrato.fecha_inicio else "No definida"
                    fecha_fin = contrato.fecha_fin.strftime("%d/%m/%Y") if contrato.fecha_fin else "No definida"
                    
                    # Formatear salario
                    salario = f"${contrato.salario_base:,}" if contrato.salario_base else "No definido"
                    
                    # Función para limpiar caracteres problemáticos
                    def limpiar_texto_para_display(texto):
                        if texto is None:
                            return "No definido"
                        # Convertir a string y limpiar caracteres problemáticos
                        texto_str = str(texto).strip()
                        if not texto_str:
                            return "No definido"
                        # Reemplazar emojis problemáticos con texto
                        replacements = {
                            '📝': '[BORR]',
                            '✅': '[ACTV]',
                            '⏰': '[VENC]',
                            '❌': '[TERM]',
                            '⏸️': '[SUSP]',
                            '📊': '[EXCEL]',
                            '📄': '[DOC]'
                        }
                        for emoji, replacement in replacements.items():
                            texto_str = texto_str.replace(emoji, replacement)
                        return texto_str
                    
                    # Estado con texto seguro
                    estado_display = {
                        'borrador': 'BORRADOR',
                        'activo': 'ACTIVO',
                        'vencido': 'VENCIDO',
                        'terminado': 'TERMINADO',
                        'suspendido': 'SUSPENDIDO'
                    }.get(contrato.estado, contrato.estado or "Sin estado")
                    
                    # Verificar archivos generados
                    archivos_info = self.verificar_archivos_contrato(contrato)
                    
                    # Limpiar todos los textos antes de mostrar
                    numero_contrato_limpio = limpiar_texto_para_display(contrato.numero_contrato or "Sin número")
                    nombre_empleado_limpio = limpiar_texto_para_display(contrato.empleado.nombre_completo)
                    cedula_limpia = limpiar_texto_para_display(contrato.empleado.cedula)
                    archivos_info_limpio = limpiar_texto_para_display(archivos_info)
                    
                    # Obtener tipo de contrato con manejo de errores
                    tipo_contrato = "No definido"
                    try:
                        if contrato.tipo_contrato_rel:
                            tipo_contrato = contrato.tipo_contrato_rel.nombre
                        elif contrato.tipo_contrato_id:
                            # Intentar obtener el tipo de contrato directamente
                            tipo_contrato_obj = self.db.query(TipoContrato).filter(TipoContrato.id == contrato.tipo_contrato_id).first()
                            if tipo_contrato_obj:
                                tipo_contrato = tipo_contrato_obj.nombre
                    except Exception as tipo_error:
                        print(f"Error obteniendo tipo de contrato para contrato {contrato.id}: {tipo_error}")
                        tipo_contrato = "Error"
                    
                    # Insertar en TreeView con textos limpios
                    self.tree.insert('', 'end', values=(
                        contrato.id,
                        numero_contrato_limpio,
                        nombre_empleado_limpio,
                        cedula_limpia,
                        limpiar_texto_para_display(tipo_contrato),
                        fecha_inicio,
                        fecha_fin,
                        salario,
                        estado_display,
                        archivos_info_limpio
                    ))
                    
                except Exception as contrato_error:
                    print(f"Error procesando contrato {contrato.id}: {contrato_error}")
                    # Insertar contrato con datos básicos
                    self.tree.insert('', 'end', values=(
                        contrato.id,
                        contrato.numero_contrato or "Sin número",
                        "Error al cargar",
                        "Error al cargar",
                        "Error",
                        "Error",
                        "Error",
                        "Error",
                        "Error",
                        "Error"
                    ))
            
            print(f"Se cargaron {len(contratos)} contratos")
            self.actualizar_estadisticas()
            
        except Exception as e:
            print(f"Error al cargar contratos: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Error al cargar contratos: {e}")
    
    def verificar_archivos_contrato(self, contrato):
        """Verificar qué archivos Excel existen para el contrato"""
        try:
            empleado = contrato.empleado
            nombre_seguro = empleado.nombre_completo.replace(" ", "_").replace("/", "_")
            cedula_segura = empleado.cedula.replace(" ", "_")
            
            contratos_dir = Path(f"empleados_data/{nombre_seguro}_{cedula_segura}/contratos")
            if contratos_dir.exists():
                excel_files = list(contratos_dir.glob("contrato_excel_*.xlsx"))
                if excel_files:
                    return f"EXCEL ({len(excel_files)} archivos)"
        except:
            pass
        
        return "Sin archivos"
    
    def actualizar_estadisticas(self):
        """Actualizar estadísticas de contratos"""
        try:
            total = self.db.query(Contrato).count()
            activos = self.db.query(Contrato).filter(Contrato.estado == 'activo').count()
            borradores = self.db.query(Contrato).filter(Contrato.estado == 'borrador').count()
            
            stats_text = f"📊 Total: {total} contratos | ✅ Activos: {activos} | 📝 Borradores: {borradores}"
            
            # Agregar información de generador disponible
            if EXCEL_GENERATOR_AVAILABLE:
                stats_text += " | 📊 Generador Excel disponible"
            
            self.stats_label.config(text=stats_text)
            
        except Exception as e:
            self.stats_label.config(text=f"❌ Error cargando estadísticas: {e}")
    
    def get_selected_contrato(self):
        """Obtener el contrato seleccionado en el TreeView"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor selecciona un contrato")
            return None
        
        contrato_id = self.tree.item(selection[0])['values'][0]
        return self.db.query(Contrato).filter(Contrato.id == contrato_id).first()
    
    def nuevo_contrato(self):
        """Abrir ventana para crear un nuevo contrato"""
        NuevoContratoWindow(self.window, self)
    
    def editar_contrato(self):
        """Abrir ventana para editar el contrato seleccionado"""
        contrato = self.get_selected_contrato()
        if contrato:
            NuevoContratoWindow(self.window, self, contrato=contrato)
    
    def ver_contrato(self):
        """Abrir ventana de detalles del contrato seleccionado"""
        contrato = self.get_selected_contrato()
        if contrato:
            DetallesContratoWindow(self.window, contrato)
    
    # =================== GENERADORES DE CONTRATOS ===================
    
    def generar_contrato_excel(self):
        """Generar contrato en formato Excel usando la plantilla"""
        contrato = self.get_selected_contrato()
        if not contrato:
            return
        
        if not EXCEL_GENERATOR_AVAILABLE:
            messagebox.showinfo(
                "Generador Excel No Disponible",
                "El generador de contratos Excel no está disponible.\n\n"
                "Verifica que:\n"
                "• El archivo utils/contrato_excel_generator.py existe\n"
                "• La librería openpyxl está instalada: pip install openpyxl\n"
                "• La plantilla Excel está en la carpeta plantillas/"
            )
            return
        
        try:
            # Abrir ventana del generador integrado
            abrir_generador_contratos_excel(self.window, self, contrato)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error abriendo generador Excel: {e}")
            print(f"Error generador Excel: {e}")
    
    def abrir_carpeta_empleado(self):
        """Abrir carpeta del empleado seleccionado"""
        contrato = self.get_selected_contrato()
        if not contrato:
            return
        
        try:
            empleado = contrato.empleado
            nombre_seguro = empleado.nombre_completo.replace(" ", "_").replace("/", "_")
            cedula_segura = empleado.cedula.replace(" ", "_")
            
            carpeta_empleado = Path(f"empleados_data/{nombre_seguro}_{cedula_segura}")
            
            if carpeta_empleado.exists():
                import subprocess
                import platform
                
                if platform.system() == "Windows":
                    os.startfile(str(carpeta_empleado))
                elif platform.system() == "Darwin":  # macOS
                    subprocess.run(["open", str(carpeta_empleado)])
                else:  # Linux
                    subprocess.run(["xdg-open", str(carpeta_empleado)])
                
                messagebox.showinfo("Éxito", f"Carpeta abierta: {carpeta_empleado}")
            else:
                messagebox.showwarning("Advertencia", f"No se encontró la carpeta: {carpeta_empleado}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error abriendo carpeta: {e}")
    
    def abrir_documento_empleado(self):
        """Abrir último documento generado del empleado"""
        contrato = self.get_selected_contrato()
        if not contrato:
            return
        
        try:
            empleado = contrato.empleado
            nombre_seguro = empleado.nombre_completo.replace(" ", "_").replace("/", "_")
            cedula_segura = empleado.cedula.replace(" ", "_")
            
            carpeta_contratos = Path(f"empleados_data/{nombre_seguro}_{cedula_segura}/contratos")
            
            if carpeta_contratos.exists():
                # Buscar el archivo más reciente
                archivos_excel = list(carpeta_contratos.glob("*.xlsx"))
                
                if archivos_excel:
                    # Ordenar por fecha de modificación (más reciente primero)
                    archivo_mas_reciente = max(archivos_excel, key=lambda x: x.stat().st_mtime)
                    
                    import subprocess
                    import platform
                    
                    if platform.system() == "Windows":
                        os.startfile(str(archivo_mas_reciente))
                    elif platform.system() == "Darwin":  # macOS
                        subprocess.run(["open", str(archivo_mas_reciente)])
                    else:  # Linux
                        subprocess.run(["xdg-open", str(archivo_mas_reciente)])
                    
                    messagebox.showinfo("Éxito", f"Documento abierto: {archivo_mas_reciente.name}")
                else:
                    messagebox.showwarning("Advertencia", "No se encontraron documentos Excel en la carpeta")
            else:
                messagebox.showwarning("Advertencia", f"No se encontró la carpeta de contratos")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error abriendo documento: {e}")
    



# =================== VENTANAS AUXILIARES ===================

class NuevoContratoWindow:
    """Ventana para crear/editar contratos"""
    def __init__(self, parent, contratos_window, contrato=None):
        self.parent = parent
        self.contratos_window = contratos_window
        self.contrato = contrato
        self.db = contratos_window.db
        
        # Crear ventana responsive
        self.window = tk.Toplevel(parent)
        self.window.title("Nuevo Contrato" if not contrato else "Editar Contrato")
        self.window.configure(bg='#f8f9fa')
        
        # Configurar ventana responsive
        self.window.resizable(True, True)
        self.window.minsize(600, 700)
        
        # Configurar tamaño inicial responsive
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Calcular tamaño inicial (65% de la pantalla)
        window_width = min(700, int(screen_width * 0.65))
        window_height = min(800, int(screen_height * 0.65))
        
        # Centrar ventana
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Configurar expansión de grid
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        
        self.window.transient(parent)
        self.window.grab_set()
        
        self.create_widgets()
        if contrato:
            self.cargar_datos_contrato()
    

    
    def create_widgets(self):
        # Frame principal responsive
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar expansión del frame principal
        main_frame.grid_rowconfigure(8, weight=1)  # Los botones se expandirán
        main_frame.grid_columnconfigure(1, weight=1)  # Los campos se expandirán
        
        # Título responsive
        title_text = "Nuevo Contrato" if not self.contrato else "Editar Contrato"
        title = tk.Label(main_frame, text=f"📋 {title_text}", 
                        font=('Segoe UI', 16, 'bold'), bg='#f8f9fa', fg='#2c3e50')
        title.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Variables
        self.numero_var = tk.StringVar()
        self.empleado_var = tk.StringVar()
        self.tipo_var = tk.StringVar()
        self.fecha_inicio_var = tk.StringVar()
        self.fecha_fin_var = tk.StringVar()
        self.salario_var = tk.StringVar()
        self.estado_var = tk.StringVar(value="borrador")
        
        # Campos del formulario
        row = 1
        
        # Número de contrato responsive
        tk.Label(main_frame, text="Número de Contrato:", font=('Segoe UI', 10, 'bold'), 
                bg='#f8f9fa').grid(row=row, column=0, sticky=tk.W, pady=5)
        tk.Entry(main_frame, textvariable=self.numero_var, font=('Segoe UI', 10)).grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 10))
        
        # Empleado responsive
        row += 1
        tk.Label(main_frame, text="Empleado:", font=('Segoe UI', 10, 'bold'), 
                bg='#f8f9fa').grid(row=row, column=0, sticky=tk.W, pady=5)
        self.empleado_combo = ttk.Combobox(main_frame, textvariable=self.empleado_var, 
                                          font=('Segoe UI', 10), state="readonly")
        self.empleado_combo.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 10))
        self.cargar_empleados()
        
        # Tipo de contrato responsive
        row += 1
        tk.Label(main_frame, text="Tipo de Contrato:", font=('Segoe UI', 10, 'bold'), 
                bg='#f8f9fa').grid(row=row, column=0, sticky=tk.W, pady=5)
        self.tipo_combo = ttk.Combobox(main_frame, textvariable=self.tipo_var, 
                                      font=('Segoe UI', 10), state="readonly")
        self.tipo_combo['values'] = ["temporal", "permanente", "temporada"]
        self.tipo_combo.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 10))
        
        # Fecha inicio con calendario responsive
        row += 1
        tk.Label(main_frame, text="Fecha de Inicio:", font=('Segoe UI', 10, 'bold'), 
                bg='#f8f9fa').grid(row=row, column=0, sticky=tk.W, pady=5)
        self.date_fecha_inicio = DateEntry(main_frame, background='darkblue', foreground='white', 
                                          borderwidth=2, date_pattern='dd/mm/yyyy', font=('Segoe UI', 10))
        self.date_fecha_inicio.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 10))
        
        # Fecha fin con calendario responsive
        row += 1
        tk.Label(main_frame, text="Fecha de Fin:", font=('Segoe UI', 10, 'bold'), 
                bg='#f8f9fa').grid(row=row, column=0, sticky=tk.W, pady=5)
        self.date_fecha_fin = DateEntry(main_frame, background='darkblue', foreground='white', 
                                       borderwidth=2, date_pattern='dd/mm/yyyy', font=('Segoe UI', 10))
        self.date_fecha_fin.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 10))
        
        # Salario responsive
        row += 1
        tk.Label(main_frame, text="Salario:", font=('Segoe UI', 10, 'bold'), 
                bg='#f8f9fa').grid(row=row, column=0, sticky=tk.W, pady=5)
        tk.Entry(main_frame, textvariable=self.salario_var, font=('Segoe UI', 10)).grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 10))
        
        # Estado responsive
        row += 1
        tk.Label(main_frame, text="Estado:", font=('Segoe UI', 10, 'bold'), 
                bg='#f8f9fa').grid(row=row, column=0, sticky=tk.W, pady=5)
        self.estado_combo = ttk.Combobox(main_frame, textvariable=self.estado_var, 
                                        font=('Segoe UI', 10), state="readonly")
        self.estado_combo['values'] = ["borrador", "activo", "finalizado", "cancelado"]
        self.estado_combo.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 10))
        
        # Botones responsive
        row += 2
        btn_frame = tk.Frame(main_frame, bg='#f8f9fa')
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20, sticky=(tk.W, tk.E))
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        
        tk.Button(btn_frame, text="💾 Guardar", command=self.guardar_contrato,
                 bg="#27ae60", fg="white", font=('Segoe UI', 12, 'bold'),
                 relief='flat', bd=0, padx=15, pady=10, cursor='hand2').grid(row=0, column=0, padx=5, sticky=(tk.W, tk.E))
        
        tk.Button(btn_frame, text="❌ Cancelar", command=self.window.destroy,
                 bg="#e74c3c", fg="white", font=('Segoe UI', 12, 'bold'),
                 relief='flat', bd=0, padx=15, pady=10, cursor='hand2').grid(row=0, column=1, padx=5, sticky=(tk.W, tk.E))
        
        # Configurar grid
        main_frame.columnconfigure(1, weight=1)
    
    def cargar_empleados(self):
        """Cargar lista de empleados activos"""
        try:
            empleados = self.db.query(Empleado).filter(Empleado.estado == True).all()
            empleados_list = [f"{emp.nombre_completo} ({emp.cedula})" for emp in empleados]
            self.empleado_combo['values'] = empleados_list
        except Exception as e:
            print(f"Error cargando empleados: {e}")
    
    def cargar_datos_contrato(self):
        """Cargar datos del contrato para edición"""
        if not self.contrato:
            return
        
        try:
            self.numero_var.set(self.contrato.numero_contrato or "")
            
            # Buscar empleado
            empleado = self.db.query(Empleado).filter(Empleado.id == self.contrato.empleado_id).first()
            if empleado:
                empleado_text = f"{empleado.nombre_completo} ({empleado.cedula})"
                self.empleado_var.set(empleado_text)
            
            # Obtener tipo de contrato
            if self.contrato.tipo_contrato_rel:
                self.tipo_var.set(self.contrato.tipo_contrato_rel.nombre)
            else:
                self.tipo_var.set("")
            
            if self.contrato.fecha_inicio:
                self.date_fecha_inicio.set_date(self.contrato.fecha_inicio)
            if self.contrato.fecha_fin:
                self.date_fecha_fin.set_date(self.contrato.fecha_fin)
            
            if self.contrato.salario_contrato:
                self.salario_var.set(str(self.contrato.salario_contrato))
            
            self.estado_var.set(self.contrato.estado or "borrador")
            
        except Exception as e:
            print(f"Error cargando datos del contrato: {e}")
    
    def guardar_contrato(self):
        """Guardar o actualizar contrato"""
        try:
            # Validar campos obligatorios
            if not self.numero_var.get().strip():
                messagebox.showerror("Error", "El número de contrato es obligatorio")
                return
            
            if not self.empleado_var.get().strip():
                messagebox.showerror("Error", "Debe seleccionar un empleado")
                return
            
            # Obtener empleado
            empleado_text = self.empleado_var.get()
            cedula = empleado_text.split('(')[-1].rstrip(')')
            empleado = self.db.query(Empleado).filter(Empleado.cedula == cedula).first()
            
            if not empleado:
                messagebox.showerror("Error", "Empleado no encontrado")
                return
            
            # Obtener tipo de contrato por ID
            tipo_nombre = self.tipo_var.get()
            tipo_contrato = self.db.query(TipoContrato).filter(TipoContrato.nombre == tipo_nombre).first()
            tipo_contrato_id = tipo_contrato.id if tipo_contrato else None
            
            # Crear o actualizar contrato
            if self.contrato:
                # Actualizar contrato existente
                self.contrato.numero_contrato = self.numero_var.get().strip()
                self.contrato.empleado_id = empleado.id
                self.contrato.tipo_contrato_id = tipo_contrato_id
                self.contrato.fecha_inicio = self.date_fecha_inicio.get_date()
                self.contrato.fecha_fin = self.date_fecha_fin.get_date()
                self.contrato.salario_contrato = int(self.salario_var.get()) if self.salario_var.get() else None
                self.contrato.estado = self.estado_var.get()
                mensaje = "Contrato actualizado exitosamente"
            else:
                # Crear nuevo contrato
                nuevo_contrato = Contrato(
                    numero_contrato=self.numero_var.get().strip(),
                    empleado_id=empleado.id,
                    tipo_contrato_id=tipo_contrato_id,
                    fecha_inicio=self.date_fecha_inicio.get_date(),
                    fecha_fin=self.date_fecha_fin.get_date(),
                    salario_contrato=int(self.salario_var.get()) if self.salario_var.get() else None,
                    estado=self.estado_var.get()
                )
                self.db.add(nuevo_contrato)
                mensaje = "Contrato creado exitosamente"
            
            self.db.commit()
            messagebox.showinfo("Éxito", mensaje)
            self.contratos_window.cargar_contratos()
            self.window.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error guardando contrato: {e}")
            print(f"Error guardando contrato: {e}")

class DetallesContratoWindow:
    """Ventana de detalles del contrato"""
    def __init__(self, parent, contrato):
        self.parent = parent
        self.contrato = contrato
        self.db = get_db()
        
        # Crear ventana responsive
        self.window = tk.Toplevel(parent)
        self.window.title(f"Detalles del Contrato - {contrato.numero_contrato or 'Sin número'}")
        self.window.configure(bg='#f8f9fa')
        
        # Configurar ventana responsive
        self.window.resizable(True, True)
        self.window.minsize(600, 500)
        
        # Configurar tamaño inicial responsive
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Calcular tamaño inicial (55% de la pantalla)
        window_width = min(700, int(screen_width * 0.55))
        window_height = min(600, int(screen_height * 0.55))
        
        # Centrar ventana
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Configurar expansión de grid
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        
        self.window.transient(parent)
        
        self.create_widgets()
        self.cargar_detalles()
    

    
    def create_widgets(self):
        # Frame principal responsive
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar expansión del frame principal
        main_frame.grid_rowconfigure(1, weight=1)  # El frame de detalles se expandirá
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Título responsive
        title = tk.Label(main_frame, text="📋 Detalles del Contrato", 
                        font=('Segoe UI', 16, 'bold'), bg='#f8f9fa', fg='#2c3e50')
        title.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Frame para detalles
        self.detalles_frame = tk.LabelFrame(main_frame, text="Información del Contrato", 
                                           font=('Segoe UI', 12, 'bold'), bg='#f8f9fa')
        self.detalles_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 20))
        
        # Botón cerrar responsive
        btn_frame = tk.Frame(main_frame, bg='#f8f9fa')
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10, sticky=(tk.W, tk.E))
        btn_frame.grid_columnconfigure(0, weight=1)
        
        tk.Button(btn_frame, text="❌ Cerrar", command=self.window.destroy,
                 bg="#e74c3c", fg="white", font=('Segoe UI', 12, 'bold'),
                 relief='flat', bd=0, padx=15, pady=10, cursor='hand2').grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # Configurar grid
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        self.detalles_frame.columnconfigure(1, weight=1)
    
    def cargar_detalles(self):
        """Cargar y mostrar detalles del contrato"""
        try:
            # Obtener empleado
            empleado = self.db.query(Empleado).filter(Empleado.id == self.contrato.empleado_id).first()
            
            # Crear etiquetas con información
            detalles = [
                ("Número de Contrato:", self.contrato.numero_contrato or "No definido"),
                ("Empleado:", f"{empleado.nombre_completo} ({empleado.cedula})" if empleado else "No encontrado"),
                ("Tipo de Contrato:", self.contrato.tipo_contrato_rel.nombre if self.contrato.tipo_contrato_rel else "No definido"),
                ("Fecha de Inicio:", self.contrato.fecha_inicio.strftime('%d/%m/%Y') if self.contrato.fecha_inicio else "No definida"),
                ("Fecha de Fin:", self.contrato.fecha_fin.strftime('%d/%m/%Y') if self.contrato.fecha_fin else "No definida"),
                ("Salario:", f"${self.contrato.salario_contrato:,}" if self.contrato.salario_contrato else "No definido"),
                ("Estado:", self.contrato.estado or "No definido"),
                ("Fecha de Creación:", self.contrato.fecha_creacion.strftime('%d/%m/%Y %H:%M') if self.contrato.fecha_creacion else "No definida")
            ]
            
            for i, (label, value) in enumerate(detalles):
                # Label
                tk.Label(self.detalles_frame, text=label, font=('Segoe UI', 10, 'bold'), 
                        bg='#f8f9fa', anchor='w').grid(row=i, column=0, sticky=tk.W, padx=10, pady=5)
                
                # Valor
                tk.Label(self.detalles_frame, text=value, font=('Segoe UI', 10), 
                        bg='#f8f9fa', anchor='w').grid(row=i, column=1, sticky=tk.W, padx=10, pady=5)
            
        except Exception as e:
            print(f"Error cargando detalles: {e}")
            tk.Label(self.detalles_frame, text=f"Error cargando detalles: {e}", 
                    font=('Segoe UI', 10), bg='#f8f9fa', fg='red').grid(row=0, column=0, columnspan=2, pady=10)
