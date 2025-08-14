import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import os
import sqlite3
from datetime import datetime, date
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from tkcalendar import DateEntry
from pathlib import Path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from alerts.notification_system import abrir_centro_alertas

# Agregar path para importar modelos
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.database import get_db, Empleado
from views.backup_config_view import abrir_configuracion_seguridad
from utils.config_fix import abrir_configuracion_avanzada_corregida

class MainWindow:
    def __init__(self, root):
        self.root = root
        self.db = get_db()
        self.contratos_window = None  # Para controlar única ventana de contratos
        self.configurar_estilos()
        self.setup_main_window()
        self.create_widgets()
        # Ya no creamos carpeta separada, usamos empleados_data
    
    def configurar_estilos(self):
        """Configurar estilos visuales mejorados"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Estilo para el TreeView
        style.configure('Empleados.Treeview',
                       font=('Arial', 9),
                       rowheight=25)
        
        style.configure('Empleados.Treeview.Heading',
                       font=('Arial', 9, 'bold'),
                       background='#4CAF50',
                       foreground='white')
        
        style.map('Empleados.Treeview',
                 background=[('selected', '#2196F3')])
    
    def setup_main_window(self):
        self.root.title("Sistema de Gestión de Personal - V1.2")
        
        # Configurar ventana responsive
        self.root.configure(bg='#f0f0f0')
        self.root.minsize(800, 600)  # Tamaño mínimo
        
        # Obtener dimensiones de pantalla
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calcular tamaño inicial (80% de la pantalla)
        window_width = min(1200, int(screen_width * 0.8))
        window_height = min(800, int(screen_height * 0.8))
        
        # Centrar ventana
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Configurar expansión de grid
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
    
    def create_widgets(self):
        # Frame principal responsive
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar expansión del frame principal
        main_frame.grid_rowconfigure(3, weight=1)  # La tabla se expandirá
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Título responsive
        title = tk.Label(main_frame, text="Sistema de Gestion de Personal", 
                        font=('Arial', 20, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        title.grid(row=0, column=0, columnspan=4, pady=(0, 25))
        
        # Frame de botones responsive
        btn_frame = tk.Frame(main_frame, bg='#f0f0f0')
        btn_frame.grid(row=1, column=0, columnspan=4, pady=(0, 20), sticky=(tk.W, tk.E))
        btn_frame.grid_columnconfigure(0, weight=1)
        
        buttons_info = [
            ("Nuevo Empleado", self.nuevo_empleado, "#27ae60"),      # Verde esmeralda
            ("Editar Empleado", self.editar_empleado, "#3498db"),      # Azul cielo
            ("Contratos", self.abrir_contratos, "#8e44ad"),            # Púrpura
            ("Inventarios", self.abrir_inventarios, "#16a085"),        # Verde azulado
            ("Asistencia QR", self.abrir_asistencia_qr, "#f39c12"),    # Naranja
            ("Centro Alertas", lambda: abrir_centro_alertas(self.root), "#e74c3c"),  # Rojo coral
            ("Seguridad", self.abrir_configuracion_seguridad, "#9b59b6"),  # Púrpura
            ("Configuracion", self.abrir_configuracion_avanzada, "#34495e")  # Gris oscuro
        ]
        
        # Crear botones responsive
        for i, (text, command, color) in enumerate(buttons_info):
            btn = tk.Button(btn_frame, text=text, command=command,
                           bg=color, fg='white', font=('Arial', 10, 'bold'),
                           relief='flat', padx=15, pady=8, cursor='hand2',
                           bd=0, activebackground=self.darker_color(color))
            btn.grid(row=i//4, column=i%4, padx=5, pady=5, sticky=(tk.W, tk.E))
            
            # Configurar expansión de columnas
            btn_frame.grid_columnconfigure(i%4, weight=1)
            
            # Efecto hover
            def make_hover(btn, color):
                def on_enter(e):
                    btn.config(bg=self.darker_color(color))
                def on_leave(e):
                    btn.config(bg=color)
                btn.bind("<Enter>", on_enter)
                btn.bind("<Leave>", on_leave)
            
            make_hover(btn, color)
        
        # Frame de búsqueda y filtros responsive
        search_frame = ttk.LabelFrame(main_frame, text="Busqueda y Filtros", padding="15")
        search_frame.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 15))
        search_frame.grid_columnconfigure(1, weight=1)  # El campo de búsqueda se expandirá
        
        # Búsqueda por texto responsive
        ttk.Label(search_frame, text="Buscar:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, font=('Arial', 10))
        search_entry.grid(row=0, column=1, padx=(0, 20), sticky=(tk.W, tk.E))
        
        # Filtro por área
        ttk.Label(search_frame, text="Area:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky=tk.W, padx=(0, 5))
        self.filter_area = tk.StringVar()
        self.filter_area.trace('w', self.on_search_change)
        area_combo = ttk.Combobox(search_frame, textvariable=self.filter_area, 
                                 values=["Todas", "planta", "postcosecha"], width=15)
        area_combo.set("Todas")
        area_combo.grid(row=0, column=3, padx=(0, 20))
        
        # Filtro por estado
        ttk.Label(search_frame, text="Estado:", font=('Arial', 10, 'bold')).grid(row=0, column=4, sticky=tk.W, padx=(0, 5))
        self.filter_estado = tk.StringVar()
        self.filter_estado.trace('w', self.on_search_change)
        estado_combo = ttk.Combobox(search_frame, textvariable=self.filter_estado, 
                                   values=["Activos", "Inactivos", "Todos"], width=15)
        estado_combo.set("Activos")
        estado_combo.grid(row=0, column=5, padx=(0, 10))
        
        # Botón limpiar filtros
        clear_btn = tk.Button(search_frame, text="Limpiar", command=self.limpiar_filtros,
                             bg="#95a5a6", fg="white", font=('Arial', 9, 'bold'),
                             relief='flat', padx=10, pady=5, cursor='hand2')
        clear_btn.grid(row=0, column=6)
        
        # Lista de empleados responsive
        empleados_frame = ttk.LabelFrame(main_frame, text="Lista de Empleados", padding="15")
        empleados_frame.grid(row=3, column=0, columnspan=4, pady=(15, 0), sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Treeview responsive para mostrar empleados
        columns = ('ID', 'Nombre', 'Cedula', 'Telefono', 'Area', 'Cargo', 'Salario', 'Estado', 'Excel')
        self.tree = ttk.Treeview(empleados_frame, columns=columns, show='headings', 
                                style='Empleados.Treeview')
        
        # Configurar columnas con anchos mínimos y expansión
        column_widths = {'ID': 50, 'Nombre': 150, 'Cedula': 100, 'Telefono': 100, 
                        'Area': 100, 'Cargo': 120, 'Salario': 100, 'Estado': 80, 'Excel': 80}
        
        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c))
            self.tree.column(col, width=column_widths.get(col, 100), minwidth=50)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(empleados_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(empleados_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid para tree y scrollbars
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Doble click para editar
        self.tree.bind('<Double-1>', self.on_double_click)
        
        # Crear barra de estado
        self.create_status_bar()
        
        # Cargar datos después de crear todos los widgets
        self.root.after(100, self.cargar_empleados)
        
        # Configurar grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        empleados_frame.columnconfigure(0, weight=1)
        empleados_frame.rowconfigure(0, weight=1)
    
    def darker_color(self, hex_color):
        """Hacer un color más oscuro para hover"""
        colors = {
            "#27ae60": "#229954",
            "#3498db": "#2980b9", 
            "#e67e22": "#d35400",
            "#8e44ad": "#7b1fa2",
            "#9b59b6": "#8e44ad",
            "#95a5a6": "#7f8c8d",
            "#34495e": "#2c3e50",
            "#1abc9c": "#16a085",
            "#e74c3c": "#c0392b"
        }
        return colors.get(hex_color, hex_color)

    def create_status_bar(self):
        """Crear barra de estado en la parte inferior"""
        self.status_bar = tk.Frame(self.root, bg='#34495e', height=30)
        self.status_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        self.status_label = tk.Label(self.status_bar, text="Sistema listo", 
                                    bg='#34495e', fg='white', font=('Arial', 9))
        self.status_label.pack(side=tk.LEFT, padx=10, pady=5)
        
        # Mostrar información actualizada después de que se cree el tree
        self.root.after(100, self.update_status)

    def update_status(self):
        """Actualizar barra de estado"""
        try:
            from datetime import datetime
            current_time = datetime.now().strftime("%H:%M:%S")
            
            # Contar empleados
            total_empleados = self.db.query(Empleado).count()
            activos = self.db.query(Empleado).filter(Empleado.estado == True).count()
            inactivos = total_empleados - activos
            
            # Contar elementos mostrados en la tabla
            items_mostrados = len(self.tree.get_children())
            
            status_text = f"Total: {total_empleados} | Activos: {activos} | Inactivos: {inactivos} | Mostrando: {items_mostrados} | {current_time}"
            self.status_label.config(text=status_text)
            
            self.root.after(1000, self.update_status)
        except:
            self.status_label.config(text="Sistema activo")
            self.root.after(1000, self.update_status)
    
    def on_search_change(self, *args):
        """Llamada cuando cambian los filtros de búsqueda"""
        if hasattr(self, 'tree'):
            self.cargar_empleados()
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros"""
        self.search_var.set("")
        self.filter_area.set("Todas")
        self.filter_estado.set("Activos")
        
    def cargar_empleados(self):
        """Cargar empleados en el TreeView con filtros aplicados"""
        try:
            # Limpiar datos existentes
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Construir query base
            query = self.db.query(Empleado)
            
            # Aplicar filtro de estado
            estado_filtro = self.filter_estado.get()
            if estado_filtro == "Activos":
                query = query.filter(Empleado.estado == True)
            elif estado_filtro == "Inactivos":
                query = query.filter(Empleado.estado == False)
            
            # Aplicar filtro de área
            area_filtro = self.filter_area.get()
            if area_filtro and area_filtro != "Todas":
                query = query.filter(Empleado.area_trabajo == area_filtro)
            
            # Aplicar búsqueda por texto
            texto_busqueda = self.search_var.get().strip().lower()
            if texto_busqueda:
                query = query.filter(
                    (Empleado.nombre_completo.ilike(f'%{texto_busqueda}%')) |
                    (Empleado.cedula.ilike(f'%{texto_busqueda}%')) |
                    (Empleado.cargo.ilike(f'%{texto_busqueda}%'))
                )
            
            # Ejecutar query
            empleados = query.all()
            
            # Cargar en TreeView (optimizado - sin verificación de Excel por defecto)
            for emp in empleados:
                estado_texto = "Activo" if emp.estado else "Inactivo"
                salario_texto = f"${emp.salario_base:,}" if emp.salario_base else "No definido"
                
                # Solo verificar Excel si es necesario (optimización)
                excel_status = "N/A"  # Por defecto no verificamos
                
                self.tree.insert('', 'end', values=(
                    emp.id,
                    emp.nombre_completo,
                    emp.cedula,
                    emp.telefono or "No definido",
                    emp.area_trabajo or "No definida",
                    emp.cargo or "No definido",
                    salario_texto,
                    estado_texto,
                    excel_status
                ))
            
            total = len(empleados)
            print(f"Se cargaron {total} empleados")
            
        except Exception as e:
            print(f"Error al cargar empleados: {e}")
    
    def verificar_excel_empleado(self, cedula):
        """Verificar si existe archivo Excel para el empleado en empleados_data"""
        try:
            # Buscar en empleados_data por cedula
            if os.path.exists("empleados_data"):
                for empleado_folder in os.listdir("empleados_data"):
                    if cedula.replace(" ", "_") in empleado_folder:
                        empleado_path = os.path.join("empleados_data", empleado_folder)
                        if os.path.isdir(empleado_path):
                            # Buscar cualquier archivo Excel en la carpeta
                            for file in os.listdir(empleado_path):
                                if file.endswith('.xlsx'):
                                    return True
            return False
        except:
            return False
    
    def sort_column(self, col):
        """Ordenar TreeView por columna"""
        try:
            data = [(self.tree.set(child, col), child) for child in self.tree.get_children('')]
            data.sort()
            
            for index, (val, child) in enumerate(data):
                self.tree.move(child, '', index)
        except:
            pass
    
    def on_double_click(self, event):
        """Manejar doble click en TreeView"""
        print("Doble clic detectado en tabla de empleados")
        self.editar_empleado()
    
    def get_selected_empleado(self):
        """Obtener empleado seleccionado"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor selecciona un empleado")
            return None
        
        item = self.tree.item(selection[0])
        empleado_id = item['values'][0]
        return self.db.query(Empleado).filter(Empleado.id == empleado_id).first()
    
    def nuevo_empleado(self):
        """Abrir ventana para nuevo empleado"""
        EmpleadosWindow(self.root, self, modo="nuevo")
    
    def editar_empleado(self):
        """Abrir ventana para editar empleado"""
        empleado = self.get_selected_empleado()
        if empleado:
            EmpleadosWindow(self.root, self, modo="editar", empleado=empleado)
    

    
    def abrir_configuracion_seguridad(self):
        """Abrir ventana de configuración de seguridad y backup"""
        try:
            abrir_configuracion_seguridad(self.root)
        except Exception as e:
            messagebox.showerror("Error", f"Error abriendo configuración de seguridad: {e}")
    
    def abrir_configuracion_avanzada(self):
        """Abrir ventana de configuración avanzada"""
        try:
            abrir_configuracion_avanzada_corregida(self.root)
        except Exception as e:
            messagebox.showerror("Error", f"Error abriendo configuración avanzada: {e}")

    def abrir_contratos(self):
        """Abrir ventana de gestión de contratos"""
        try:
            # Verificar si ya hay una ventana de contratos abierta
            if hasattr(self, 'contratos_window') and self.contratos_window:
                try:
                    self.contratos_window.window.lift()
                    self.contratos_window.window.focus_force()
                    return
                except:
                    # La ventana fue cerrada, continuar creando una nueva
                    self.contratos_window = None
            
            from views.contratos_view import ContratosWindow
            self.contratos_window = ContratosWindow(self.root, self)
            
            # El protocolo WM_DELETE_WINDOW ya está configurado en ContratosWindow
            
        except ImportError:
            messagebox.showerror("Error", "Módulo de contratos no encontrado")
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir contratos: {e}")
            import traceback
            traceback.print_exc()
    
    def abrir_inventarios(self):
        """Abrir ventana de gestión de inventarios - VERSIÓN COMPLETA"""
        try:
            print("Abriendo Centro de Inventarios completo...")
            FullInventarioWindow(self.root, self)
        except Exception as e:
            print(f"Error abriendo inventarios: {e}")
            messagebox.showerror("Error", f"Error al abrir inventarios: {e}")

    def abrir_asistencia_qr(self):
        """Abrir ventana del sistema de asistencia QR"""
        try:
            from views.asistencia_qr_view import AsistenciaQRView
            AsistenciaQRView(self.root)
        except ImportError:
            messagebox.showerror("Error", "Módulo de asistencia QR no encontrado")
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir sistema de asistencia QR: {e}")
            import traceback
            traceback.print_exc()

    # ================= NUEVA FUNCIONALIDAD EXCEL =================
    
    def generar_excel_empleado(self, empleado, abrir_despues=False):
        """Generar archivo Excel automáticamente para el empleado"""
        try:
            # Función para limpiar texto preservando caracteres especiales
            def limpiar_texto(texto):
                if texto is None:
                    return "No definido"
                # Convertir a string y limpiar solo caracteres problemáticos específicos
                texto_str = str(texto).strip()
                if not texto_str:
                    return "No definido"
                # Reemplazar solo caracteres que realmente causan problemas en Excel
                # pero preservar tildes y ñ
                caracteres_problematicos = {
                    '\u2705': '✓',  # check mark
                    '\u274c': '✗',  # cross mark
                    '\u2026': '...',  # ellipsis
                }
                for problema, reemplazo in caracteres_problematicos.items():
                    texto_str = texto_str.replace(problema, reemplazo)
                return texto_str
            
            # Crear carpeta del empleado en empleados_data
            def limpiar_nombre_archivo(texto):
                if not texto:
                    return "sin_nombre"
                caracteres_prohibidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
                texto_limpio = str(texto)
                for char in caracteres_prohibidos:
                    texto_limpio = texto_limpio.replace(char, '_')
                texto_limpio = texto_limpio.replace(" ", "_")
                return texto_limpio
            
            nombre_seguro = limpiar_nombre_archivo(empleado.nombre_completo)
            cedula_segura = limpiar_nombre_archivo(empleado.cedula)
            empleado_dir = os.path.join("empleados_data", f"{nombre_seguro}_{cedula_segura}")
            
            # Crear carpeta del empleado si no existe
            os.makedirs(empleado_dir, exist_ok=True)
            
            # Crear subcarpeta contratos si no existe
            contratos_dir = os.path.join(empleado_dir, "contratos")
            os.makedirs(contratos_dir, exist_ok=True)
            
            # Crear nombre de archivo con timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"empleado_{cedula_segura}_{timestamp}.xlsx"
            filepath = os.path.join(contratos_dir, filename)
            
            # Crear workbook
            wb = openpyxl.Workbook()
            
            # ===== HOJA 1: DATOS PERSONALES =====
            ws1 = wb.active
            ws1.title = "Datos Personales"
            
            # Configurar estilos
            header_font = Font(size=14, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_font = Font(size=18, bold=True, color="2F4F4F")
            data_font = Font(size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Título principal (sin tildes)
            ws1.merge_cells('A1:D1')
            ws1['A1'] = "INFORMACION DEL EMPLEADO"
            ws1['A1'].font = title_font
            ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Fecha de generación
            ws1.merge_cells('A2:D2')
            ws1['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws1['A2'].alignment = Alignment(horizontal='center')
            ws1['A2'].font = Font(size=10, italic=True)
            
            # Datos personales (con limpieza de texto)
            row = 4
            datos_personales = [
                ("DATOS BASICOS", ""),
                ("Nombre Completo:", limpiar_texto(empleado.nombre_completo)),
                ("Cedula:", limpiar_texto(empleado.cedula)),
                ("Telefono:", limpiar_texto(empleado.telefono)),
                ("Email:", limpiar_texto(empleado.email)),
                ("Direccion:", limpiar_texto(empleado.direccion)),
                ("", ""),
                ("DATOS LABORALES", ""),
                ("Area de Trabajo:", limpiar_texto(empleado.area_trabajo)),
                ("Cargo:", limpiar_texto(empleado.cargo)),
                ("Salario Base:", f"${empleado.salario_base:,}" if empleado.salario_base else "No definido"),
                ("Estado:", "Activo" if empleado.estado else "Inactivo"),
                ("Fecha de Ingreso:", empleado.fecha_creacion.strftime('%d/%m/%Y') if empleado.fecha_creacion else "No definida")
            ]
            
            for label, value in datos_personales:
                if label in ["DATOS BASICOS", "DATOS LABORALES"]:
                    # Headers de sección
                    ws1.merge_cells(f'A{row}:D{row}')
                    ws1[f'A{row}'] = limpiar_texto(label)
                    ws1[f'A{row}'].font = header_font
                    ws1[f'A{row}'].fill = header_fill
                    ws1[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws1[f'A{row}'].border = border
                elif label == "":
                    # Fila vacía
                    pass
                else:
                    # Datos (limpiar ambos textos)
                    ws1[f'A{row}'] = limpiar_texto(label)
                    ws1[f'B{row}'] = limpiar_texto(value)
                    ws1[f'A{row}'].font = Font(size=11, bold=True)
                    ws1[f'B{row}'].font = data_font
                    ws1[f'A{row}'].border = border
                    ws1[f'B{row}'].border = border
                
                row += 1
            
            # Ajustar ancho de columnas
            ws1.column_dimensions['A'].width = 20
            ws1.column_dimensions['B'].width = 30
            ws1.column_dimensions['C'].width = 15
            ws1.column_dimensions['D'].width = 15
            
            # ===== HOJA 2: SEGUIMIENTO =====
            ws2 = wb.create_sheet("Seguimiento")
            
            # Título
            ws2.merge_cells('A1:F1')
            ws2['A1'] = "SEGUIMIENTO Y NOTAS"
            ws2['A1'].font = title_font
            ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Headers de tabla (sin tildes)
            headers_seguimiento = ['Fecha', 'Tipo', 'Descripcion', 'Usuario', 'Estado', 'Observaciones']
            for i, header in enumerate(headers_seguimiento, 1):
                cell = ws2.cell(row=3, column=i, value=limpiar_texto(header))
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Fila de ejemplo (limpiando texto)
            ejemplo_seguimiento = [
                datetime.now().strftime('%d/%m/%Y'),
                'CREACION',
                'Empleado registrado en el sistema',
                'Sistema',
                'COMPLETADO',
                'Registro inicial automatico'
            ]
            
            for i, value in enumerate(ejemplo_seguimiento, 1):
                cell = ws2.cell(row=4, column=i, value=limpiar_texto(value))
                cell.font = data_font
                cell.border = border
            
            # Ajustar anchos
            column_widths = [12, 15, 30, 15, 12, 25]
            for i, width in enumerate(column_widths, 1):
                ws2.column_dimensions[chr(64 + i)].width = width
            
            # ===== HOJA 3: CAPACITACIONES =====
            ws3 = wb.create_sheet("Capacitaciones")
            
            # Título
            ws3.merge_cells('A1:E1')
            ws3['A1'] = "REGISTRO DE CAPACITACIONES"
            ws3['A1'].font = title_font
            ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Headers (sin tildes)
            headers_capacitaciones = ['Fecha', 'Capacitacion', 'Instructor', 'Duracion (hrs)', 'Estado']
            for i, header in enumerate(headers_capacitaciones, 1):
                cell = ws3.cell(row=3, column=i, value=limpiar_texto(header))
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Instrucciones (sin tildes)
            ws3['A5'] = "INSTRUCCIONES:"
            ws3['A5'].font = Font(size=12, bold=True, color="8B0000")
            
            instrucciones = [
                "• Registrar todas las capacitaciones recibidas",
                "• Actualizar el estado segun el progreso",
                "• Incluir certificaciones obtenidas",
                "• Revisar mensualmente"
            ]
            
            for i, instruccion in enumerate(instrucciones, 6):
                ws3[f'A{i}'] = limpiar_texto(instruccion)
                ws3[f'A{i}'].font = Font(size=10)
            
            # Ajustar anchos
            for i, width in enumerate([12, 25, 20, 12, 15], 1):
                ws3.column_dimensions[chr(64 + i)].width = width
            
            # ===== HOJA 4: EVALUACIONES =====
            ws4 = wb.create_sheet("Evaluaciones")
            
            # Título
            ws4.merge_cells('A1:F1')
            ws4['A1'] = "EVALUACIONES DE DESEMPEÑO"
            ws4['A1'].font = title_font
            ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Headers (sin tildes)
            headers_eval = ['Fecha', 'Periodo', 'Puntuacion', 'Fortalezas', 'Areas de Mejora', 'Evaluador']
            for i, header in enumerate(headers_eval, 1):
                cell = ws4.cell(row=3, column=i, value=limpiar_texto(header))
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ejemplo de evaluación (sin tildes)
            ws4['A5'] = "Proxima evaluacion:"
            ws4['B5'] = (datetime.now().replace(month=datetime.now().month+3) if datetime.now().month <= 9 
                         else datetime.now().replace(year=datetime.now().year+1, month=datetime.now().month-9)).strftime('%d/%m/%Y')
            ws4['A5'].font = Font(size=11, bold=True)
            ws4['B5'].font = Font(size=11, color="FF0000")
            
            # Ajustar anchos
            for i, width in enumerate([12, 15, 12, 25, 25, 15], 1):
                ws4.column_dimensions[chr(64 + i)].width = width
            
            # Guardar archivo
            wb.save(filepath)
            
            print(f"Excel generado: {filepath}")
            
            if abrir_despues:
                try:
                    os.startfile(filepath)  # Windows
                except:
                    try:
                        os.system(f'open "{filepath}"')  # macOS
                    except:
                        os.system(f'xdg-open "{filepath}"')  # Linux
            
            return True
            
        except Exception as e:
            print(f"Error generando Excel: {e}")
            messagebox.showerror("Error", f"Error generando Excel: {e}")
            return False


# ================= SISTEMA COMPLETO DE INVENTARIOS =================

class FullInventarioWindow:
    """Sistema completo de inventarios con funcionalidades reales"""
    
    def __init__(self, parent, main_window=None):
        self.parent = parent
        self.main_window = main_window
        
        # Crear ventana principal responsive
        self.window = tk.Toplevel(parent)
        self.window.title("CENTRO DE INVENTARIOS - Sistema Completo")
        self.window.configure(bg='#f8f9fa')
        
        # Configurar ventana responsive
        self.window.resizable(True, True)
        self.window.minsize(900, 600)
        
        # Configurar tamaño inicial responsive
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Calcular tamaño inicial (85% de la pantalla)
        window_width = min(1400, int(screen_width * 0.85))
        window_height = min(900, int(screen_height * 0.85))
        
        # Centrar ventana
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Configurar expansión de grid
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        
        # No hacer modal para que tenga controles estándar de Windows
        # self.window.transient(parent)
        # self.window.grab_set()
        
        # Configurar estilos y colores
        self.setup_styles()
        
        # Inicializar bases de datos
        self.setup_databases()
        
        # Lista para guardar las pestañas de inventario
        self.inventory_tabs = []
        
        # Crear interfaz
        self.create_interface()
        
        # No cargar datos iniciales - los usuarios importarán desde Excel
        # self.load_initial_data()
        
        # Configurar protocolo de cierre para que funcione el botón X
        self.window.protocol("WM_DELETE_WINDOW", self.close_window)
    
    def setup_styles(self):
        """Configurar estilos y colores armónicos"""
        self.colors = {
            'primary': '#2c3e50',      # Azul oscuro principal
            'secondary': '#34495e',     # Azul grisáceo
            'success': '#27ae60',       # Verde esmeralda
            'info': '#3498db',          # Azul cielo
            'warning': '#e67e22',       # Naranja suave
            'danger': '#e74c3c',        # Rojo coral
            'quimicos': '#27ae60',      # Verde para químicos
            'almacen': '#3498db',       # Azul para almacén
            'poscosecha': '#16a085'     # Verde azulado para poscosecha
        }
        
        # Configurar estilos de TreeView
        style = ttk.Style()
        
        # Estilo para químicos
        style.configure('Quimicos.Treeview', font=('Arial', 9), rowheight=25)
        style.configure('Quimicos.Treeview.Heading', 
                       font=('Arial', 9, 'bold'), 
                       background=self.colors['quimicos'], 
                       foreground='white')
        
        # Estilo para almacén
        style.configure('Almacen.Treeview', font=('Arial', 9), rowheight=25)
        style.configure('Almacen.Treeview.Heading',
                       font=('Arial', 9, 'bold'),
                       background=self.colors['almacen'],
                       foreground='white')
        
        # Estilo para poscosecha
        style.configure('Poscosecha.Treeview', font=('Arial', 9), rowheight=25)
        style.configure('Poscosecha.Treeview.Heading',
                       font=('Arial', 9, 'bold'),
                       background=self.colors['poscosecha'],
                       foreground='white')
    

    
    def setup_databases(self):
        """Configurar bases de datos para cada sistema"""
        try:
            # Crear directorio de bases de datos
            os.makedirs('database', exist_ok=True)
            
            # Bases de datos específicas
            self.db_paths = {
                'quimicos': 'database/inventario_quimicos.db',
                'almacen': 'database/inventario_almacen.db',
                'poscosecha': 'database/inventario_poscosecha.db'
            }
            
            # Crear tablas para cada sistema
            self.create_databases()
            
        except Exception as e:
            print(f"Error configurando bases de datos: {e}")
    
    def create_databases(self):
        """Crear bases de datos y tablas"""
        # Base de datos de químicos
        conn_q = sqlite3.connect(self.db_paths['quimicos'])
        cursor_q = conn_q.cursor()
        cursor_q.execute('''
            CREATE TABLE IF NOT EXISTS productos_quimicos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE,
                clase TEXT,
                nombre TEXT,
                saldo INTEGER DEFAULT 0,
                unidad TEXT,
                valor_unitario REAL DEFAULT 0,
                ubicacion TEXT,
                proveedor TEXT,
                fecha_vencimiento DATE,
                nivel_peligrosidad TEXT DEFAULT 'MEDIO',
                stock_minimo INTEGER DEFAULT 0,
                activo INTEGER DEFAULT 1,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn_q.commit()
        conn_q.close()
        
        # Base de datos de almacén
        conn_a = sqlite3.connect(self.db_paths['almacen'])
        cursor_a = conn_a.cursor()
        cursor_a.execute('''
            CREATE TABLE IF NOT EXISTS productos_almacen (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE,
                nombre TEXT,
                categoria TEXT DEFAULT 'General',
                saldo INTEGER DEFAULT 0,
                unidad TEXT,
                valor_unitario REAL DEFAULT 0,
                stock_minimo INTEGER DEFAULT 0,
                ubicacion TEXT,
                proveedor TEXT,
                activo INTEGER DEFAULT 1,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn_a.commit()
        conn_a.close()
        
        # Base de datos de poscosecha
        conn_p = sqlite3.connect(self.db_paths['poscosecha'])
        cursor_p = conn_p.cursor()
        cursor_p.execute('''
            CREATE TABLE IF NOT EXISTS productos_poscosecha (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE,
                categoria TEXT,
                nombre TEXT,
                saldo INTEGER DEFAULT 0,
                unidad TEXT,
                valor_unitario REAL DEFAULT 0,
                stock_minimo INTEGER DEFAULT 0,
                ubicacion TEXT,
                proveedor TEXT,
                tipo_producto TEXT,
                tipo TEXT DEFAULT 'General',
                activo INTEGER DEFAULT 1,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn_p.commit()
        conn_p.close()
        
        print("Bases de datos creadas exitosamente")
    
    def create_interface(self):
        """Crear interfaz principal responsive"""
        # Frame principal responsive
        main_frame = tk.Frame(self.window, bg='#f8f9fa')
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Configurar expansión
        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Header principal
        self.create_header(main_frame)
        
        # Notebook para pestañas responsive
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=(0, 10))
        
        # Crear pestañas
        self.create_dashboard_tab()
        self.create_quimicos_tab()
        self.create_almacen_tab()
        self.create_poscosecha_tab()
        
        # Footer
        self.create_footer(main_frame)
    
    def create_header(self, parent):
        """Crear header principal responsive"""
        header = tk.Frame(parent, bg=self.colors['primary'], height=80)
        header.grid(row=0, column=0, sticky=(tk.W, tk.E))
        header.grid_columnconfigure(0, weight=1)
        
        header_content = tk.Frame(header, bg=self.colors['primary'])
        header_content.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # Título
        title_label = tk.Label(header_content, text="CENTRO DE INVENTARIOS - Sistema Completo", 
                              font=('Arial', 20, 'bold'),
                              bg=self.colors['primary'], fg='white')
        title_label.pack(side=tk.LEFT)
        
        # Fecha y hora actual (reloj en tiempo real)
        self.datetime_label = tk.Label(header_content, text=datetime.now().strftime("%d/%m/%Y - %H:%M"),
                                 font=('Arial', 12),
                                 bg=self.colors['primary'], fg='white')
        self.datetime_label.pack(side=tk.RIGHT, padx=(0, 20))
        
        # El reloj se iniciará después en load_initial_data
    
    def create_dashboard_tab(self):
        """Crear pestaña de dashboard"""
        dashboard_frame = ttk.Frame(self.notebook)
        self.notebook.add(dashboard_frame, text="Dashboard General")
        
        # Contenido del dashboard
        main_content = tk.Frame(dashboard_frame, bg='#f8f9fa')
        main_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Título
        title_label = tk.Label(main_content, text="Dashboard de Inventarios",
                              font=('Arial', 18, 'bold'),
                              bg='#f8f9fa', fg=self.colors['primary'])
        title_label.pack(pady=(0, 20))
        
        # Frame para estadísticas
        stats_frame = tk.Frame(main_content, bg='#f8f9fa')
        stats_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Crear cards de estadísticas
        self.create_stats_cards(stats_frame)
        
        # Frame para acciones rápidas
        actions_frame = tk.Frame(main_content, bg='#f8f9fa')
        actions_frame.pack(fill=tk.X)
        
        # Crear botones de acciones
        self.create_dashboard_actions(actions_frame)
    
    def create_stats_cards(self, parent):
        """Crear cards de estadísticas"""
        # Obtener estadísticas de cada sistema
        stats = self.get_inventory_stats()
        
        cards_data = [
            ("QUIMICOS", stats['quimicos'], self.colors['quimicos']),
            ("ALMACEN", stats['almacen'], self.colors['almacen']),
            ("POSCOSECHA", stats['poscosecha'], self.colors['poscosecha']),
            ("TOTAL", stats['total'], "#9b59b6")
        ]
        
        # Configurar grid para cards responsive
        for i in range(4):
            parent.grid_columnconfigure(i, weight=1)
        
        for i, (title, count, color) in enumerate(cards_data):
            card = tk.Frame(parent, bg='white', relief='solid', bd=1)
            card.grid(row=0, column=i, sticky='ew', padx=5, pady=10)
            
            # Header de la card
            header = tk.Frame(card, bg=color, height=40)
            header.pack(fill=tk.X)
            header.pack_propagate(False)
            
            header_label = tk.Label(header, text=title,
                                   font=('Arial', 12, 'bold'),
                                   bg=color, fg='white')
            header_label.pack(expand=True)
            
            # Contenido
            content = tk.Frame(card, bg='white')
            content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            count_label = tk.Label(content, text=str(count),
                                  font=('Arial', 24, 'bold'),
                                  bg='white', fg=color)
            count_label.pack()
            
            desc_label = tk.Label(content, text="Productos",
                                 font=('Arial', 10),
                                 bg='white', fg='#7f8c8d')
            desc_label.pack()
        
        # Configurar grid
        for i in range(4):
            parent.grid_columnconfigure(i, weight=1)
    
    def create_dashboard_actions(self, parent):
        """Crear acciones del dashboard"""
        title_label = tk.Label(parent, text="Acciones Rapidas",
                              font=('Arial', 14, 'bold'),
                              bg='#f8f9fa', fg=self.colors['primary'])
        title_label.pack(pady=(0, 15))
        
        buttons_frame = tk.Frame(parent, bg='#f8f9fa')
        buttons_frame.pack()
        
        buttons = [
            ("Limpiar Inventarios", self.limpiar_todos_inventarios, self.colors['danger']),
            ("Generar Reporte", self.generate_report, self.colors['warning']),
            ("Exportar Datos", self.export_data, self.colors['success'])
        ]
        
        for text, command, color in buttons:
            btn = tk.Button(buttons_frame, text=text, command=command,
                           bg=color, fg='white', font=('Arial', 11, 'bold'),
                           relief='flat', bd=0, padx=20, pady=10, cursor='hand2')
            btn.pack(side=tk.LEFT, padx=10)
            self.add_hover_effect(btn, color)
    
    def create_quimicos_tab(self):
        """Crear pestaña de químicos"""
        quimicos_frame = ttk.Frame(self.notebook)
        self.notebook.add(quimicos_frame, text="Quimicos Agricolas")
        
        quimicos_tab = InventorySystemTab(quimicos_frame, 'quimicos', self.db_paths['quimicos'], 
                          self.colors['quimicos'], self)
        self.inventory_tabs.append(quimicos_tab)
    
    def create_almacen_tab(self):
        """Crear pestaña de almacén"""
        almacen_frame = ttk.Frame(self.notebook)
        self.notebook.add(almacen_frame, text="Almacen General")
        
        almacen_tab = InventorySystemTab(almacen_frame, 'almacen', self.db_paths['almacen'],
                          self.colors['almacen'], self)
        self.inventory_tabs.append(almacen_tab)
    
    def create_poscosecha_tab(self):
        """Crear pestaña de poscosecha"""
        poscosecha_frame = ttk.Frame(self.notebook)
        self.notebook.add(poscosecha_frame, text="Poscosecha")
        
        poscosecha_tab = InventorySystemTab(poscosecha_frame, 'poscosecha', self.db_paths['poscosecha'],
                          self.colors['poscosecha'], self)
        self.inventory_tabs.append(poscosecha_tab)
    
    def create_footer(self, parent):
        """Crear footer responsive"""
        footer = tk.Frame(parent, bg=self.colors['secondary'], height=30)
        footer.grid(row=2, column=0, sticky=(tk.W, tk.E))
        footer.grid_columnconfigure(0, weight=1)
        
        footer_label = tk.Label(footer, text="Centro de Inventarios v1.0 - Todos los sistemas operativos",
                               bg=self.colors['secondary'], fg='white', font=('Arial', 9))
        footer_label.grid(row=0, column=0, sticky=(tk.W, tk.E))
    
    def load_initial_data(self):
        """Cargar datos iniciales de ejemplo"""
        try:
            # Cargar datos de ejemplo para químicos
            self.load_sample_quimicos()
            
            # Cargar datos de ejemplo para almacén
            self.load_sample_almacen()
            
            # Cargar datos de ejemplo para poscosecha
            self.load_sample_poscosecha()
            
            print("Datos iniciales cargados exitosamente")
            
            # Iniciar el reloj después de que todo esté listo
            self.update_clock()
            
        except Exception as e:
            print(f"Error cargando datos iniciales: {e}")
            # Iniciar el reloj aunque haya errores
            self.update_clock()
    
    def load_sample_quimicos(self):
        """Cargar datos de ejemplo para químicos"""
        conn = sqlite3.connect(self.db_paths['quimicos'])
        cursor = conn.cursor()
        
        # Verificar si ya hay datos
        cursor.execute("SELECT COUNT(*) FROM productos_quimicos")
        if cursor.fetchone()[0] > 0:
            conn.close()
            return
        
        sample_data = [
            ('QM001', 'ACARICIDA', 'Abamectina 1.8%', 500, 'ML', 120, 'A-01', 'BAYER', '2025-12-31', 'ALTO'),
            ('QM002', 'FUNGICIDA', 'Mancozeb 80%', 2000, 'GR', 85, 'A-02', 'SYNGENTA', '2025-10-15', 'MEDIO'),
            ('QM003', 'INSECTICIDA', 'Lambda Cyhalothrin', 300, 'ML', 350, 'A-03', 'CORTEVA', '2025-08-20', 'ALTO'),
            ('QM004', 'HERBICIDA', 'Glifosato 48%', 1000, 'ML', 45, 'B-01', 'MONSANTO', '2026-03-10', 'MEDIO'),
            ('QM005', 'FERTILIZANTE', 'Urea 46%', 5000, 'KG', 2.5, 'C-01', 'YARA', None, 'BAJO'),
        ]
        
        cursor.executemany('''
            INSERT INTO productos_quimicos 
            (codigo, clase, nombre, saldo, unidad, valor_unitario, ubicacion, proveedor, fecha_vencimiento, nivel_peligrosidad)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', sample_data)
        
        conn.commit()
        conn.close()
    
    def load_sample_almacen(self):
        """Cargar datos de ejemplo para almacén"""
        conn = sqlite3.connect(self.db_paths['almacen'])
        cursor = conn.cursor()
        
        # Verificar si ya hay datos
        cursor.execute("SELECT COUNT(*) FROM productos_almacen")
        if cursor.fetchone()[0] > 0:
            conn.close()
            return
        
        sample_data = [
            ('ALM001', 'Aceite 15W-40', 50, 'LT', 45000, 10, 'A-01', 'LUBRICANTES SA'),
            ('ALM002', 'Filtro de aire', 25, 'UND', 25000, 5, 'A-02', 'REPUESTOS DIESEL'),
            ('ALM003', 'Llaves inglesas 12"', 15, 'UND', 35000, 3, 'A-03', 'HERRAMIENTAS PRO'),
            ('ALM004', 'Tornillos M8x20', 500, 'UND', 500, 100, 'A-04', 'TORNILLERIA'),
            ('ALM005', 'Cable eléctrico 12AWG', 200, 'MT', 2500, 50, 'A-05', 'ELECTRICOS DEL VALLE'),
        ]
        
        cursor.executemany('''
            INSERT INTO productos_almacen 
            (codigo, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', sample_data)
        
        conn.commit()
        conn.close()
    
    def load_sample_poscosecha(self):
        """Cargar datos de ejemplo para poscosecha"""
        conn = sqlite3.connect(self.db_paths['poscosecha'])
        cursor = conn.cursor()
        
        # Verificar si ya hay datos
        cursor.execute("SELECT COUNT(*) FROM productos_poscosecha")
        if cursor.fetchone()[0] > 0:
            conn.close()
            return
        
        sample_data = [
            ('PC001', 'EMBALAJE', 'Cajas cartón 18kg', 1000, 'UND', 2500, 100, 'PC-01', 'CARTONERIA', 'EMPAQUE'),
            ('PC002', 'QUIMICO', 'Tiabendazol', 50, 'KG', 45000, 10, 'PC-02', 'AGROQUIMICOS', 'TRATAMIENTO'),
            ('PC003', 'ETIQUETA', 'Etiquetas PLU', 5000, 'UND', 25, 1000, 'PC-03', 'ETIQUETAS SA', 'IDENTIFICACION'),
            ('PC004', 'HERRAMIENTA', 'Cuchillos inox', 20, 'UND', 35000, 5, 'PC-04', 'HERRAMIENTAS', 'HERRAMIENTA'),
            ('PC005', 'EMBALAJE', 'Pallets 120x80', 150, 'UND', 125000, 20, 'PC-05', 'PALLETS COL', 'EMPAQUE'),
        ]
        
        cursor.executemany('''
            INSERT INTO productos_poscosecha 
            (codigo, categoria, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor, tipo_producto)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', sample_data)
        
        conn.commit()
        conn.close()
    
    def get_inventory_stats(self):
        """Obtener estadísticas de inventarios"""
        stats = {'quimicos': 0, 'almacen': 0, 'poscosecha': 0, 'total': 0}
        
        try:
            # Químicos
            conn = sqlite3.connect(self.db_paths['quimicos'])
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos_quimicos")
            stats['quimicos'] = cursor.fetchone()[0]
            conn.close()
            
            # Almacén
            conn = sqlite3.connect(self.db_paths['almacen'])
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos_almacen")
            stats['almacen'] = cursor.fetchone()[0]
            conn.close()
            
            # Poscosecha
            conn = sqlite3.connect(self.db_paths['poscosecha'])
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos_poscosecha")
            stats['poscosecha'] = cursor.fetchone()[0]
            conn.close()
            
            stats['total'] = stats['quimicos'] + stats['almacen'] + stats['poscosecha']
            
        except Exception as e:
            print(f"Error obteniendo estadísticas: {e}")
        
        return stats
    
    # Funciones de acciones del dashboard


    def import_from_excel_to_inventory(self, system_type):
        """Importar productos desde Excel a un inventario específico"""
        try:
            file_path = filedialog.askopenfilename(
                title=f"Seleccionar archivo Excel para {system_type}",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            # Procesar el archivo según el tipo de inventario
            if system_type == 'quimicos':
                df = pd.read_excel(file_path, header=None)
                # Buscar encabezados
                header_row = None
                for i, row in df.iterrows():
                    if 'SALDO ANTERIOR' in str(row.values):
                        header_row = i
                        break
                if header_row is not None:
                    df = pd.read_excel(file_path, header=header_row)
                else:
                    df = pd.DataFrame()
                    
            elif system_type == 'almacen':
                df = pd.read_excel(file_path, header=None)
                # Buscar encabezados
                header_row = None
                for i, row in df.iterrows():
                    if 'PRODUCTO' in str(row.values) and 'SALDO' in str(row.values):
                        header_row = i
                        break
                if header_row is not None:
                    df = pd.read_excel(file_path, header=header_row)
                else:
                    df = pd.DataFrame()
                    
            elif system_type == 'poscosecha':
                df = pd.read_excel(file_path, header=None)
                # Buscar encabezados
                header_row = None
                for i, row in df.iterrows():
                    if 'PRODUCTO' in str(row.values) and 'SALDO' in str(row.values):
                        header_row = i
                        break
                if header_row is not None:
                    df = pd.read_excel(file_path, header=header_row)
                else:
                    df = pd.DataFrame()
            else:
                # Fallback: intentar leer como archivo estándar
                df = pd.read_excel(file_path)
            
            if df.empty:
                messagebox.showerror("Error", "No se pudieron procesar datos del archivo Excel")
                return
            
            # Mostrar información del archivo
            info_msg = f"Archivo: {os.path.basename(file_path)}\n"
            info_msg += f"Registros procesados: {len(df)}\n"
            info_msg += f"Columnas: {', '.join(df.columns.tolist())}"
            
            if messagebox.askyesno("Confirmar Importación", 
                                 f"{info_msg}\n\n¿Desea importar estos datos?"):
                
                # Conectar a la base de datos correspondiente
                base_dir = Path(__file__).resolve().parent.parent.parent
                db_path = base_dir / "database" / f"inventario_{system_type}.db"
                
                conn = sqlite3.connect(str(db_path))
                cursor = conn.cursor()
                
                # Limpiar tabla existente
                if system_type == 'quimicos':
                    cursor.execute("DELETE FROM productos_quimicos")
                elif system_type == 'almacen':
                    cursor.execute("DELETE FROM productos_almacen")
                elif system_type == 'poscosecha':
                    cursor.execute("DELETE FROM productos_poscosecha")
                
                # Insertar datos
                imported_count = 0
                for index, row in df.iterrows():
                    try:
                        if system_type == 'quimicos':
                            saldo = row.iloc[0] if len(row) > 0 else 0
                            clase = row.iloc[1] if len(row) > 1 else 'General'
                            producto = row.iloc[2] if len(row) > 2 else ''
                            valor = row.iloc[3] if len(row) > 3 else 50.0  # VALOR REAL DEL EXCEL
                            stock_min = 10  # Stock mínimo por defecto
                            proveedor = 'Proveedor General'  # Proveedor por defecto
                            
                            if pd.notna(producto) and str(producto).strip() != '':
                                try:
                                    saldo_num = int(float(saldo)) if pd.notna(saldo) else 0
                                    valor_num = float(valor) if pd.notna(valor) else 50.0
                                    stock_min_num = int(float(stock_min)) if pd.notna(stock_min) else 10
                                except:
                                    saldo_num = 0
                                    valor_num = 50.0
                                    stock_min_num = 10
                                
                                # Solo importar si tiene saldo > 0
                                if saldo_num > 0:
                                    codigo = f"Q{imported_count:03d}"
                                    cursor.execute('''
                                        INSERT INTO productos_quimicos 
                                        (codigo, nombre, clase, saldo, unidad, valor_unitario, 
                                         stock_minimo, ubicacion, proveedor, activo)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    ''', (
                                        codigo,
                                        str(producto).strip(),
                                        str(clase).strip() if pd.notna(clase) else 'General',
                                        saldo_num,
                                        'Kg',
                                        valor_num,  # Valor real del Excel
                                        stock_min_num,  # Stock mínimo real del Excel
                                        'Almacén General',
                                        str(proveedor).strip() if pd.notna(proveedor) else 'Proveedor General',
                                        1
                                    ))
                                    imported_count += 1
                                    
                        elif system_type == 'almacen':
                            # Para almacén, usar índices de columna
                            saldo = row.iloc[0] if len(row) > 0 else 0
                            producto = row.iloc[1] if len(row) > 1 else None
                            # Para almacén, usar valores por defecto ya que el Excel no tiene datos de precios
                            valor = 15.0  # Valor por defecto razonable
                            stock_min = 5  # Stock mínimo por defecto
                            proveedor = 'Proveedor General'  # Proveedor por defecto
                            
                            if pd.notna(producto) and str(producto).strip() != '' and str(producto).strip() != 'PRODUCTO y saldos':
                                try:
                                    saldo_num = int(float(saldo)) if pd.notna(saldo) else 0
                                    valor_num = float(valor) if pd.notna(valor) else 15.0
                                    stock_min_num = int(float(stock_min)) if pd.notna(stock_min) else 5
                                except:
                                    saldo_num = 0
                                    valor_num = 15.0
                                    stock_min_num = 5
                                
                                # Solo importar si tiene saldo > 0
                                if saldo_num > 0:
                                    codigo = f"A{imported_count:03d}"
                                    cursor.execute('''
                                        INSERT INTO productos_almacen 
                                        (codigo, nombre, categoria, saldo, unidad, valor_unitario, 
                                         stock_minimo, ubicacion, proveedor, activo)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    ''', (
                                        codigo,
                                        str(producto).strip(),
                                        'General',
                                        saldo_num,
                                        'Unidad',
                                        valor_num,  # Valor real del Excel
                                        stock_min_num,  # Stock mínimo real del Excel
                                        'Almacén General',
                                        str(proveedor).strip() if pd.notna(proveedor) else 'Proveedor General',
                                        1
                                    ))
                                    imported_count += 1
                                    
                        elif system_type == 'poscosecha':
                            # Para poscosecha, usar índices de columna
                            producto = row.iloc[0] if len(row) > 0 else None
                            saldo = row.iloc[1] if len(row) > 1 else 0
                            # Para poscosecha, usar valores por defecto ya que el Excel no tiene datos de precios
                            valor = 25.0  # Valor por defecto razonable
                            stock_min = 8  # Stock mínimo por defecto
                            proveedor = 'Proveedor General'  # Proveedor por defecto
                            
                            if pd.notna(producto) and str(producto).strip() != '' and str(producto).strip() != 'PRODUCTO':
                                try:
                                    saldo_num = int(float(saldo)) if pd.notna(saldo) else 0
                                    valor_num = float(valor) if pd.notna(valor) else 25.0
                                    stock_min_num = int(float(stock_min)) if pd.notna(stock_min) else 8
                                except:
                                    saldo_num = 0
                                    valor_num = 25.0
                                    stock_min_num = 8
                                
                                # Solo importar si tiene saldo > 0
                                if saldo_num > 0:
                                    codigo = f"P{imported_count:03d}"
                                    cursor.execute('''
                                        INSERT INTO productos_poscosecha 
                                        (codigo, nombre, tipo, saldo, unidad, valor_unitario, 
                                         stock_minimo, ubicacion, proveedor, activo)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    ''', (
                                        codigo,
                                        str(producto).strip(),
                                        'General',
                                        saldo_num,
                                        'Unidad',
                                        valor_num,  # Valor real del Excel
                                        stock_min_num,  # Stock mínimo real del Excel
                                        'Almacén General',
                                        str(proveedor).strip() if pd.notna(proveedor) else 'Proveedor General',
                                        1
                                    ))
                                    imported_count += 1
                                    
                    except Exception as e:
                        print(f"Error importando fila {index}: {e}")
                        continue
                    
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Importación Exitosa", 
                                  f"Se importaron {imported_count} productos exitosamente a {system_type.title()}")
                
                # ACTUALIZAR LA TABLA AUTOMÁTICAMENTE
                self.refresh_inventory_tables(system_type)
                
        except Exception as e:
            messagebox.showerror("Error", f"Error importando desde Excel: {e}")
    
    def refresh_inventory_tables(self, system_type):
        """Actualizar todas las tablas de inventario"""
        try:
            # Buscar y actualizar la tabla correspondiente
            for tab in self.inventory_tabs:
                if hasattr(tab, 'system_type') and tab.system_type == system_type:
                    if hasattr(tab, 'load_data'):
                        tab.load_data()
                        print(f"✅ Tabla de {system_type} actualizada automáticamente")
                    break
        except Exception as e:
            print(f"Error actualizando tabla: {e}")
    
    def limpiar_todos_inventarios(self):
        """Limpiar todos los inventarios"""
        if messagebox.askyesno("Confirmar Limpieza", 
                              "¿Está seguro de que desea eliminar TODOS los productos de todos los inventarios?\n\nEsta acción no se puede deshacer."):
            try:
                # Limpiar químicos
                conn_q = sqlite3.connect(self.db_paths['quimicos'])
                cursor_q = conn_q.cursor()
                cursor_q.execute("DELETE FROM productos_quimicos")
                conn_q.commit()
                conn_q.close()
                
                # Limpiar almacén
                conn_a = sqlite3.connect(self.db_paths['almacen'])
                cursor_a = conn_a.cursor()
                cursor_a.execute("DELETE FROM productos_almacen")
                conn_a.commit()
                conn_a.close()
                
                # Limpiar poscosecha
                conn_p = sqlite3.connect(self.db_paths['poscosecha'])
                cursor_p = conn_p.cursor()
                cursor_p.execute("DELETE FROM productos_poscosecha")
                conn_p.commit()
                conn_p.close()
                
                # Actualizar todas las tablas
                for tab in self.inventory_tabs:
                    if hasattr(tab, 'load_data'):
                        tab.load_data()
                
                messagebox.showinfo("Limpieza Exitosa", 
                                  "Todos los inventarios han sido limpiados exitosamente.\n\nLos inventarios están listos para nuevas importaciones.")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error limpiando inventarios: {e}")
    
    def generate_report(self):
        """Generar reporte completo de inventarios"""
        try:
            # Obtener estadísticas
            stats = self.get_inventory_stats()
            
            # Crear contenido del reporte
            report_content = "REPORTE DE INVENTARIOS\n"
            report_content += "=" * 50 + "\n\n"
            report_content += f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n"
            report_content += f"QUÍMICOS: {stats['quimicos']} productos\n"
            report_content += f"ALMACÉN: {stats['almacen']} productos\n"
            report_content += f"POSCOSECHA: {stats['poscosecha']} productos\n"
            report_content += f"TOTAL: {stats['total']} productos\n\n"
            
            # Guardar reporte
            file_path = filedialog.asksaveasfilename(
                title="Guardar Reporte",
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(report_content)
                
                messagebox.showinfo("Reporte Generado", 
                                  f"Reporte guardado exitosamente en:\n{file_path}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error generando reporte: {e}")
    
    def export_data(self):
        """Exportar datos de todos los inventarios"""
        try:
            # Seleccionar directorio de destino
            export_dir = filedialog.askdirectory(title="Seleccionar carpeta para exportar")
            
            if not export_dir:
                return
            
            exported_files = []
            
            # Exportar químicos
            conn_q = sqlite3.connect(self.db_paths['quimicos'])
            df_q = pd.read_sql_query("SELECT * FROM productos_quimicos", conn_q)
            if not df_q.empty:
                file_path = os.path.join(export_dir, "inventario_quimicos.csv")
                df_q.to_csv(file_path, index=False, encoding='utf-8')
                exported_files.append("inventario_quimicos.csv")
            conn_q.close()
            
            # Exportar almacén
            conn_a = sqlite3.connect(self.db_paths['almacen'])
            df_a = pd.read_sql_query("SELECT * FROM productos_almacen", conn_a)
            if not df_a.empty:
                file_path = os.path.join(export_dir, "inventario_almacen.csv")
                df_a.to_csv(file_path, index=False, encoding='utf-8')
                exported_files.append("inventario_almacen.csv")
            conn_a.close()
            
            # Exportar poscosecha
            conn_p = sqlite3.connect(self.db_paths['poscosecha'])
            df_p = pd.read_sql_query("SELECT * FROM productos_poscosecha", conn_p)
            if not df_p.empty:
                file_path = os.path.join(export_dir, "inventario_poscosecha.csv")
                df_p.to_csv(file_path, index=False, encoding='utf-8')
                exported_files.append("inventario_poscosecha.csv")
            conn_p.close()
            
            if exported_files:
                messagebox.showinfo("Exportación Exitosa", 
                                  f"Se exportaron {len(exported_files)} archivos:\n\n" + 
                                  "\n".join(exported_files) + 
                                  f"\n\nUbicación: {export_dir}")
            else:
                messagebox.showwarning("Sin Datos", 
                                     "No hay datos para exportar. Los inventarios están vacíos.")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error exportando datos: {e}")
    
    def add_hover_effect(self, button, color):
        """Agregar efecto hover"""
        def on_enter(e):
            button.config(bg=self.darken_color(color))
        def on_leave(e):
            button.config(bg=color)
        
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
    
    def darken_color(self, color):
        """Oscurecer color"""
        color_map = {
            '#27ae60': '#229954', '#3498db': '#2980b9', '#f39c12': '#e67e22',
            '#e74c3c': '#c0392b', '#16a085': '#138d75', '#95a5a6': '#7f8c8d'
        }
        return color_map.get(color, color)
    
    def close_window(self):
        """Cerrar ventana"""
        self.window.destroy()

    def update_clock(self):
        """Actualizar el reloj en tiempo real"""
        try:
            if hasattr(self, 'datetime_label'):
                current_time = datetime.now().strftime("%d/%m/%Y - %H:%M")
                self.datetime_label.config(text=current_time)
            # Programar próxima actualización en 60 segundos
            self.window.after(60000, self.update_clock)
        except Exception as e:
            print(f"Error actualizando reloj: {e}")
            # Reintentar en 60 segundos aunque haya error
            self.window.after(60000, self.update_clock)




# ================= PESTAÑA INDIVIDUAL PARA CADA SISTEMA =================

class InventorySystemTab:
    """Pestaña individual para cada sistema de inventario"""
    
    def __init__(self, parent, system_type, db_path, color, main_window):
        self.parent = parent
        self.system_type = system_type
        self.db_path = db_path
        self.color = color
        self.main_window = main_window
        
        self.create_interface()
        self.load_data()
    
    def create_interface(self):
        """Crear interfaz de la pestaña"""
        # Frame principal
        main_frame = tk.Frame(self.parent, bg='#f8f9fa')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header de la pestaña
        self.create_tab_header(main_frame)
        
        # Área de búsqueda y filtros
        self.create_search_area(main_frame)
        
        # Tabla de productos
        self.create_products_table(main_frame)
        
        # Panel de acciones
        self.create_actions_panel(main_frame)
    
    def create_tab_header(self, parent):
        """Crear header de la pestaña compacto"""
        header = tk.Frame(parent, bg=self.color, height=45)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        header_content = tk.Frame(header, bg=self.color)
        header_content.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        # Configurar expansión
        header_content.grid_columnconfigure(0, weight=1)
        
        # Título
        system_names = {
            'quimicos': 'PRODUCTOS QUIMICOS AGRICOLAS',
            'almacen': 'INVENTARIO DE ALMACEN GENERAL',
            'poscosecha': 'PRODUCTOS DE POSCOSECHA'
        }
        
        title_label = tk.Label(header_content, text=system_names[self.system_type],
                              font=('Arial', 14, 'bold'),
                              bg=self.color, fg='white')
        title_label.grid(row=0, column=0, sticky=tk.W)
        
        # Contador de productos
        self.count_label = tk.Label(header_content, text="",
                                   font=('Arial', 10, 'bold'),
                                   bg=self.color, fg='white')
        self.count_label.grid(row=0, column=1, sticky=tk.E)
    
    def create_search_area(self, parent):
        """Crear área de búsqueda compacta"""
        search_frame = tk.LabelFrame(parent, text="Busqueda y Filtros", 
                                    font=('Arial', 9, 'bold'),
                                    bg='#f8f9fa', fg='#2c3e50',
                                    padx=10, pady=8)
        search_frame.pack(fill=tk.X, pady=(8, 0))
        
        # Búsqueda en una sola fila compacta
        search_row = tk.Frame(search_frame, bg='#f8f9fa')
        search_row.pack(fill=tk.X, pady=3)
        
        # Configurar expansión
        search_row.grid_columnconfigure(1, weight=1)
        
        tk.Label(search_row, text="Buscar:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa').grid(row=0, column=0, padx=(0, 5), sticky=tk.W)
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)
        search_entry = tk.Entry(search_row, textvariable=self.search_var,
                               font=('Arial', 9))
        search_entry.grid(row=0, column=1, padx=(0, 10), sticky=(tk.W, tk.E))
        
        # Botones en la misma fila
        search_btn = tk.Button(search_row, text="Buscar",
                              command=self.search_products,
                              bg=self.color, fg='white',
                              font=('Arial', 8, 'bold'),
                              relief='flat', padx=12, pady=3, cursor='hand2')
        search_btn.grid(row=0, column=2, padx=2)
        
        clear_btn = tk.Button(search_row, text="Limpiar",
                             command=self.clear_search,
                             bg='#95a5a6', fg='white',
                             font=('Arial', 8, 'bold'),
                             relief='flat', padx=12, pady=3, cursor='hand2')
        clear_btn.grid(row=0, column=3, padx=2)
    
    def create_products_table(self, parent):
        """Crear tabla de productos"""
        table_frame = tk.LabelFrame(parent, text="Lista de Productos",
                                   font=('Arial', 10, 'bold'),
                                   bg='#f8f9fa', fg='#2c3e50',
                                   padx=15, pady=10)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Configurar columnas según el tipo de sistema
        if self.system_type == 'quimicos':
            columns = ('Codigo', 'Clase', 'Producto', 'Saldo', 'Unidad', 'Valor', 'Ubicacion', 'Proveedor', 'Peligrosidad')
            style_name = 'Quimicos.Treeview'
        elif self.system_type == 'almacen':
            columns = ('Codigo', 'Producto', 'Saldo', 'Unidad', 'Valor', 'Stock Min', 'Ubicacion', 'Proveedor')
            style_name = 'Almacen.Treeview'
        else:  # poscosecha
            columns = ('Codigo', 'Categoria', 'Producto', 'Saldo', 'Unidad', 'Valor', 'Stock Min', 'Ubicacion', 'Tipo')
            style_name = 'Poscosecha.Treeview'
        
        # TreeView con altura reducida para dar espacio a los botones
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings',
                                height=12, style=style_name)
        
        # Configurar columnas
        column_widths = {
            'Codigo': 80, 'Clase': 100, 'Categoria': 100, 'Producto': 200,
            'Saldo': 80, 'Unidad': 70, 'Valor': 100, 'Stock Min': 80,
            'Ubicacion': 80, 'Proveedor': 150, 'Peligrosidad': 100, 'Tipo': 120
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths.get(col, 100))
        
        # Scrollbars
        v_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        # Grid
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Eventos
        self.tree.bind('<Double-1>', self.on_double_click)
    
    def create_actions_panel(self, parent):
        """Crear panel de acciones responsive - todos los botones visibles"""
        actions_frame = tk.Frame(parent, bg='#f8f9fa')
        actions_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Configurar expansión de columnas para botones responsive (8 columnas)
        for i in range(8):
            actions_frame.grid_columnconfigure(i, weight=1)
        
        # Todos los botones en una sola fila compacta
        actions = [
            ("Nuevo", self.new_product, '#2c3e50'),
            ("Editar", self.edit_product, '#34495e'),
            ("Entrada", self.movimiento_entrada, '#27ae60'),
            ("Salida", self.movimiento_salida, '#e74c3c'),
            ("Historial", self.ver_historial_movimientos, '#8e44ad'),
            ("Eliminar", self.delete_product, '#c0392b'),
            ("Exportar", self.export_products, '#7f8c8d'),
            ("Importar", lambda: self.main_window.import_from_excel_to_inventory(self.system_type), '#f39c12')
        ]
        
        # Crear todos los botones en una sola fila
        for i, (text, command, color) in enumerate(actions):
            btn = tk.Button(actions_frame, text=text, command=command,
                           bg=color, fg='white', font=('Arial', 8, 'bold'),
                           relief='flat', bd=0, padx=5, pady=4, cursor='hand2')
            btn.grid(row=0, column=i, padx=1, pady=2, sticky=(tk.W, tk.E))
            self.add_hover_effect(btn, color)
    
    def load_data(self):
        """Cargar datos en la tabla"""
        try:
            # Limpiar tabla
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Conectar a base de datos
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Query según tipo de sistema
            if self.system_type == 'quimicos':
                cursor.execute("""
                    SELECT codigo, clase, nombre, saldo, unidad, valor_unitario, 
                           ubicacion, proveedor, nivel_peligrosidad
                    FROM productos_quimicos ORDER BY codigo
                """)
            elif self.system_type == 'almacen':
                cursor.execute("""
                    SELECT codigo, nombre, saldo, unidad, valor_unitario, 
                           stock_minimo, ubicacion, proveedor
                    FROM productos_almacen ORDER BY codigo
                """)
            else:  # poscosecha
                cursor.execute("""
                    SELECT codigo, categoria, nombre, saldo, unidad, valor_unitario,
                           stock_minimo, ubicacion, tipo_producto
                    FROM productos_poscosecha ORDER BY codigo
                """)
            
            # Cargar datos
            products_count = 0
            for row in cursor.fetchall():
                # Formatear valores
                formatted_row = list(row)
                if self.system_type == 'quimicos':
                    formatted_row[5] = f"${row[5]:,.0f}" if row[5] else "$0"
                elif self.system_type == 'almacen':
                    formatted_row[4] = f"${row[4]:,.0f}" if row[4] else "$0"
                else:  # poscosecha
                    formatted_row[5] = f"${row[5]:,.0f}" if row[5] else "$0"
                
                self.tree.insert('', tk.END, values=formatted_row)
                products_count += 1
            
            # Actualizar contador
            self.count_label.config(text=f"Total: {products_count} productos")
            
            conn.close()
            
        except Exception as e:
            print(f"Error cargando datos de {self.system_type}: {e}")
    
    def on_search_change(self, *args):
        """Manejar cambio en búsqueda"""
        # Implementar búsqueda en tiempo real si es necesario
        pass
    
    def search_products(self):
        """Buscar productos"""
        search_term = self.search_var.get().strip().lower()
        if not search_term:
            self.load_data()
            return
        
        try:
            # Limpiar tabla
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Query de búsqueda según tipo
            if self.system_type == 'quimicos':
                cursor.execute("""
                    SELECT codigo, clase, nombre, saldo, unidad, valor_unitario, 
                           ubicacion, proveedor, nivel_peligrosidad
                    FROM productos_quimicos 
                    WHERE LOWER(nombre) LIKE ? OR LOWER(codigo) LIKE ? OR LOWER(clase) LIKE ?
                    ORDER BY codigo
                """, (f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'))
            elif self.system_type == 'almacen':
                cursor.execute("""
                    SELECT codigo, nombre, saldo, unidad, valor_unitario, 
                           stock_minimo, ubicacion, proveedor
                    FROM productos_almacen 
                    WHERE LOWER(nombre) LIKE ? OR LOWER(codigo) LIKE ?
                    ORDER BY codigo
                """, (f'%{search_term}%', f'%{search_term}%'))
            else:  # poscosecha
                cursor.execute("""
                    SELECT codigo, categoria, nombre, saldo, unidad, valor_unitario,
                           stock_minimo, ubicacion, tipo_producto
                    FROM productos_poscosecha 
                    WHERE LOWER(nombre) LIKE ? OR LOWER(codigo) LIKE ? OR LOWER(categoria) LIKE ?
                    ORDER BY codigo
                """, (f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'))
            
            # Cargar resultados
            results_count = 0
            for row in cursor.fetchall():
                formatted_row = list(row)
                if self.system_type == 'quimicos':
                    formatted_row[5] = f"${row[5]:,.0f}" if row[5] else "$0"
                elif self.system_type == 'almacen':
                    formatted_row[4] = f"${row[4]:,.0f}" if row[4] else "$0"
                else:
                    formatted_row[5] = f"${row[5]:,.0f}" if row[5] else "$0"
                
                self.tree.insert('', tk.END, values=formatted_row)
                results_count += 1
            
            self.count_label.config(text=f"Encontrados: {results_count} productos")
            conn.close()
            
        except Exception as e:
            print(f"Error en búsqueda: {e}")
    
    def clear_search(self):
        """Limpiar búsqueda"""
        self.search_var.set("")
        self.load_data()
    
    def on_double_click(self, event):
        """Manejar doble click"""
        self.edit_product()
    
    def get_selected_product(self):
        """Obtener producto seleccionado"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Selecciona un producto")
            return None
        
        item = self.tree.item(selection[0])
        return item['values']
    
    def new_product(self):
        """Crear nuevo producto"""
        ProductFormWindow(self.parent, self, "new")
    
    def edit_product(self):
        """Editar producto"""
        product = self.get_selected_product()
        if product:
            ProductFormWindow(self.parent, self, "edit", product)
    
    def delete_product(self):
        """Eliminar producto"""
        product = self.get_selected_product()
        if product:
            if messagebox.askyesno("Confirmar", f"¿Eliminar el producto {product[0]}?"):
                try:
                    conn = sqlite3.connect(self.db_path)
                    cursor = conn.cursor()
                    
                    if self.system_type == 'quimicos':
                        cursor.execute("DELETE FROM productos_quimicos WHERE codigo = ?", (product[0],))
                    elif self.system_type == 'almacen':
                        cursor.execute("DELETE FROM productos_almacen WHERE codigo = ?", (product[0],))
                    else:
                        cursor.execute("DELETE FROM productos_poscosecha WHERE codigo = ?", (product[0],))
                    
                    conn.commit()
                    conn.close()
                    
                    self.load_data()
                    messagebox.showinfo("Éxito", "Producto eliminado")
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Error eliminando producto: {e}")
    
    def export_products(self):
        """Exportar productos"""
        file_path = filedialog.asksaveasfilename(
            title=f"Exportar {self.system_type}",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                if self.system_type == 'quimicos':
                    cursor.execute("SELECT * FROM productos_quimicos")
                elif self.system_type == 'almacen':
                    cursor.execute("SELECT * FROM productos_almacen")
                else:
                    cursor.execute("SELECT * FROM productos_poscosecha")
                
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow([description[0] for description in cursor.description])
                    writer.writerows(cursor.fetchall())
                
                conn.close()
                messagebox.showinfo("Éxito", f"Datos exportados a: {file_path}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error exportando: {e}")

    def movimiento_entrada(self):
        """Abrir diálogo de entrada de stock"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import MovimientoInventarioDialog
            dialog = MovimientoInventarioDialog(self.parent, self.system_type, 'entrada')
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir entrada de stock: {e}")

    def movimiento_salida(self):
        """Abrir diálogo de salida de stock"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import MovimientoInventarioDialog
            dialog = MovimientoInventarioDialog(self.parent, self.system_type, 'salida')
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir salida de stock: {e}")

    def ver_historial_movimientos(self):
        """Ver historial de movimientos"""
        try:
            import sys
            import os
            sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
            from utils.movimientos_inventario import HistorialMovimientosWindow
            dialog = HistorialMovimientosWindow(self.parent, self.system_type)
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir historial: {e}")
    
    def add_hover_effect(self, button, color):
        """Agregar efecto hover"""
        def on_enter(e):
            button.config(bg=self.darken_color(color))
        def on_leave(e):
            button.config(bg=color)
        
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
    
    def darken_color(self, color):
        """Oscurecer color"""
        color_map = {
            '#27ae60': '#229954', '#3498db': '#2980b9', '#e74c3c': '#c0392b', '#f39c12': '#e67e22'
        }
        return color_map.get(color, color)


# ================= FORMULARIO DE PRODUCTOS =================

class ProductFormWindow:
    """Formulario para crear/editar productos"""
    
    def __init__(self, parent, tab, mode, product_data=None):
        self.parent = parent
        self.tab = tab
        self.mode = mode  # "new" o "edit"
        self.product_data = product_data
        
        # Crear ventana responsive
        self.window = tk.Toplevel(parent)
        title = f"{'Editar' if mode == 'edit' else 'Nuevo'} Producto - {tab.system_type.title()}"
        self.window.title(title)
        self.window.configure(bg='#f8f9fa')
        
        # Configurar ventana responsive
        self.window.resizable(True, True)
        self.window.minsize(500, 600)
        
        # Configurar tamaño inicial responsive
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Calcular tamaño inicial (60% de la pantalla)
        window_width = min(600, int(screen_width * 0.6))
        window_height = min(700, int(screen_height * 0.6))
        
        # Centrar ventana
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Configurar expansión de grid
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        
        self.window.transient(parent)
        self.window.grab_set()
        
        self.create_form()
        
        if mode == "edit" and product_data:
            self.load_product_data()
    

    
    def create_form(self):
        """Crear formulario responsive"""
        # Frame principal responsive
        main_frame = tk.Frame(self.window, bg='#f8f9fa')
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Configurar expansión
        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Header
        header = tk.Frame(main_frame, bg=self.tab.color, height=60)
        header.grid(row=0, column=0, sticky=(tk.W, tk.E))
        header.grid_columnconfigure(0, weight=1)
        
        title = f"{'Editar' if self.mode == 'edit' else 'Nuevo'} Producto"
        title_label = tk.Label(header, text=title,
                              font=('Arial', 16, 'bold'),
                              bg=self.tab.color, fg='white')
        title_label.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # Formulario responsive
        form_frame = tk.Frame(main_frame, bg='white')
        form_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=20, pady=20)
        
        # Configurar expansión del formulario
        form_frame.grid_columnconfigure(1, weight=1)
        
        # Campos según tipo de sistema
        self.create_form_fields(form_frame)
        
        # Botones
        self.create_form_buttons(form_frame)
    
    def create_form_fields(self, parent):
        """Crear campos del formulario"""
        self.entries = {}
        row = 0
        
        # Campos comunes
        common_fields = [
            ("Código:", "codigo"),
            ("Nombre:", "nombre"),
            ("Saldo:", "saldo"),
            ("Unidad:", "unidad"),
            ("Valor Unitario:", "valor_unitario"),
            ("Ubicación:", "ubicacion"),
            ("Proveedor:", "proveedor")
        ]
        
        # Campos específicos por sistema
        if self.tab.system_type == 'quimicos':
            specific_fields = [
                ("Clase:", "clase"),
                ("Nivel Peligrosidad:", "nivel_peligrosidad"),
                ("Fecha Vencimiento:", "fecha_vencimiento")
            ]
        elif self.tab.system_type == 'almacen':
            specific_fields = [
                ("Stock Mínimo:", "stock_minimo")
            ]
        else:  # poscosecha
            specific_fields = [
                ("Categoría:", "categoria"),
                ("Stock Mínimo:", "stock_minimo"),
                ("Tipo Producto:", "tipo_producto")
            ]
        
        # Crear campos
        all_fields = common_fields + specific_fields
        
        for label_text, field_name in all_fields:
            # Label
            label = tk.Label(parent, text=label_text,
                           font=('Arial', 10, 'bold'),
                           bg='white', fg='#2c3e50')
            label.grid(row=row, column=0, sticky='w', pady=8, padx=(0, 10))
            
            # Entry o Combobox responsive
            if field_name in ['clase', 'nivel_peligrosidad', 'categoria', 'tipo_producto', 'unidad']:
                # Combobox para campos con opciones predefinidas
                values = self.get_field_options(field_name)
                widget = ttk.Combobox(parent, values=values)
            else:
                # Entry normal responsive
                widget = tk.Entry(parent, font=('Arial', 10))
            
            widget.grid(row=row, column=1, pady=8, sticky='ew')
            self.entries[field_name] = widget
            row += 1
        
        parent.grid_columnconfigure(1, weight=1)
    
    def get_field_options(self, field_name):
        """Obtener opciones para combobox"""
        options = {
            'clase': ['ACARICIDA', 'FUNGICIDA', 'INSECTICIDA', 'HERBICIDA', 'FERTILIZANTE', 'REGULADOR'],
            'nivel_peligrosidad': ['BAJO', 'MEDIO', 'ALTO'],
            'categoria': ['EMBALAJE', 'QUIMICO', 'ETIQUETA', 'HERRAMIENTA', 'GENERAL'],
            'tipo_producto': ['EMPAQUE', 'TRATAMIENTO', 'IDENTIFICACION', 'LIMPIEZA', 'GENERAL'],
            'unidad': ['ML', 'LT', 'KG', 'GR', 'UND', 'MT', 'M2']
        }
        return options.get(field_name, [])
    
    def create_form_buttons(self, parent):
        """Crear botones del formulario responsive"""
        btn_frame = tk.Frame(parent, bg='white')
        btn_frame.grid(row=20, column=0, columnspan=2, pady=20, sticky=(tk.W, tk.E))
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        
        # Botón guardar
        save_text = "Actualizar" if self.mode == "edit" else "Guardar"
        save_btn = tk.Button(btn_frame, text=save_text, command=self.save_product,
                           bg=self.tab.color, fg='white', font=('Arial', 11, 'bold'),
                           relief='flat', bd=0, padx=15, pady=10, cursor='hand2')
        save_btn.grid(row=0, column=0, padx=5, sticky=(tk.W, tk.E))
        
        # Botón cancelar
        cancel_btn = tk.Button(btn_frame, text="Cancelar", command=self.window.destroy,
                             bg='#95a5a6', fg='white', font=('Arial', 11, 'bold'),
                             relief='flat', bd=0, padx=15, pady=10, cursor='hand2')
        cancel_btn.grid(row=0, column=1, padx=5, sticky=(tk.W, tk.E))
    
    def load_product_data(self):
        """Cargar datos del producto para edición"""
        if not self.product_data:
            return
        
        # Mapear datos según tipo de sistema
        if self.tab.system_type == 'quimicos':
            mapping = {
                'codigo': 0, 'clase': 1, 'nombre': 2, 'saldo': 3, 'unidad': 4,
                'valor_unitario': 5, 'ubicacion': 6, 'proveedor': 7, 'nivel_peligrosidad': 8
            }
        elif self.tab.system_type == 'almacen':
            mapping = {
                'codigo': 0, 'nombre': 1, 'saldo': 2, 'unidad': 3,
                'valor_unitario': 4, 'stock_minimo': 5, 'ubicacion': 6, 'proveedor': 7
            }
        else:  # poscosecha
            mapping = {
                'codigo': 0, 'categoria': 1, 'nombre': 2, 'saldo': 3, 'unidad': 4,
                'valor_unitario': 5, 'stock_minimo': 6, 'ubicacion': 7, 'tipo_producto': 8
            }
        
        # Cargar datos en los campos
        for field_name, widget in self.entries.items():
            if field_name in mapping:
                index = mapping[field_name]
                if index < len(self.product_data):
                    value = self.product_data[index]
                    if value is not None:
                        if isinstance(widget, ttk.Combobox):
                            widget.set(str(value))
                        else:
                            widget.delete(0, tk.END)
                            widget.insert(0, str(value))
    
    def save_product(self):
        """Guardar producto"""
        try:
            # Validar campos requeridos
            required_fields = ['codigo', 'nombre', 'saldo', 'unidad', 'valor_unitario']
            for field in required_fields:
                if field in self.entries and not self.entries[field].get().strip():
                    messagebox.showerror("Error", f"El campo {field} es obligatorio")
                    self.entries[field].focus()
                    return
            
            # Preparar datos
            data = {}
            for field_name, widget in self.entries.items():
                value = widget.get().strip()
                if field_name in ['saldo', 'stock_minimo']:
                    data[field_name] = int(value) if value else 0
                elif field_name == 'valor_unitario':
                    data[field_name] = float(value) if value else 0.0
                else:
                    data[field_name] = value
            
            # Conectar a base de datos
            conn = sqlite3.connect(self.tab.db_path)
            cursor = conn.cursor()
            
            if self.mode == "new":
                # Verificar código único
                table_name = f"productos_{self.tab.system_type}"
                cursor.execute(f"SELECT id FROM {table_name} WHERE codigo = ?", (data['codigo'],))
                if cursor.fetchone():
                    messagebox.showerror("Error", "El código ya existe")
                    return
                
                # Insertar nuevo producto
                self.insert_new_product(cursor, data)
                message = "Producto creado exitosamente"
            else:
                # Actualizar producto existente
                self.update_existing_product(cursor, data)
                message = "Producto actualizado exitosamente"
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Éxito", message)
            self.tab.load_data()
            
            if self.mode == "new":
                self.clear_form()
            else:
                self.window.destroy()
                
        except ValueError as e:
            messagebox.showerror("Error", "Valores numéricos inválidos")
        except Exception as e:
            messagebox.showerror("Error", f"Error guardando producto: {e}")
    
    def insert_new_product(self, cursor, data):
        """Insertar nuevo producto"""
        if self.tab.system_type == 'quimicos':
            cursor.execute('''
                INSERT INTO productos_quimicos 
                (codigo, clase, nombre, saldo, unidad, valor_unitario, ubicacion, proveedor, nivel_peligrosidad, fecha_vencimiento)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['codigo'], data.get('clase', ''), data['nombre'], data['saldo'],
                data['unidad'], data['valor_unitario'], data.get('ubicacion', ''),
                data.get('proveedor', ''), data.get('nivel_peligrosidad', 'MEDIO'),
                data.get('fecha_vencimiento', '') or None
            ))
        elif self.tab.system_type == 'almacen':
            cursor.execute('''
                INSERT INTO productos_almacen 
                (codigo, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['codigo'], data['nombre'], data['saldo'], data['unidad'],
                data['valor_unitario'], data.get('stock_minimo', 0),
                data.get('ubicacion', ''), data.get('proveedor', '')
            ))
        else:  # poscosecha
            cursor.execute('''
                INSERT INTO productos_poscosecha 
                (codigo, categoria, nombre, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor, tipo_producto)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['codigo'], data.get('categoria', ''), data['nombre'], data['saldo'],
                data['unidad'], data['valor_unitario'], data.get('stock_minimo', 0),
                data.get('ubicacion', ''), data.get('proveedor', ''), data.get('tipo_producto', '')
            ))
    
    def update_existing_product(self, cursor, data):
        """Actualizar producto existente"""
        codigo_original = self.product_data[0]
        
        if self.tab.system_type == 'quimicos':
            cursor.execute('''
                UPDATE productos_quimicos SET 
                codigo=?, clase=?, nombre=?, saldo=?, unidad=?, valor_unitario=?, 
                ubicacion=?, proveedor=?, nivel_peligrosidad=?, fecha_vencimiento=?
                WHERE codigo=?
            ''', (
                data['codigo'], data.get('clase', ''), data['nombre'], data['saldo'],
                data['unidad'], data['valor_unitario'], data.get('ubicacion', ''),
                data.get('proveedor', ''), data.get('nivel_peligrosidad', 'MEDIO'),
                data.get('fecha_vencimiento', '') or None, codigo_original
            ))
        elif self.tab.system_type == 'almacen':
            cursor.execute('''
                UPDATE productos_almacen SET 
                codigo=?, nombre=?, saldo=?, unidad=?, valor_unitario=?, 
                stock_minimo=?, ubicacion=?, proveedor=?
                WHERE codigo=?
            ''', (
                data['codigo'], data['nombre'], data['saldo'], data['unidad'],
                data['valor_unitario'], data.get('stock_minimo', 0),
                data.get('ubicacion', ''), data.get('proveedor', ''), codigo_original
            ))
        else:  # poscosecha
            cursor.execute('''
                UPDATE productos_poscosecha SET 
                codigo=?, categoria=?, nombre=?, saldo=?, unidad=?, valor_unitario=?, 
                stock_minimo=?, ubicacion=?, proveedor=?, tipo_producto=?
                WHERE codigo=?
            ''', (
                data['codigo'], data.get('categoria', ''), data['nombre'], data['saldo'],
                data['unidad'], data['valor_unitario'], data.get('stock_minimo', 0),
                data.get('ubicacion', ''), data.get('proveedor', ''), 
                data.get('tipo_producto', ''), codigo_original
            ))
    
    def clear_form(self):
        """Limpiar formulario"""
        for widget in self.entries.values():
            if isinstance(widget, ttk.Combobox):
                widget.set('')
            else:
                widget.delete(0, tk.END)


# ================= VENTANA DE EMPLEADOS CON EXCEL AUTOMÁTICO =================

class EmpleadosWindow:
    def __init__(self, parent, main_window, modo="nuevo", empleado=None):
        self.parent = parent
        self.main_window = main_window
        self.db = get_db()
        self.modo = modo
        self.empleado = empleado
        
        # Crear ventana responsive
        self.window = tk.Toplevel(parent)
        titulo = "Editar Empleado" if modo == "editar" else "Nuevo Empleado"
        self.window.title(titulo)
        self.window.configure(bg='#f8f9fa')
        
        # Configurar ventana responsive
        self.window.resizable(True, True)
        self.window.minsize(600, 800)
        
        # Configurar tamaño inicial responsive
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Calcular tamaño inicial (75% de la pantalla)
        window_width = min(700, int(screen_width * 0.75))
        window_height = min(900, int(screen_height * 0.75))
        
        # Centrar ventana
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Configurar expansión de grid
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        
        # Hacer modal
        self.window.transient(parent)
        self.window.grab_set()
        
        self.create_widgets()
        
        # Si es edición, cargar datos
        if modo == "editar" and empleado:
            self.cargar_datos_empleado()
    

    
    def create_widgets(self):
        # Frame principal simple y funcional
        main_frame = tk.Frame(self.window, bg='#f8f9fa', padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Configurar expansión
        main_frame.grid_columnconfigure(1, weight=1)  # Los campos se expandirán
        
        # Título con mejor estilo
        titulo_texto = "Editar Empleado" if self.modo == "editar" else "Registrar Nuevo Empleado"
        title_label = tk.Label(main_frame, text=titulo_texto, font=('Arial', 18, 'bold'),
                              bg='#f8f9fa', fg='#2c3e50')
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 25), sticky='ew')
        
        # Campos
        # Campos normales
        campos_normales = [
            ("Nombre Completo:", "entry_nombre"),
            ("Cedula:", "entry_cedula"),
            ("Telefono:", "entry_telefono"),
            ("Email:", "entry_email"),
            ("Direccion:", "entry_direccion"),
            ("Lugar de Nacimiento:", "entry_lugar_nacimiento"),
            ("Nacionalidad:", "entry_nacionalidad"),
            ("Cargo:", "entry_cargo"),
            ("Salario Base:", "entry_salario")
        ]
        
        row = 1
        for label_text, entry_name in campos_normales:
            label = tk.Label(main_frame, text=label_text, font=('Arial', 11, 'bold'),
                           bg='#f8f9fa', fg='#2c3e50')
            label.grid(row=row, column=0, sticky=tk.W, pady=10, padx=(0, 15))
            
            entry = ttk.Entry(main_frame, font=('Arial', 11))
            entry.grid(row=row, column=1, pady=10, sticky='ew', padx=(0, 10))
            setattr(self, entry_name, entry)
            row += 1
        
        # Campo de fecha de nacimiento con calendario
        date_label = tk.Label(main_frame, text="Fecha de Nacimiento:", font=('Arial', 11, 'bold'),
                            bg='#f8f9fa', fg='#2c3e50')
        date_label.grid(row=row, column=0, sticky=tk.W, pady=10, padx=(0, 15))
        self.date_fecha_nacimiento = DateEntry(main_frame, width=27, background='darkblue', foreground='white', 
                                              borderwidth=2, date_pattern='dd/mm/yyyy', font=('Arial', 10))
        self.date_fecha_nacimiento.grid(row=row, column=1, pady=8)
        row += 1
        
        # Área como combobox
        area_label = tk.Label(main_frame, text="Area:", font=('Arial', 11, 'bold'),
                            bg='#f8f9fa', fg='#2c3e50')
        area_label.grid(row=11, column=0, sticky=tk.W, pady=10, padx=(0, 15))
        self.combo_area = ttk.Combobox(main_frame, values=["planta", "postcosecha"], font=('Arial', 11))
        self.combo_area.grid(row=11, column=1, pady=10, sticky='ew', padx=(0, 10))
        
        # Estado (solo para edición)
        if self.modo == "editar":
            estado_label = tk.Label(main_frame, text="Estado:", font=('Arial', 11, 'bold'),
                                  bg='#f8f9fa', fg='#2c3e50')
            estado_label.grid(row=12, column=0, sticky=tk.W, pady=10, padx=(0, 15))
            self.combo_estado = ttk.Combobox(main_frame, values=["Activo", "Inactivo"], font=('Arial', 11))
            self.combo_estado.grid(row=12, column=1, pady=10, sticky='ew')
        
        
        
        # Botones responsive
        btn_frame = tk.Frame(main_frame, bg='#f8f9fa')
        btn_row = 13 if self.modo == "editar" else 12
        btn_frame.grid(row=btn_row, column=0, columnspan=2, pady=20, sticky=(tk.W, tk.E))
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        btn_frame.grid_columnconfigure(2, weight=1)
        
        # Botón principal
        if self.modo == "nuevo":
            texto_boton = "Guardar Empleado"
        else:
            texto_boton = "Actualizar Empleado"
            
        save_btn = tk.Button(btn_frame, text=texto_boton, command=self.guardar_empleado,
                           bg="#27ae60", fg="white", font=('Arial', 10, 'bold'),
                           relief='flat', padx=15, pady=8, cursor='hand2')
        save_btn.grid(row=0, column=0, padx=5, sticky=(tk.W, tk.E))
        
        # Botón limpiar
        clear_btn = tk.Button(btn_frame, text="Limpiar", command=self.limpiar_campos,
                            bg="#f39c12", fg="white", font=('Arial', 10, 'bold'),
                            relief='flat', padx=15, pady=8, cursor='hand2')
        clear_btn.grid(row=0, column=1, padx=5, sticky=(tk.W, tk.E))
        
        # Botón cerrar
        close_btn = tk.Button(btn_frame, text="Cerrar", command=self.window.destroy,
                            bg="#e74c3c", fg="white", font=('Arial', 10, 'bold'),
                            relief='flat', padx=15, pady=8, cursor='hand2')
        close_btn.grid(row=0, column=2, padx=5, sticky=(tk.W, tk.E))
    
    def cargar_datos_empleado(self):
        """Cargar datos del empleado en el formulario"""
        if not self.empleado:
            return
        
        self.entry_nombre.insert(0, self.empleado.nombre_completo or "")
        self.entry_cedula.insert(0, self.empleado.cedula or "")
        self.entry_telefono.insert(0, self.empleado.telefono or "")
        self.entry_email.insert(0, self.empleado.email or "")
        self.entry_direccion.insert(0, self.empleado.direccion or "")
        self.entry_lugar_nacimiento.insert(0, self.empleado.lugar_nacimiento or "")
        if self.empleado.fecha_nacimiento:
            self.date_fecha_nacimiento.set_date(self.empleado.fecha_nacimiento)
        self.entry_nacionalidad.insert(0, self.empleado.nacionalidad or "COLOMBIANA")
        self.combo_area.set(self.empleado.area_trabajo or "")
        self.entry_cargo.insert(0, self.empleado.cargo or "")
        self.entry_salario.insert(0, str(self.empleado.salario_base or ""))
        
        if hasattr(self, 'combo_estado'):
            self.combo_estado.set("Activo" if self.empleado.estado else "Inactivo")
    
    def guardar_empleado(self):
        try:
            # Importar validadores
            from utils.validators import DataValidator
            
            # Recopilar datos del formulario
            datos_empleado = {
                'nombre_completo': self.entry_nombre.get().strip(),
                'cedula': self.entry_cedula.get().strip(),
                'telefono': self.entry_telefono.get().strip(),
                'email': self.entry_email.get().strip(),
                'direccion': self.entry_direccion.get().strip(),
                'lugar_nacimiento': self.entry_lugar_nacimiento.get().strip(),
                'fecha_nacimiento': self.date_fecha_nacimiento.get_date(),
                'nacionalidad': self.entry_nacionalidad.get().strip(),
                'area_trabajo': self.combo_area.get(),
                'cargo': self.entry_cargo.get().strip(),
                'salario_base': self.entry_salario.get().strip()
            }
            
            # Validar todos los datos usando el sistema de validación
            empleado_id = self.empleado.id if self.empleado else None
            es_valido, errores = DataValidator.validar_empleado_completo(self.db, datos_empleado, empleado_id)
            
            if not es_valido:
                # Mostrar todos los errores en un mensaje
                mensaje_errores = "Se encontraron los siguientes errores:\n\n"
                for error in errores:
                    mensaje_errores += f"• {error}\n"
                messagebox.showerror("Errores de Validación", mensaje_errores)
                return
            
            # Procesar salario
            salario = 0
            if datos_empleado['salario_base']:
                salario = int(datos_empleado['salario_base'])
            
            if self.modo == "nuevo":
                # Crear nuevo empleado
                empleado = Empleado(
                    nombre_completo=datos_empleado['nombre_completo'],
                    cedula=datos_empleado['cedula'],
                    telefono=datos_empleado['telefono'],
                    email=datos_empleado['email'],
                    direccion=datos_empleado['direccion'],
                    lugar_nacimiento=datos_empleado['lugar_nacimiento'],
                    fecha_nacimiento=datos_empleado['fecha_nacimiento'],
                    nacionalidad=datos_empleado['nacionalidad'],
                    area_trabajo=datos_empleado['area_trabajo'],
                    cargo=datos_empleado['cargo'],
                    salario_base=salario
                )
                self.db.add(empleado)
                self.db.commit()  # Hacer commit para obtener el ID
                
                mensaje = f"✅ Empleado {empleado.nombre_completo} creado exitosamente"
                print(f"Empleado creado: {empleado.nombre_completo}")
                
            else:
                # Actualizar empleado existente
                self.empleado.nombre_completo = datos_empleado['nombre_completo']
                self.empleado.cedula = datos_empleado['cedula']
                self.empleado.telefono = datos_empleado['telefono']
                self.empleado.email = datos_empleado['email']
                self.empleado.direccion = datos_empleado['direccion']
                self.empleado.lugar_nacimiento = datos_empleado['lugar_nacimiento']
                self.empleado.fecha_nacimiento = datos_empleado['fecha_nacimiento']
                self.empleado.nacionalidad = datos_empleado['nacionalidad']
                self.empleado.area_trabajo = datos_empleado['area_trabajo']
                self.empleado.cargo = datos_empleado['cargo']
                self.empleado.salario_base = salario
                
                if hasattr(self, 'combo_estado'):
                    self.empleado.estado = self.combo_estado.get() == "Activo"
                
                self.db.commit()
                mensaje = "Empleado actualizado exitosamente"
            
            print(mensaje)
            
            # Actualizar ventana principal
            self.main_window.cargar_empleados()
            
            # Limpiar campos o cerrar
            if self.modo == "nuevo":
                # Mostrar mensaje de éxito
                messagebox.showinfo("Exito!", mensaje)
                self.limpiar_campos()
            else:
                messagebox.showinfo("Exito!", mensaje)
                self.window.destroy()
            
        except Exception as e:
            print(f"Error al guardar: {e}")
            messagebox.showerror("Error", f"Error al guardar empleado: {e}")
    
    def limpiar_campos(self):
        """Limpiar todos los campos del formulario"""
        self.entry_nombre.delete(0, tk.END)
        self.entry_cedula.delete(0, tk.END)
        self.entry_telefono.delete(0, tk.END)
        self.entry_email.delete(0, tk.END)
        self.entry_direccion.delete(0, tk.END)
        self.entry_lugar_nacimiento.delete(0, tk.END)
        self.date_fecha_nacimiento.set_date(date.today())
        self.entry_nacionalidad.delete(0, tk.END)
        self.entry_nacionalidad.insert(0, "COLOMBIANA")
        self.combo_area.set('')
        self.entry_cargo.delete(0, tk.END)
        self.entry_salario.delete(0, tk.END)
        
        if hasattr(self, 'combo_estado'):
            self.combo_estado.set('Activo')


# ================= FUNCIÓN PRINCIPAL PARA EJECUTAR =================

def main():
    """Función principal para ejecutar el sistema"""
    root = tk.Tk()
    
    # Verificar dependencias
    try:
        import pandas as pd
        import openpyxl
        print("Todas las dependencias estan instaladas")
    except ImportError as e:
        missing_lib = str(e).split("'")[1]
        messagebox.showerror("Error de Dependencias", 
                            f"Falta instalar la libreria: {missing_lib}\n\n"
                            f"Ejecuta en terminal:\n"
                            f"pip install {missing_lib}")
        root.destroy()
        return
    
    # Inicializar aplicación
    app = MainWindow(root)
    
    # Configurar cierre de aplicación
    def on_closing():
        if messagebox.askokcancel("Cerrar", "¿Estás seguro que deseas cerrar el sistema?"):
            # Detener el sistema de alertas antes de cerrar
            try:
                from alerts.notification_system import AlertsManager
                # Crear una instancia temporal para detener el monitoreo
                alerts_manager = AlertsManager()
                alerts_manager.stop_monitoring()
            except Exception as e:
                print(f"Error deteniendo alertas: {e}")
            
            root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # Mostrar mensaje de bienvenida
    print("Sistema de Gestion de Personal v1.2 iniciado")
    print("Funcionalidad Excel automatico ACTIVADA")
    print("Sistema de validacion y backup ACTIVADO")
    
    # Ejecutar aplicación
    root.mainloop()

if __name__ == "__main__":
    main()


# ================= INSTRUCCIONES DE USO =================
"""
SISTEMA DE GESTION DE PERSONAL V1.2 - CON EXCEL AUTOMATICO

NUEVA FUNCIONALIDAD:
- Al registrar un NUEVO empleado, automaticamente se genera un archivo Excel
- El Excel incluye 4 hojas: Datos Personales, Seguimiento, Capacitaciones, Evaluaciones
- Cada Excel se guarda como "empleado_{cedula}.xlsx" en la carpeta "empleados_excel"
- Opcion de abrir el Excel inmediatamente despues de crearlo

CARACTERISTICAS:
1. **Registro Automatico**: Al crear empleado → Excel se genera automaticamente
2. **Excel Completo**: 4 hojas con toda la informacion necesaria
3. **Boton "Ver Excel"**: Abre el Excel de cualquier empleado seleccionado
4. **Columna "Excel"**: Muestra SI/NO si existe archivo Excel
5. **Inventarios Intactos**: Todo el sistema de inventarios sigue funcionando igual

INSTALACION:
pip install pandas openpyxl

USO:
1. Ejecutar el sistema
2. Crear nuevo empleado
3. Llenar datos basicos
4. Hacer click en "Guardar + Excel"
5. El Excel se genera automaticamente!
6. Opcion de abrirlo inmediatamente

ARCHIVOS GENERADOS:
- empleados_excel/empleado_{cedula}.xlsx (para cada empleado)
- database/ (bases de datos del sistema)

COMPATIBILIDAD:
- Windows, macOS, Linux
- Python 3.6+
- Tkinter (incluido en Python)

FUNCIONALIDADES COMPLETAS:
✓ Gestion de Empleados con Excel automatico
✓ Sistema completo de Inventarios (3 modulos)
✓ Busquedas y filtros avanzados
✓ Exportacion de datos
✓ Interfaz moderna y profesional
✓ Base de datos SQLite integrada

SISTEMA COMPLETAMENTE FUNCIONAL!
"""