# views/inventario_view.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import os
from datetime import datetime, date
import sqlite3
import csv

class ModernInventarioWindow:
    def __init__(self, parent, main_window=None):
        self.parent = parent
        self.main_window = main_window
        self.setup_database()
        
        # Crear ventana principal
        self.window = tk.Toplevel(parent)
        self.window.title("📦 Sistema de Inventarios")
        self.window.geometry("1200x800")
        self.window.configure(bg='#f0f0f0')
        
        # Configurar ventana
        self.window.resizable(True, True)
        self.window.minsize(1000, 600)
        
        # Centrar ventana
        self.center_window()
        
        # Hacer modal
        self.window.transient(parent)
        self.window.grab_set()
        
        # Configurar estilos
        self.configurar_estilos()
        
        # Crear interfaz
        self.create_interface()
        
        # Cargar datos automáticamente y mostrar mensaje si es exitoso
        total_cargados = self.cargar_inventario()
        
        # Mostrar información de carga una sola vez
        if total_cargados and total_cargados > 20:  # Si cargó productos desde Excel
            self.window.after(2000, lambda: self.mostrar_info_carga_exitosa(total_cargados))
    
    def configurar_estilos(self):
        """Configurar estilos minimalistas"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Estilo para el TreeView de inventario - más neutral
        style.configure('Inventario.Treeview',
                       background='#ffffff',
                       foreground='#2c3e50',
                       fieldbackground='#ffffff',
                       font=('Arial', 9),
                       rowheight=25)
        
        # Estilo para los headers - más sobrio
        style.configure('Inventario.Treeview.Heading',
                       background='#34495e',
                       foreground='white',
                       font=('Arial', 9, 'bold'),
                       relief='flat')
        
        # Efectos de selección
        style.map('Inventario.Treeview',
                 background=[('selected', '#bdc3c7')],
                 foreground=[('selected', '#2c3e50')])
    
    def center_window(self):
        """Centrar ventana en pantalla"""
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (1200 // 2)
        y = (self.window.winfo_screenheight() // 2) - (800 // 2)
        self.window.geometry(f"1200x800+{x}+{y}")
    
    def setup_database(self):
        """Configurar base de datos de inventarios"""
        try:
            # Crear carpeta database si no existe
            db_dir = 'database'
            os.makedirs(db_dir, exist_ok=True)
            
            db_path = os.path.join(db_dir, 'inventario.db')
            self.conn = sqlite3.connect(db_path)
            self.crear_tablas()
            self.poblar_datos_iniciales()
        except Exception as e:
            print(f"Error BD: {e}")
            # Fallback
            self.conn = sqlite3.connect('inventario.db')
            self.crear_tablas()
            self.poblar_datos_iniciales()
    
    def crear_tablas(self):
        """Crear tablas de inventario"""
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE NOT NULL,
                nombre TEXT NOT NULL,
                categoria TEXT NOT NULL,
                saldo INTEGER DEFAULT 0,
                unidad TEXT NOT NULL,
                valor_unitario REAL NOT NULL,
                stock_minimo INTEGER DEFAULT 0,
                ubicacion TEXT,
                proveedor TEXT,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS movimientos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                producto_id INTEGER,
                tipo TEXT NOT NULL,
                cantidad INTEGER NOT NULL,
                fecha DATE NOT NULL,
                factura TEXT,
                proveedor TEXT,
                destino TEXT,
                valor_total REAL,
                observaciones TEXT,
                fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (producto_id) REFERENCES productos (id)
            )
        ''')
        
        self.conn.commit()
    
    def poblar_datos_iniciales(self):
        """Cargar automáticamente TODOS los datos desde Excel al iniciar"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM productos")
        
        if cursor.fetchone()[0] == 0:
            print("🔄 Cargando datos desde archivos Excel...")
            
            # SIEMPRE intentar cargar desde Excel primero
            if self.cargar_datos_desde_excel():
                print("✅ Datos reales cargados exitosamente desde Excel")
                self.mostrar_resumen_carga()
                return
            
            # Solo usar datos de ejemplo si fallan completamente los Excel
            print("⚠️ No se pudieron cargar archivos Excel, usando datos de ejemplo...")
            productos_ejemplo = [
                ('ALM001', 'ACEITE 15W 40 AL', 'almacen', 3, 'GALONES', 45000, 2, 'A-01', 'LUBRICANTES SA'),
                ('ALM002', 'ACOPLE ESPIGO 3/4', 'almacen', 13, 'UND', 8500, 5, 'A-02', 'FERRETERIA CENTRAL'),
                ('ALM003', 'ACOPLES FOSTER 17$ BRONCE', 'almacen', 2, 'UND', 17000, 3, 'A-03', 'FOSTER LTDA'),
                ('PC001', 'TAPAS DE CUARTOS IMBLLOM', 'postcosecha', 0, 'UND', 3500, 5, 'PC-01', 'EMPAQUES DEL SUR'),
                ('PC002', 'OCTAVOS INBLLOM', 'postcosecha', 32, 'UND', 15000, 10, 'PC-02', 'PAPELERIA INDUSTRIAL'),
                ('QM001', '[ACARICIDA] Abafed', 'quimicos', 32000, 'C.C', 80, 5000, 'QM-01', 'AGROQUIMICOS ANDINOS'),
                ('QM002', '[ACARICIDA] ADN MILBE', 'quimicos', 0, 'C.C.', 120, 100, 'QM-02', 'PRODUCTOS QUIMICOS SA'),
                ('ALM005', 'ARANDELA CUADRADA 4*4', 'almacen', 0, 'UND', 500, 10, 'A-05', 'TORNILLERIA BOGOTA')
            ]
            
            cursor.executemany('''
                INSERT INTO productos (codigo, nombre, categoria, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', productos_ejemplo)
            
            self.conn.commit()
            print("📋 Datos de ejemplo cargados como respaldo")
    
    def mostrar_resumen_carga(self):
        """Mostrar resumen de datos cargados desde Excel"""
        cursor = self.conn.cursor()
        
        # Contar por categoría
        categorias = ['almacen', 'postcosecha', 'quimicos']
        resumen = {}
        total_general = 0
        
        for categoria in categorias:
            cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria = ?", (categoria,))
            count = cursor.fetchone()[0]
            resumen[categoria] = count
            total_general += count
        
        print(f"📊 RESUMEN DE CARGA DESDE EXCEL:")
        print(f"   🏭 ALMACÉN: {resumen['almacen']} productos")
        print(f"   🌿 POSTCOSECHA: {resumen['postcosecha']} productos") 
        print(f"   🧪 QUÍMICOS: {resumen['quimicos']} productos")
        print(f"   📦 TOTAL: {total_general} productos cargados")
        
        # Calcular valor total del inventario
        cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos")
        valor_total = cursor.fetchone()[0] or 0
        print(f"   💰 VALOR TOTAL: ${valor_total:,.0f}")
        
        return total_general
    
    def cargar_datos_desde_excel(self):
        """Cargar TODOS los datos desde los archivos Excel reales - MEJORADO"""
        try:
            archivos_excel = [
                ('INVENTARIO ALMACEN JUNIO  2025 xlsx.xlsx', 'almacen', '🏭'),
                ('INV QUIMICOS JUNIO 2025 1.xlsx', 'quimicos', '🧪'),
                ('SALDOS POSCOSECHA  JUNIO  2025.xlsx', 'postcosecha', '🌿')
            ]
            
            productos_cargados = 0
            cursor = self.conn.cursor()
            archivos_exitosos = 0
            
            print("🔍 Buscando archivos Excel...")
            
            for archivo, categoria, emoji in archivos_excel:
                if not os.path.exists(archivo):
                    print(f"❌ {emoji} {archivo} - NO ENCONTRADO")
                    continue
                    
                try:
                    print(f"📖 {emoji} Procesando {archivo}...")
                    productos = self.procesar_archivo_excel(archivo, categoria)
                    
                    if productos:
                        productos_insertados = 0
                        for producto in productos:
                            try:
                                # Verificar si ya existe
                                cursor.execute("SELECT id FROM productos WHERE codigo = ?", (producto[0],))
                                if cursor.fetchone():
                                    # Actualizar existente
                                    cursor.execute('''
                                        UPDATE productos SET 
                                        nombre=?, categoria=?, saldo=?, unidad=?, valor_unitario=?, 
                                        stock_minimo=?, ubicacion=?, proveedor=?
                                        WHERE codigo=?
                                    ''', producto[1:] + (producto[0],))
                                else:
                                    # Insertar nuevo
                                    cursor.execute('''
                                        INSERT INTO productos 
                                        (codigo, nombre, categoria, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    ''', producto)
                                
                                productos_insertados += 1
                                    
                            except Exception as e:
                                print(f"⚠️ Error con producto {producto[1]}: {e}")
                                continue
                        
                        productos_cargados += productos_insertados
                        archivos_exitosos += 1
                        print(f"✅ {emoji} {archivo}: {productos_insertados} productos cargados")
                        
                    else:
                        print(f"⚠️ {emoji} {archivo}: No se encontraron productos válidos")
                    
                except Exception as e:
                    print(f"❌ {emoji} Error procesando {archivo}: {e}")
                    continue
            
            if productos_cargados > 0:
                self.conn.commit()
                print(f"\n🎉 CARGA EXITOSA:")
                print(f"   📁 Archivos procesados: {archivos_exitosos}/3")
                print(f"   📦 Total productos: {productos_cargados}")
                
                return True
            else:
                print("❌ No se pudieron cargar productos desde Excel")
                return False
            
        except Exception as e:
            print(f"❌ Error general cargando Excel: {e}")
            return False
    
    def procesar_archivo_excel(self, archivo, categoria):
        """Procesar archivo Excel específico según su estructura"""
        if not os.path.exists(archivo):
            return []
        
        try:
            # Leer archivo usando openpyxl si está disponible, sino usar csv
            try:
                import openpyxl
                return self.procesar_con_openpyxl(archivo, categoria)
            except ImportError:
                print(f"openpyxl no disponible para {archivo}")
                return []
                
        except Exception as e:
            print(f"Error procesando {archivo}: {e}")
            return []
    
    def procesar_con_openpyxl(self, archivo, categoria):
        """Procesar Excel con openpyxl"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(archivo, data_only=True)
            
            # Seleccionar la primera hoja o hoja específica
            if categoria == 'almacen':
                sheet = workbook['Almacen'] if 'Almacen' in workbook.sheetnames else workbook.active
                return self.procesar_almacen(sheet)
            elif categoria == 'quimicos':
                sheet = workbook.active
                return self.procesar_quimicos(sheet)
            elif categoria == 'postcosecha':
                sheet = workbook.active
                return self.procesar_postcosecha(sheet)
                
        except Exception as e:
            print(f"Error con openpyxl en {archivo}: {e}")
            return []
    
    def procesar_almacen(self, sheet):
        """Procesar hoja de almacén - PROCESAMIENTO AGRESIVO PARA TODOS LOS PRODUCTOS"""
        productos = []
        contador = 1
        
        print("   📋 Procesando productos de ALMACÉN...")
        print(f"   📏 Total de filas en hoja: {sheet.max_row}")
        
        # Buscar AGRESIVAMENTE en todas las filas posibles
        productos_encontrados = 0
        for row_num in range(1, min(sheet.max_row + 1, 500)):  # Aumentar rango de búsqueda
            try:
                # Revisar múltiples columnas para encontrar productos
                posibles_nombres = []
                posibles_saldos = []
                posibles_unidades = []
                
                # Revisar las primeras 6 columnas buscando datos
                for col in range(1, 7):
                    cell_value = sheet.cell(row=row_num, column=col).value
                    if cell_value:
                        if isinstance(cell_value, str) and len(str(cell_value).strip()) > 3:
                            posibles_nombres.append((col, str(cell_value).strip()))
                        elif isinstance(cell_value, (int, float)) and cell_value >= 0:
                            posibles_saldos.append((col, cell_value))
                
                # Buscar productos válidos en esta fila
                for col_nombre, nombre in posibles_nombres:
                    # Filtrar nombres que claramente son productos
                    nombre_clean = nombre.upper()
                    if (len(nombre) > 3 and 
                        not any(skip in nombre_clean for skip in ['FECHA', 'FACTURA', 'PROVE', 'CANT', 'VALOR', 'TOTAL', 'DIA', 'SALDO MES', 'PRODUCTO Y SALDOS']) and
                        not nombre_clean.startswith('DIA ') and
                        nombre_clean not in ['FECHA', 'N. FACTURA', 'PROVE', 'CANT', 'VALOR UND', 'VALOR TOTAL']):
                        
                        # Buscar saldo en columnas adyacentes
                        saldo = 0
                        unidad = 'UND'
                        
                        # Buscar saldo en las siguientes columnas
                        for saldo_col in range(col_nombre + 1, min(col_nombre + 4, 7)):
                            saldo_val = sheet.cell(row=row_num, column=saldo_col).value
                            if isinstance(saldo_val, (int, float)) and saldo_val >= 0:
                                saldo = int(saldo_val)
                                
                                # Buscar unidad en la siguiente columna
                                unidad_val = sheet.cell(row=row_num, column=saldo_col + 1).value
                                if isinstance(unidad_val, str) and len(str(unidad_val).strip()) <= 10:
                                    unidad = str(unidad_val).strip().upper()
                                break
                        
                        # Crear el producto si parece válido
                        if nombre and len(nombre) > 2:
                            codigo = self.crear_codigo_automatico('almacen', contador)
                            valor_unitario = self.estimar_valor_unitario(nombre, categoria='almacen')
                            stock_minimo = max(1, int(saldo * 0.2)) if saldo > 0 else 5
                            
                            productos.append((
                                codigo, nombre, 'almacen', saldo, unidad,
                                valor_unitario, stock_minimo, f'ALM-{contador:02d}', 'PROVEEDOR ALMACEN'
                            ))
                            
                            contador += 1
                            productos_encontrados += 1
                            
                            # Debug para los primeros productos
                            if productos_encontrados <= 10:
                                print(f"     • {nombre} - {saldo} {unidad} (fila {row_num}, col {col_nombre})")
                            
                            break  # Solo un producto por fila
                    
            except Exception as e:
                continue
        
        print(f"   ✅ {len(productos)} productos de ALMACÉN procesados")
        return productos
    
    def procesar_quimicos(self, sheet):
        """Procesar hoja de químicos - PROCESAMIENTO AGRESIVO"""
        productos = []
        contador = 1
        
        print("   📋 Procesando productos de QUÍMICOS...")
        print(f"   📏 Total de filas en hoja: {sheet.max_row}")
        
        productos_encontrados = 0
        # Buscar en un rango amplio
        for row_num in range(1, min(sheet.max_row + 1, 400)):
            try:
                # Revisar múltiples patrones de columnas para químicos
                # Patrón 1: saldo_anterior, clase, nombre, saldo_real, unidad
                clase1 = sheet.cell(row=row_num, column=2).value
                nombre1 = sheet.cell(row=row_num, column=3).value
                saldo1 = sheet.cell(row=row_num, column=4).value
                unidad1 = sheet.cell(row=row_num, column=5).value
                
                # Patrón 2: buscar en otras columnas también
                nombre2 = sheet.cell(row=row_num, column=2).value
                saldo2 = sheet.cell(row=row_num, column=3).value
                
                # Evaluar patrón 1 (más probable)
                if (nombre1 and isinstance(nombre1, str) and len(str(nombre1).strip()) > 2):
                    nombre = str(nombre1).strip()
                    clase = str(clase1).strip() if clase1 else ''
                    
                    # Filtrar headers y texto no deseado
                    nombre_upper = nombre.upper()
                    if (nombre_upper not in ['PRODUCTO', 'CLASE', 'SALDO REAL', 'FECHA', 'FACTURA', 'PROVE', 'CANT', 'VALOR', 'TOTAL'] and
                        not nombre_upper.startswith('DIA ') and
                        len(nombre) > 2):
                        
                        # Combinar clase y nombre si hay clase válida
                        if (clase and len(clase) > 1 and 
                            clase.upper() not in ['CLASE', 'SALDO ANTERIOR', 'PRODUCTO']):
                            nombre_completo = f"[{clase}] {nombre}"
                        else:
                            nombre_completo = nombre
                        
                        # Obtener saldo
                        try:
                            saldo_real = int(saldo1) if saldo1 is not None else 0
                        except (ValueError, TypeError):
                            saldo_real = 0
                        
                        # Obtener unidad
                        if isinstance(unidad1, str) and len(str(unidad1).strip()) <= 10:
                            unidad = str(unidad1).strip()
                        else:
                            unidad = 'C.C'
                        
                        # Crear producto
                        codigo = self.crear_codigo_automatico('quimicos', contador)
                        valor_unitario = self.estimar_valor_unitario(nombre_completo, categoria='quimicos')
                        stock_minimo = max(100, int(saldo_real * 0.1)) if saldo_real > 1000 else 50
                        
                        productos.append((
                            codigo, nombre_completo, 'quimicos', saldo_real, unidad,
                            valor_unitario, stock_minimo, f'QM-{contador:02d}', 'PROVEEDOR QUIMICOS'
                        ))
                        
                        contador += 1
                        productos_encontrados += 1
                        
                        # Debug
                        if productos_encontrados <= 10:
                            print(f"     • {nombre_completo} - {saldo_real} {unidad} (fila {row_num})")
                
                # Evaluar patrón 2 si el patrón 1 no funcionó
                elif (nombre2 and isinstance(nombre2, str) and len(str(nombre2).strip()) > 2):
                    nombre = str(nombre2).strip()
                    nombre_upper = nombre.upper()
                    
                    if (nombre_upper not in ['PRODUCTO', 'CLASE', 'SALDO', 'FECHA'] and
                        len(nombre) > 2):
                        
                        try:
                            saldo_real = int(saldo2) if saldo2 is not None else 0
                        except (ValueError, TypeError):
                            saldo_real = 0
                        
                        codigo = self.crear_codigo_automatico('quimicos', contador)
                        valor_unitario = self.estimar_valor_unitario(nombre, categoria='quimicos')
                        stock_minimo = max(50, int(saldo_real * 0.1)) if saldo_real > 500 else 25
                        
                        productos.append((
                            codigo, nombre, 'quimicos', saldo_real, 'C.C',
                            valor_unitario, stock_minimo, f'QM-{contador:02d}', 'PROVEEDOR QUIMICOS'
                        ))
                        
                        contador += 1
                        productos_encontrados += 1
                        
            except Exception as e:
                continue
        
        print(f"   ✅ {len(productos)} productos QUÍMICOS procesados")
        return productos
    
    def procesar_postcosecha(self, sheet):
        """Procesar hoja de postcosecha - PROCESAMIENTO AGRESIVO"""
        productos = []
        contador = 1
        
        print("   📋 Procesando productos de POSTCOSECHA...")
        print(f"   📏 Total de filas en hoja: {sheet.max_row}")
        
        productos_encontrados = 0
        # Buscar en un rango amplio
        for row_num in range(1, min(sheet.max_row + 1, 300)):
            try:
                # Patrón típico: nombre, saldo, unidad
                nombre = sheet.cell(row=row_num, column=1).value
                saldo = sheet.cell(row=row_num, column=2).value
                unidad = sheet.cell(row=row_num, column=3).value
                
                # También revisar otras posiciones
                if not nombre:
                    nombre = sheet.cell(row=row_num, column=2).value
                    saldo = sheet.cell(row=row_num, column=3).value
                    unidad = sheet.cell(row=row_num, column=4).value
                
                if (nombre and isinstance(nombre, str) and len(str(nombre).strip()) > 2):
                    nombre_clean = str(nombre).strip()
                    nombre_upper = nombre_clean.upper()
                    
                    # Filtrar headers y texto no válido
                    if (nombre_upper not in ['PRODUCTO', 'SALDO', 'UND', 'UNIDAD', 'FECHA', 'FACTURA', 'PROVE', 'CANT', 'VALOR'] and
                        not nombre_upper.startswith('DIA ') and
                        len(nombre_clean) > 2 and
                        nombre_upper != 'ENTRADA' and
                        'SALIDAS' not in nombre_upper):
                        
                        # Obtener saldo
                        try:
                            saldo_val = int(saldo) if saldo is not None else 0
                        except (ValueError, TypeError):
                            saldo_val = 0
                        
                        # Obtener unidad
                        if isinstance(unidad, str) and len(str(unidad).strip()) <= 10:
                            unidad_val = str(unidad).strip().upper()
                        else:
                            unidad_val = 'UND'
                        
                        # Crear producto
                        codigo = self.crear_codigo_automatico('postcosecha', contador)
                        valor_unitario = self.estimar_valor_unitario(nombre_clean, categoria='postcosecha')
                        stock_minimo = max(10, int(saldo_val * 0.15)) if saldo_val > 0 else 20
                        
                        productos.append((
                            codigo, nombre_clean, 'postcosecha', saldo_val, unidad_val,
                            valor_unitario, stock_minimo, f'PC-{contador:02d}', 'PROVEEDOR POSTCOSECHA'
                        ))
                        
                        contador += 1
                        productos_encontrados += 1
                        
                        # Debug
                        if productos_encontrados <= 10:
                            print(f"     • {nombre_clean} - {saldo_val} {unidad_val} (fila {row_num})")
                        
            except Exception as e:
                continue
        
        print(f"   ✅ {len(productos)} productos POSTCOSECHA procesados")
        return productos
    
    def estimar_valor_unitario(self, nombre, categoria):
        """Estimar valor unitario basado en nombre y categoría"""
        nombre_lower = nombre.lower()
        
        if categoria == 'almacen':
            if 'aceite' in nombre_lower:
                return 45000
            elif 'acople' in nombre_lower or 'foster' in nombre_lower:
                return 15000
            elif 'arandela' in nombre_lower:
                return 500
            else:
                return 8500
                
        elif categoria == 'quimicos':
            if 'acaricida' in nombre_lower:
                return 80
            elif 'fungicida' in nombre_lower:
                return 120
            elif 'insecticida' in nombre_lower:
                return 100
            else:
                return 95
                
        elif categoria == 'postcosecha':
            if 'tapa' in nombre_lower or 'base' in nombre_lower:
                return 500
            elif 'octavo' in nombre_lower:
                return 300
            else:
                return 400
                
        return 1000  # Valor por defecto
    
    def crear_codigo_automatico(self, categoria, contador):
        """Crear código automático para productos"""
        prefijos = {
            'almacen': 'ALM',
            'quimicos': 'QM',
            'postcosecha': 'PC'
        }
        
        prefijo = prefijos.get(categoria, 'GEN')
        return f"{prefijo}{contador:03d}"
    
    def create_interface(self):
        """Crear interfaz moderna y atractiva"""
        # Header elegante con gradiente
        self.create_modern_header()
        
        # Contenedor principal con layout moderno
        main_container = tk.Frame(self.window, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Layout de dos columnas: sidebar + contenido principal
        self.create_sidebar(main_container)
        self.create_main_content(main_container)
    
    def create_modern_header(self):
        """Crear header moderno con gradiente"""
        # Header principal
        header = tk.Frame(self.window, bg='#2c3e50', height=70)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        # Contenido del header
        header_content = tk.Frame(header, bg='#2c3e50')
        header_content.pack(fill=tk.BOTH, expand=True, padx=30, pady=15)
        
        # Logo y título
        title_frame = tk.Frame(header_content, bg='#2c3e50')
        title_frame.pack(side=tk.LEFT)
        
        # Icono del sistema
        icon_label = tk.Label(title_frame, text="📦", font=('Arial', 24), 
                             bg='#2c3e50', fg='#ecf0f1')
        icon_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # Título y subtítulo
        title_container = tk.Frame(title_frame, bg='#2c3e50')
        title_container.pack(side=tk.LEFT)
        
        title_label = tk.Label(title_container, text="Sistema de Inventarios", 
                              font=('Segoe UI', 18, 'bold'), bg='#2c3e50', fg='#ecf0f1')
        title_label.pack(anchor='w')
        
        subtitle = tk.Label(title_container, text="Gestión integral de productos", 
                           font=('Segoe UI', 10), bg='#2c3e50', fg='#bdc3c7')
        subtitle.pack(anchor='w')
        
        # Botones de acción del header
        self.create_header_actions(header_content)
    
    def create_header_actions(self, parent):
        """Crear botones de acción en el header"""
        actions_frame = tk.Frame(parent, bg='#2c3e50')
        actions_frame.pack(side=tk.RIGHT)
        
        # Botones principales
        btn_configs = [
            ("+ Nuevo", self.nuevo_producto, "#27ae60", "Crear nuevo producto"),
            ("↓ Entrada", lambda: self.movimiento_dialog('entrada'), "#16a085", "Registrar entrada"),
            ("↑ Salida", lambda: self.movimiento_dialog('salida'), "#f39c12", "Registrar salida"),
            ("📊", self.generar_reporte_completo, "#8e44ad", "Generar reporte"),
            ("✕", self.cerrar_ventana, "#e74c3c", "Cerrar sistema")
        ]
        
        for text, command, color, tooltip in btn_configs:
            btn = tk.Button(actions_frame, text=text, command=command,
                           bg=color, fg='white', font=('Segoe UI', 9, 'bold'),
                           relief='flat', bd=0, padx=12, pady=6,
                           cursor='hand2', activebackground=self.darker_color(color))
            btn.pack(side=tk.LEFT, padx=3)
            
            # Efecto hover
            self.add_hover_effect(btn, color)
    
    def create_sidebar(self, parent):
        """Crear sidebar moderno con estadísticas"""
        sidebar = tk.Frame(parent, bg='#ffffff', width=280)
        sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=(20, 0), pady=20)
        sidebar.pack_propagate(False)
        
        # Título del sidebar
        sidebar_title = tk.Label(sidebar, text="📊 Dashboard", 
                                font=('Segoe UI', 16, 'bold'), 
                                bg='#ffffff', fg='#2c3e50')
        sidebar_title.pack(anchor='w', padx=20, pady=(20, 15))
        
        # Estadísticas en cards modernas
        self.create_stats_cards(sidebar)
        
        # Filtros avanzados
        self.create_advanced_filters(sidebar)
        
        # Acciones rápidas
        self.create_quick_actions_sidebar(sidebar)
    
    def create_stats_cards(self, parent):
        """Crear tarjetas de estadísticas modernas"""
        stats_container = tk.Frame(parent, bg='#ffffff')
        stats_container.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Obtener estadísticas
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM productos")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM productos WHERE saldo <= stock_minimo AND saldo > 0")
        bajo = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM productos WHERE saldo = 0")
        agotado = cursor.fetchone()[0]
        
        cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos")
        valor = cursor.fetchone()[0] or 0
        
        # Crear cards modernas
        cards_data = [
            ("Total Productos", str(total), "#3498db", "#ecf5ff"),
            ("Stock Bajo", str(bajo), "#f39c12", "#fff8e1"),
            ("Sin Stock", str(agotado), "#e74c3c", "#ffebee"),
            ("Valor Total", f"${valor:,.0f}", "#27ae60", "#e8f5e8")
        ]
        
        for i, (title, value, color, bg_color) in enumerate(cards_data):
            # Card container con efecto de elevación simulado
            card = tk.Frame(stats_container, bg=bg_color, relief='solid', bd=1)
            card.pack(fill=tk.X, pady=8)
            
            # Contenido de la card
            card_content = tk.Frame(card, bg=bg_color)
            card_content.pack(fill=tk.BOTH, padx=15, pady=12)
            
            # Valor grande
            value_label = tk.Label(card_content, text=value, 
                                  font=('Segoe UI', 18, 'bold'),
                                  bg=bg_color, fg=color)
            value_label.pack(anchor='w')
            
            # Título
            title_label = tk.Label(card_content, text=title,
                                  font=('Segoe UI', 10),
                                  bg=bg_color, fg='#7f8c8d')
            title_label.pack(anchor='w')
    
    def create_advanced_filters(self, parent):
        """Crear filtros avanzados en el sidebar"""
        filters_frame = tk.Frame(parent, bg='#ffffff')
        filters_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Título de filtros
        tk.Label(filters_frame, text="🔍 Filtros", 
                font=('Segoe UI', 14, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(anchor='w', pady=(0, 10))
        
        # Búsqueda
        search_frame = tk.Frame(filters_frame, bg='#ffffff')
        search_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(search_frame, text="Buscar producto:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var,
                               font=('Segoe UI', 10), relief='solid', bd=1,
                               bg='#f8f9fa', fg='#2c3e50')
        search_entry.pack(fill=tk.X, pady=(3, 0))
        
        # Filtro categoría
        cat_frame = tk.Frame(filters_frame, bg='#ffffff')
        cat_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(cat_frame, text="Categoría:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.filter_categoria = tk.StringVar(value="Todas")
        self.filter_categoria.trace('w', self.on_search_change)
        cat_combo = ttk.Combobox(cat_frame, textvariable=self.filter_categoria,
                                values=["Todas", "almacen", "postcosecha", "quimicos"],
                                state="readonly", font=('Segoe UI', 10))
        cat_combo.pack(fill=tk.X, pady=(3, 0))
        
        # Filtro estado
        state_frame = tk.Frame(filters_frame, bg='#ffffff')
        state_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(state_frame, text="Estado del stock:", 
                font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w')
        
        self.filter_stock = tk.StringVar(value="Todos")
        self.filter_stock.trace('w', self.on_search_change)
        state_combo = ttk.Combobox(state_frame, textvariable=self.filter_stock,
                                  values=["Todos", "Normal", "Bajo", "Agotado"],
                                  state="readonly", font=('Segoe UI', 10))
        state_combo.pack(fill=tk.X, pady=(3, 0))
        
        # Botón limpiar filtros
        clear_btn = tk.Button(filters_frame, text="🗑️ Limpiar filtros", 
                             command=self.limpiar_filtros,
                             bg='#95a5a6', fg='white', font=('Segoe UI', 9),
                             relief='flat', bd=0, pady=8, cursor='hand2')
        clear_btn.pack(fill=tk.X, pady=(10, 0))
    
    def create_quick_actions_sidebar(self, parent):
        """Crear acciones rápidas en el sidebar"""
        actions_frame = tk.Frame(parent, bg='#ffffff')
        actions_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Título
        tk.Label(actions_frame, text="⚡ Acciones Rápidas", 
                font=('Segoe UI', 14, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(anchor='w', pady=(0, 15))
        
        # Botones de acción
        actions = [
            ("✏️ Editar Seleccionado", self.editar_producto, "#3498db"),
            ("⚠️ Ver Alertas", self.mostrar_alertas, "#e74c3c"),
            ("📈 Estadísticas Detalladas", self.mostrar_estadisticas_detalladas, "#9b59b6")
        ]
        
        for text, command, color in actions:
            btn = tk.Button(actions_frame, text=text, command=command,
                           bg=color, fg='white', font=('Segoe UI', 10),
                           relief='flat', bd=0, pady=8, cursor='hand2')
            btn.pack(fill=tk.X, pady=3)
            self.add_hover_effect(btn, color)
    
    def create_main_content(self, parent):
        """Crear área de contenido principal"""
        content_area = tk.Frame(parent, bg='#f8f9fa')
        content_area.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Card principal para la tabla
        table_card = tk.Frame(content_area, bg='#ffffff', relief='solid', bd=1)
        table_card.pack(fill=tk.BOTH, expand=True)
        
        # Header de la tabla
        table_header = tk.Frame(table_card, bg='#ffffff', height=50)
        table_header.pack(fill=tk.X)
        table_header.pack_propagate(False)
        
        # Título de la tabla
        tk.Label(table_header, text="📋 Lista de Productos", 
                font=('Segoe UI', 16, 'bold'),
                bg='#ffffff', fg='#2c3e50').pack(side=tk.LEFT, padx=20, pady=15)
        
        # Contador de productos
        self.products_count_label = tk.Label(table_header, text="", 
                                           font=('Segoe UI', 10),
                                           bg='#ffffff', fg='#7f8c8d')
        self.products_count_label.pack(side=tk.RIGHT, padx=20, pady=15)
        
        # Tabla moderna
        self.create_modern_table(table_card)
    
    def create_modern_table(self, parent):
        """Crear tabla moderna y elegante"""
        table_container = tk.Frame(parent, bg='#ffffff')
        table_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Configurar estilo moderno para el treeview
        self.configure_modern_treeview_style()
        
        # Crear treeview
        columns = ('Código', 'Nombre', 'Categoría', 'Saldo', 'Unidad', 'Valor Unit.', 'Stock Min.', 'Estado')
        self.tree = ttk.Treeview(table_container, columns=columns, show='headings',
                                height=20, style='Modern.Treeview')
        
        # Configurar columnas con mejor distribución
        column_config = {
            'Código': {'width': 80, 'anchor': 'center'},
            'Nombre': {'width': 300, 'anchor': 'w'},
            'Categoría': {'width': 100, 'anchor': 'center'},
            'Saldo': {'width': 80, 'anchor': 'center'},
            'Unidad': {'width': 80, 'anchor': 'center'},
            'Valor Unit.': {'width': 100, 'anchor': 'e'},
            'Stock Min.': {'width': 80, 'anchor': 'center'},
            'Estado': {'width': 100, 'anchor': 'center'}
        }
        
        for col, config in column_config.items():
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c))
            self.tree.column(col, width=config['width'], anchor=config['anchor'])
        
        # Scrollbars modernas
        v_scrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(table_container, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        table_container.grid_rowconfigure(0, weight=1)
        table_container.grid_columnconfigure(0, weight=1)
        
        # Eventos
        self.tree.bind('<Double-1>', lambda e: self.editar_producto())
        self.tree.bind('<Button-3>', self.mostrar_menu_contextual)
    
    def configure_modern_treeview_style(self):
        """Configurar estilo moderno para el treeview"""
        style = ttk.Style()
        
        # Estilo principal del treeview
        style.configure('Modern.Treeview',
                       background='#ffffff',
                       foreground='#2c3e50',
                       fieldbackground='#ffffff',
                       font=('Segoe UI', 10),
                       rowheight=32,
                       borderwidth=0)
        
        # Estilo de los headers
        style.configure('Modern.Treeview.Heading',
                       background='#34495e',
                       foreground='#ffffff',
                       font=('Segoe UI', 10, 'bold'),
                       relief='flat',
                       borderwidth=1)
        
        # Efectos de selección y hover
        style.map('Modern.Treeview',
                 background=[('selected', '#3498db')],
                 foreground=[('selected', '#ffffff')])
        
        style.map('Modern.Treeview.Heading',
                 background=[('active', '#2c3e50')])
    
    def add_hover_effect(self, button, original_color):
        """Agregar efecto hover elegante a los botones"""
        darker = self.darker_color(original_color)
        
        def on_enter(e):
            button.config(bg=darker)
        
        def on_leave(e):
            button.config(bg=original_color)
        
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
    
    def create_action_buttons(self, parent):
        """Crear botones de acción con diseño minimalista"""
        btn_frame = tk.Frame(parent, bg='#f0f0f0')
        btn_frame.grid(row=1, column=0, columnspan=4, pady=(0, 20))
        
        buttons_info = [
            ("+ Nuevo", self.nuevo_producto, "#2c3e50"),
            ("✏ Editar", self.editar_producto, "#34495e"),
            ("↓ Entrada", lambda: self.movimiento_dialog('entrada'), "#27ae60"),
            ("↑ Salida", lambda: self.movimiento_dialog('salida'), "#e67e22"),
            ("! Alertas", self.mostrar_alertas, "#e74c3c"),
            ("📊 Reporte", self.generar_reporte_completo, "#8e44ad"),
            ("✕ Cerrar", self.cerrar_ventana, "#7f8c8d")
        ]
        
        for i, (text, command, color) in enumerate(buttons_info):
            btn = tk.Button(btn_frame, text=text, command=command,
                           bg=color, fg='white', font=('Arial', 9, 'bold'),
                           relief='flat', padx=12, pady=6, cursor='hand2',
                           bd=0, activebackground=self.darker_color(color),
                           width=10)
            btn.pack(side=tk.LEFT, padx=3)
            
            # Efecto hover
            self.make_hover(btn, color)
    
    def darker_color(self, hex_color):
        """Colores hover simplificados"""
        colors = {
            "#2c3e50": "#1a252f",
            "#34495e": "#2c3e50", 
            "#27ae60": "#229954",
            "#e67e22": "#d35400",
            "#e74c3c": "#c0392b",
            "#8e44ad": "#7d3c98",
            "#7f8c8d": "#6c7b7d"
        }
        return colors.get(hex_color, hex_color)
    
    def make_hover(self, button, color):
        """Compatibilidad con función anterior - redirige a add_hover_effect"""
        self.add_hover_effect(button, color)
    
    def create_search_frame(self, parent):
        """Crear frame de búsqueda simple"""
        search_frame = ttk.LabelFrame(parent, text="🔍 Búsqueda", padding="10")
        search_frame.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Búsqueda por texto
        ttk.Label(search_frame, text="Buscar:", font=('Arial', 9)).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=20, font=('Arial', 9))
        search_entry.grid(row=0, column=1, padx=(0, 15))
        
        # Filtro por categoría
        ttk.Label(search_frame, text="Categoría:", font=('Arial', 9)).grid(row=0, column=2, sticky=tk.W, padx=(0, 5))
        self.filter_categoria = tk.StringVar()
        self.filter_categoria.trace('w', self.on_search_change)
        categoria_combo = ttk.Combobox(search_frame, textvariable=self.filter_categoria,
                                     values=["Todas", "almacen", "postcosecha", "quimicos"], width=12)
        categoria_combo.set("Todas")
        categoria_combo.grid(row=0, column=3, padx=(0, 15))
        
        # Filtro por estado de stock
        ttk.Label(search_frame, text="Estado:", font=('Arial', 9)).grid(row=0, column=4, sticky=tk.W, padx=(0, 5))
        self.filter_stock = tk.StringVar()
        self.filter_stock.trace('w', self.on_search_change)
        stock_combo = ttk.Combobox(search_frame, textvariable=self.filter_stock,
                                 values=["Todos", "Normal", "Bajo", "Agotado"], width=10)
        stock_combo.set("Todos")
        stock_combo.grid(row=0, column=5, padx=(0, 10))
        
        # Botón limpiar
        clear_btn = tk.Button(search_frame, text="Limpiar", command=self.limpiar_filtros,
                             bg="#95a5a6", fg="white", font=('Arial', 8, 'bold'),
                             relief='flat', padx=8, pady=3, cursor='hand2')
        clear_btn.grid(row=0, column=6)
    
    def create_simple_stats(self, parent):
        """Crear estadísticas simples y minimalistas"""
        stats_frame = tk.Frame(parent, bg='#f0f0f0')
        stats_frame.grid(row=3, column=0, columnspan=4, pady=(0, 15))
        
        # Obtener estadísticas
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM productos")
        total_productos = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM productos WHERE saldo <= stock_minimo AND saldo > 0")
        stock_bajo = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM productos WHERE saldo = 0")
        sin_stock = cursor.fetchone()[0]
        
        cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos")
        valor_total = cursor.fetchone()[0] or 0
        
        # Crear tarjetas simples
        stats = [
            ("Total", str(total_productos), "#ecf0f1"),
            ("Stock Bajo", str(stock_bajo), "#ecf0f1"),
            ("Sin Stock", str(sin_stock), "#ecf0f1"),
            ("Valor Total", f"${valor_total:,.0f}", "#ecf0f1")
        ]
        
        for i, (title, value, bg_color) in enumerate(stats):
            stat_frame = tk.Frame(stats_frame, bg=bg_color, relief='solid', bd=1)
            stat_frame.pack(side=tk.LEFT, padx=8, ipadx=15, ipady=8)
            
            tk.Label(stat_frame, text=title, bg=bg_color, fg='#2c3e50',
                    font=('Arial', 9, 'bold')).pack()
            tk.Label(stat_frame, text=value, bg=bg_color, fg='#2c3e50',
                    font=('Arial', 11, 'bold')).pack()
    
    def create_inventory_table(self, parent):
        """Crear tabla de inventario minimalista"""
        inventario_frame = ttk.LabelFrame(parent, text="Lista de Productos", padding="10")
        inventario_frame.grid(row=4, column=0, columnspan=4, pady=(10, 0), sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Treeview para mostrar inventario
        columns = ('Código', 'Nombre', 'Categoría', 'Saldo', 'Unidad', 'Valor Unit.', 'Stock Min.', 'Estado')
        self.tree = ttk.Treeview(inventario_frame, columns=columns, show='headings',
                                height=18, style='Inventario.Treeview')
        
        # Configurar columnas
        column_widths = {
            'Código': 70, 'Nombre': 220, 'Categoría': 90, 'Saldo': 70,
            'Unidad': 70, 'Valor Unit.': 90, 'Stock Min.': 70, 'Estado': 90
        }
        
        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c))
            self.tree.column(col, width=column_widths.get(col, 100))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(inventario_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(inventario_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid para tree y scrollbars
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Eventos
        self.tree.bind('<Double-1>', lambda e: self.editar_producto())
        self.tree.bind('<Button-3>', self.mostrar_menu_contextual)
        
        # Configurar grid weights
        inventario_frame.columnconfigure(0, weight=1)
        inventario_frame.rowconfigure(0, weight=1)
    
    def mostrar_menu_contextual(self, event):
        """Mostrar menú contextual"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            
            menu = tk.Menu(self.window, tearoff=0)
            menu.add_command(label="✏️ Editar producto", command=self.editar_producto)
            menu.add_command(label="📥 Registrar entrada", command=lambda: self.movimiento_dialog('entrada'))
            menu.add_command(label="📤 Registrar salida", command=lambda: self.movimiento_dialog('salida'))
            menu.add_separator()
            menu.add_command(label="🗑️ Eliminar producto", command=self.eliminar_producto)
            
            menu.post(event.x_root, event.y_root)
    
    # =================== FUNCIONES DE FILTRADO Y BÚSQUEDA ===================
    
    def on_search_change(self, *args):
        """Llamada cuando cambian los filtros"""
        self.cargar_inventario()
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros"""
        self.search_var.set("")
        self.filter_categoria.set("Todas")
        self.filter_stock.set("Todos")
    
    def cargar_inventario(self):
        """Cargar inventario con la nueva interfaz moderna"""
        try:
            # Limpiar datos existentes
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Construir query con filtros
            query = "SELECT codigo, nombre, categoria, saldo, unidad, valor_unitario, stock_minimo, ubicacion FROM productos WHERE 1=1"
            params = []
            
            # Aplicar filtros
            if hasattr(self, 'search_var'):
                texto_busqueda = self.search_var.get().strip().lower()
                if texto_busqueda:
                    query += " AND (LOWER(nombre) LIKE ? OR LOWER(codigo) LIKE ?)"
                    params.extend([f'%{texto_busqueda}%', f'%{texto_busqueda}%'])
            
            if hasattr(self, 'filter_categoria'):
                categoria_filtro = self.filter_categoria.get()
                if categoria_filtro and categoria_filtro != "Todas":
                    query += " AND categoria = ?"
                    params.append(categoria_filtro)
            
            if hasattr(self, 'filter_stock'):  
                stock_filtro = self.filter_stock.get()
                if stock_filtro == "Bajo":
                    query += " AND saldo <= stock_minimo AND saldo > 0"
                elif stock_filtro == "Agotado":
                    query += " AND saldo = 0"
                elif stock_filtro == "Normal":
                    query += " AND saldo > stock_minimo"
            
            query += " ORDER BY categoria, codigo"
            
            # Ejecutar query
            cursor = self.conn.cursor()
            cursor.execute(query, params)
            productos_mostrados = 0
            
            for row in cursor.fetchall():
                codigo, nombre, categoria, saldo, unidad, valor_unit, stock_min, ubicacion = row
                
                # Determinar estado con indicadores modernos
                if saldo == 0:
                    estado = "• AGOTADO"
                    tags = ('agotado',)
                elif saldo <= stock_min:
                    estado = "• STOCK BAJO"
                    tags = ('bajo',)
                else:
                    estado = "• NORMAL"
                    tags = ('normal',)
                
                valor_formateado = f"${valor_unit:,.0f}"
                
                self.tree.insert('', tk.END, values=(
                    codigo, nombre, categoria.title(), saldo, unidad,
                    valor_formateado, stock_min, estado
                ), tags=tags)
                productos_mostrados += 1
            
            # Configurar colores modernos para los estados
            self.tree.tag_configure('agotado', background='#ffebee', foreground='#d32f2f')
            self.tree.tag_configure('bajo', background='#fff3e0', foreground='#f57c00')
            self.tree.tag_configure('normal', background='#ffffff', foreground='#2c3e50')
            
            # Actualizar contador en el header de la tabla
            if hasattr(self, 'products_count_label'):
                cursor.execute("SELECT COUNT(*) FROM productos")
                total_productos = cursor.fetchone()[0]
                self.products_count_label.config(
                    text=f"Mostrando {productos_mostrados} de {total_productos} productos"
                )
            
            # Actualizar estadísticas del sidebar
            self.update_sidebar_stats()
            
            return total_productos
            
        except Exception as e:
            print(f"Error al cargar inventario: {e}")
            return 0
    
    def update_sidebar_stats(self):
        """Actualizar estadísticas del sidebar"""
        try:
            # Esta función se ejecuta automáticamente cuando se filtra
            # Las estadísticas se actualizan en create_stats_cards
            pass
        except Exception as e:
            print(f"Error actualizando estadísticas: {e}")
    
    def on_search_change(self, *args):
        """Actualizar tabla cuando cambian los filtros"""
        self.cargar_inventario()
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros con la nueva interfaz"""
        if hasattr(self, 'search_var'):
            self.search_var.set("")
        if hasattr(self, 'filter_categoria'):
            self.filter_categoria.set("Todas")
        if hasattr(self, 'filter_stock'):
            self.filter_stock.set("Todos")
    
    def mostrar_info_carga_exitosa(self, total_productos):
        """Mostrar información sobre carga exitosa desde Excel"""
        cursor = self.conn.cursor()
        
        # Obtener estadísticas por categoría
        stats = []
        for categoria in ['almacen', 'quimicos', 'postcosecha']:
            cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria = ?", (categoria,))
            count = cursor.fetchone()[0]
            if count > 0:
                stats.append(f"• {categoria.upper()}: {count} productos")
        
        # Calcular valor total
        cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos")
        valor_total = cursor.fetchone()[0] or 0
        
        messagebox.showinfo(
            "🎉 Datos Cargados Exitosamente", 
            f"¡Sistema de Inventarios listo!\n\n" +
            f"📦 Total de productos cargados: {total_productos}\n\n" +
            f"📊 Desglose por categoría:\n" + "\n".join(stats) + 
            f"\n\n💰 Valor total del inventario: ${valor_total:,.0f}\n\n" +
            "Los datos fueron cargados automáticamente desde\n" +
            "los archivos Excel del sistema. ¡Ya puedes empezar\n" +
            "a gestionar tu inventario!"
        )
    
    def sort_column(self, col):
        """Ordenar TreeView por columna"""
        try:
            data = [(self.tree.set(child, col), child) for child in self.tree.get_children('')]
            data.sort()
            
            for index, (val, child) in enumerate(data):
                self.tree.move(child, '', index)
        except:
            pass
    
    # =================== FUNCIONES PRINCIPALES ===================
    
    def get_selected_producto(self):
        """Obtener producto seleccionado"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor selecciona un producto")
            return None
        
        item = self.tree.item(selection[0])
        codigo = item['values'][0]
        
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM productos WHERE codigo = ?", (codigo,))
        return cursor.fetchone()
    
    def nuevo_producto(self):
        """Crear nuevo producto"""
        ProductoWindow(self.window, self, modo="nuevo")
    
    def editar_producto(self):
        """Editar producto seleccionado"""
        producto = self.get_selected_producto()
        if producto:
            ProductoWindow(self.window, self, modo="editar", producto=producto)
    
    def eliminar_producto(self):
        """Eliminar producto seleccionado"""
        producto = self.get_selected_producto()
        if not producto:
            return
        
        codigo, nombre = producto[1], producto[2]
        
        if messagebox.askyesno("Confirmar eliminación",
                              f"¿Estás seguro de eliminar el producto:\n{codigo} - {nombre}?"):
            try:
                cursor = self.conn.cursor()
                cursor.execute("DELETE FROM productos WHERE codigo = ?", (codigo,))
                self.conn.commit()
                messagebox.showinfo("Éxito", "Producto eliminado correctamente")
                self.cargar_inventario()
            except Exception as e:
                messagebox.showerror("Error", f"Error al eliminar producto: {e}")
    
    def movimiento_dialog(self, tipo):
        """Diálogo para registrar movimientos"""
        MovimientoWindow(self.window, self, tipo)
    
    def mostrar_alertas(self):
        """Mostrar alertas de stock bajo"""
        AlertasWindow(self.window, self)
    
    def importar_excel_completo(self):
        """Reimportar datos desde los archivos Excel del sistema"""
        respuesta = messagebox.askyesno(
            "🔄 Reimportar desde Excel", 
            "¿Deseas RECARGAR todos los productos desde los archivos Excel?\n\n" + 
            "⚠️ ATENCIÓN: Esto eliminará todos los datos actuales y los\n" +
            "reemplazará con los datos más recientes de los archivos Excel.\n\n" +
            "Archivos que se procesarán:\n" +
            "• INVENTARIO ALMACEN JUNIO 2025.xlsx\n" + 
            "• INV QUIMICOS JUNIO 2025 1.xlsx\n" +
            "• SALDOS POSCOSECHA JUNIO 2025.xlsx\n\n" +
            "¿Continuar con la reimportación?"
        )
        
        if not respuesta:
            return
            
        try:
            print("\n🔄 INICIANDO REIMPORTACIÓN DESDE EXCEL...")
            
            # Hacer backup de conteo actual
            cursor = self.conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos")
            productos_anteriores = cursor.fetchone()[0]
            print(f"📊 Productos actuales en BD: {productos_anteriores}")
            
            # Limpiar tabla actual
            cursor.execute("DELETE FROM productos")
            print("🗑️ Base de datos limpiada")
            
            # Cargar datos desde Excel
            if self.cargar_datos_desde_excel():
                messagebox.showinfo("✅ Reimportación Exitosa", 
                    "Datos reimportados correctamente desde archivos Excel.\n\n" +
                    f"Productos anteriores: {productos_anteriores}\n" +
                    f"Productos nuevos: {self.obtener_total_productos()}\n\n" +
                    "La tabla de inventario ha sido actualizada con los datos más recientes.")
                self.cargar_inventario()  # Refrescar la vista
            else:
                # Restaurar datos de ejemplo si falla
                print("⚠️ Reimportación falló, restaurando datos de ejemplo...")
                cursor.execute("DELETE FROM productos")
                self.poblar_datos_iniciales()
                messagebox.showwarning("⚠️ Error en Reimportación", 
                    "No se pudieron cargar los archivos Excel.\n\n" +
                    "Posibles causas:\n" +
                    "• Los archivos no están en la misma carpeta\n" +
                    "• Los nombres de archivo no coinciden exactamente\n" +
                    "• Los archivos están abiertos en Excel\n\n" +
                    "Se han restaurado los datos de ejemplo.")
                self.cargar_inventario()
                
        except Exception as e:
            print(f"❌ Error crítico en reimportación: {e}")
            messagebox.showerror("❌ Error Crítico", f"Error al reimportar datos: {e}")
            # Restaurar datos de ejemplo en caso de error crítico
            try:
                cursor = self.conn.cursor()
                cursor.execute("DELETE FROM productos")
                self.poblar_datos_iniciales()
                self.cargar_inventario()
            except:
                pass
    
    def obtener_total_productos(self):
        """Obtener total de productos actual"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos")
            return cursor.fetchone()[0]
        except:
            return 0
    
    def generar_reporte_completo(self):
        """Generar reporte completo con estadísticas por categoría"""
        try:
            archivo = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialvalue=f"inventario_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if archivo:
                cursor = self.conn.cursor()
                
                # Reporte principal
                cursor.execute("""
                    SELECT codigo, nombre, categoria, saldo, unidad, valor_unitario, 
                           stock_minimo, ubicacion, proveedor,
                           (saldo * valor_unitario) as valor_total
                    FROM productos 
                    ORDER BY categoria, codigo
                """)
                productos = cursor.fetchall()
                
                with open(archivo, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    
                    # Headers
                    writer.writerow(['Código', 'Nombre', 'Categoría', 'Saldo', 'Unidad',
                                   'Valor Unitario', 'Stock Mínimo', 'Ubicación', 'Proveedor', 'Valor Total'])
                    
                    # Datos
                    for producto in productos:
                        writer.writerow(producto)
                    
                    # Estadísticas por categoría
                    writer.writerow([])
                    writer.writerow(['=== ESTADÍSTICAS POR CATEGORÍA ==='])
                    writer.writerow(['Categoría', 'Total Productos', 'Valor Total', 'Stock Bajo', 'Sin Stock'])
                    
                    for categoria in ['almacen', 'postcosecha', 'quimicos']:
                        cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria = ?", (categoria,))
                        total = cursor.fetchone()[0]
                        
                        cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos WHERE categoria = ?", (categoria,))
                        valor = cursor.fetchone()[0] or 0
                        
                        cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria = ? AND saldo <= stock_minimo AND saldo > 0", (categoria,))
                        bajo = cursor.fetchone()[0]
                        
                        cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria = ? AND saldo = 0", (categoria,))
                        sin_stock = cursor.fetchone()[0]
                        
                        writer.writerow([categoria.upper(), total, f"${valor:,.0f}", bajo, sin_stock])
                
                messagebox.showinfo("Éxito", f"Reporte completo generado: {archivo}")
                
                # Preguntar si desea abrir
                if messagebox.askyesno("Abrir archivo", "¿Deseas abrir el archivo generado?"):
                    try:
                        os.startfile(archivo)
                    except:
                        pass
                        
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar reporte: {e}")
    
    def mostrar_estadisticas_detalladas(self):
        """Mostrar ventana con estadísticas detalladas"""
        EstadisticasWindow(self.window, self)
    

    
    def cerrar_ventana(self):
        """Cerrar ventana y limpiar recursos"""
        if self.conn:
            self.conn.close()
        self.window.destroy()


# =================== VENTANAS DE DIÁLOGO ===================

class ProductoWindow:
    """Ventana para crear/editar productos"""
    
    def __init__(self, parent, main_window, modo="nuevo", producto=None):
        self.parent = parent
        self.main_window = main_window
        self.modo = modo
        self.producto = producto
        
        # Crear ventana
        self.window = tk.Toplevel(parent)
        titulo = "✏️ Editar Producto" if modo == "editar" else "➕ Nuevo Producto"
        self.window.title(titulo)
        self.window.geometry("500x550")
        self.window.configure(bg='#ecf0f1')
        self.window.resizable(False, False)
        
        # Centrar y hacer modal
        self.center_window()
        self.window.transient(parent)
        self.window.grab_set()
        
        # Crear interfaz
        self.create_widgets()
        
        # Si es edición, cargar datos
        if modo == "editar" and producto:
            self.cargar_datos_producto()
    
    def center_window(self):
        """Centrar ventana"""
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (500 // 2)
        y = (self.window.winfo_screenheight() // 2) - (550 // 2)
        self.window.geometry(f"500x550+{x}+{y}")
    
    def create_widgets(self):
        """Crear widgets de la ventana"""
        # Frame principal
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Título
        titulo_texto = "✏️ Editar Producto" if self.modo == "editar" else "➕ Nuevo Producto"
        title_label = tk.Label(main_frame, text=titulo_texto, font=('Arial', 16, 'bold'),
                              bg='#ecf0f1', fg='#2c3e50')
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Campos del formulario
        campos = [
            ("Código:", "entry_codigo"),
            ("Nombre:", "entry_nombre"),
            ("Categoría:", "combo_categoria"),
            ("Saldo:", "entry_saldo"),
            ("Unidad:", "entry_unidad"),
            ("Valor Unitario:", "entry_valor"),
            ("Stock Mínimo:", "entry_stock_min"),
            ("Ubicación:", "entry_ubicacion"),
            ("Proveedor:", "entry_proveedor")
        ]
        
        for i, (label_text, widget_name) in enumerate(campos, 1):
            label = ttk.Label(main_frame, text=label_text, font=('Arial', 10, 'bold'))
            label.grid(row=i, column=0, sticky=tk.W, pady=8, padx=(0, 10))
            
            if widget_name == "combo_categoria":
                widget = ttk.Combobox(main_frame, values=["almacen", "postcosecha", "quimicos"],
                                    width=27, font=('Arial', 10))
            else:
                widget = ttk.Entry(main_frame, width=30, font=('Arial', 10))
            
            widget.grid(row=i, column=1, pady=8)
            setattr(self, widget_name, widget)
        
        # Botones
        btn_frame = tk.Frame(main_frame, bg='#ecf0f1')
        btn_frame.grid(row=10, column=0, columnspan=2, pady=25)
        
        # Botón guardar
        texto_boton = "💾 Actualizar" if self.modo == "editar" else "💾 Guardar"
        save_btn = tk.Button(btn_frame, text=texto_boton, command=self.guardar_producto,
                           bg="#27ae60", fg="white", font=('Arial', 10, 'bold'),
                           relief='flat', padx=20, pady=8, cursor='hand2')
        save_btn.pack(side=tk.LEFT, padx=5)
        
        # Botón limpiar
        clear_btn = tk.Button(btn_frame, text="🗑️ Limpiar", command=self.limpiar_campos,
                            bg="#f39c12", fg="white", font=('Arial', 10, 'bold'),
                            relief='flat', padx=20, pady=8, cursor='hand2')
        clear_btn.pack(side=tk.LEFT, padx=5)
        
        # Botón cerrar
        close_btn = tk.Button(btn_frame, text="❌ Cerrar", command=self.window.destroy,
                            bg="#e74c3c", fg="white", font=('Arial', 10, 'bold'),
                            relief='flat', padx=20, pady=8, cursor='hand2')
        close_btn.pack(side=tk.LEFT, padx=5)
    
    def cargar_datos_producto(self):
        """Cargar datos del producto en el formulario"""
        if not self.producto:
            return
        
        self.entry_codigo.insert(0, self.producto[1] or "")
        self.entry_nombre.insert(0, self.producto[2] or "")
        self.combo_categoria.set(self.producto[3] or "")
        self.entry_saldo.insert(0, str(self.producto[4] or ""))
        self.entry_unidad.insert(0, self.producto[5] or "")
        self.entry_valor.insert(0, str(self.producto[6] or ""))
        self.entry_stock_min.insert(0, str(self.producto[7] or ""))
        self.entry_ubicacion.insert(0, self.producto[8] or "")
        self.entry_proveedor.insert(0, self.producto[9] or "")
        
        # Deshabilitar código en edición
        if self.modo == "editar":
            self.entry_codigo.config(state='disabled')
    
    def guardar_producto(self):
        """Guardar producto"""
        try:
            # Validar campos obligatorios
            if not all([self.entry_codigo.get().strip(), self.entry_nombre.get().strip(),
                       self.combo_categoria.get(), self.entry_unidad.get().strip()]):
                messagebox.showerror("Error", "Completa todos los campos obligatorios")
                return
            
            # Validar números
            try:
                saldo = int(self.entry_saldo.get() or 0)
                valor_unitario = float(self.entry_valor.get() or 0)
                stock_minimo = int(self.entry_stock_min.get() or 0)
            except ValueError:
                messagebox.showerror("Error", "Los valores numéricos no son válidos")
                return
            
            cursor = self.main_window.conn.cursor()
            
            if self.modo == "nuevo":
                # Verificar que no exista el código
                cursor.execute("SELECT id FROM productos WHERE codigo = ?", (self.entry_codigo.get().strip(),))
                if cursor.fetchone():
                    messagebox.showerror("Error", "Ya existe un producto con ese código")
                    return
                
                # Insertar nuevo producto
                cursor.execute('''
                    INSERT INTO productos 
                    (codigo, nombre, categoria, saldo, unidad, valor_unitario, stock_minimo, ubicacion, proveedor)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    self.entry_codigo.get().strip(),
                    self.entry_nombre.get().strip(),
                    self.combo_categoria.get(),
                    saldo,
                    self.entry_unidad.get().strip(),
                    valor_unitario,
                    stock_minimo,
                    self.entry_ubicacion.get().strip(),
                    self.entry_proveedor.get().strip()
                ))
                mensaje = "Producto creado exitosamente"
            else:
                # Actualizar producto existente
                cursor.execute('''
                    UPDATE productos SET 
                    nombre=?, categoria=?, saldo=?, unidad=?, valor_unitario=?, 
                    stock_minimo=?, ubicacion=?, proveedor=?
                    WHERE codigo=?
                ''', (
                    self.entry_nombre.get().strip(),
                    self.combo_categoria.get(),
                    saldo,
                    self.entry_unidad.get().strip(),
                    valor_unitario,
                    stock_minimo,
                    self.entry_ubicacion.get().strip(),
                    self.entry_proveedor.get().strip(),
                    self.entry_codigo.get().strip()
                ))
                mensaje = "Producto actualizado exitosamente"
            
            self.main_window.conn.commit()
            messagebox.showinfo("Éxito", mensaje)
            
            # Actualizar tabla principal
            self.main_window.cargar_inventario()
            
            # Limpiar o cerrar
            if self.modo == "nuevo":
                self.limpiar_campos()
            else:
                self.window.destroy()
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar producto: {e}")
    
    def limpiar_campos(self):
        """Limpiar todos los campos"""
        for widget_name in ['entry_codigo', 'entry_nombre', 'entry_saldo', 'entry_unidad',
                           'entry_valor', 'entry_stock_min', 'entry_ubicacion', 'entry_proveedor']:
            widget = getattr(self, widget_name)
            widget.delete(0, tk.END)
        self.combo_categoria.set('')


class MovimientoWindow:
    """Ventana para registrar movimientos"""
    
    def __init__(self, parent, main_window, tipo):
        self.parent = parent
        self.main_window = main_window
        self.tipo = tipo
        
        # Crear ventana
        self.window = tk.Toplevel(parent)
        titulo = f"📥 Registrar Entrada" if tipo == 'entrada' else f"📤 Registrar Salida"
        self.window.title(titulo)
        self.window.geometry("450x400")
        self.window.configure(bg='#ecf0f1')
        self.window.resizable(False, False)
        
        # Centrar y hacer modal
        self.center_window()
        self.window.transient(parent)
        self.window.grab_set()
        
        # Crear interfaz
        self.create_widgets()
    
    def center_window(self):
        """Centrar ventana"""
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (450 // 2)
        y = (self.window.winfo_screenheight() // 2) - (400 // 2)
        self.window.geometry(f"450x400+{x}+{y}")
    
    def create_widgets(self):
        """Crear widgets"""
        # Frame principal
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Título
        icon = "📥" if self.tipo == 'entrada' else "📤"
        titulo = f"{icon} Registrar {'Entrada' if self.tipo == 'entrada' else 'Salida'}"
        title_label = tk.Label(main_frame, text=titulo, font=('Arial', 16, 'bold'),
                              bg='#ecf0f1', fg='#2c3e50')
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Producto
        ttk.Label(main_frame, text="Producto:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.W, pady=8, padx=(0, 10))
        
        self.combo_producto = ttk.Combobox(main_frame, width=35, font=('Arial', 10))
        self.combo_producto.grid(row=1, column=1, pady=8)
        self.cargar_productos()
        
        # Cantidad
        ttk.Label(main_frame, text="Cantidad:", font=('Arial', 10, 'bold')).grid(
            row=2, column=0, sticky=tk.W, pady=8, padx=(0, 10))
        
        self.entry_cantidad = ttk.Entry(main_frame, width=38, font=('Arial', 10))
        self.entry_cantidad.grid(row=2, column=1, pady=8)
        
        # Campos adicionales según el tipo
        row = 3
        if self.tipo == 'entrada':
            # Factura
            ttk.Label(main_frame, text="Factura:", font=('Arial', 10, 'bold')).grid(
                row=row, column=0, sticky=tk.W, pady=8, padx=(0, 10))
            self.entry_factura = ttk.Entry(main_frame, width=38, font=('Arial', 10))
            self.entry_factura.grid(row=row, column=1, pady=8)
            row += 1
            
            # Proveedor
            ttk.Label(main_frame, text="Proveedor:", font=('Arial', 10, 'bold')).grid(
                row=row, column=0, sticky=tk.W, pady=8, padx=(0, 10))
            self.entry_proveedor = ttk.Entry(main_frame, width=38, font=('Arial', 10))
            self.entry_proveedor.grid(row=row, column=1, pady=8)
        else:
            # Destino
            ttk.Label(main_frame, text="Destino:", font=('Arial', 10, 'bold')).grid(
                row=row, column=0, sticky=tk.W, pady=8, padx=(0, 10))
            self.entry_destino = ttk.Entry(main_frame, width=38, font=('Arial', 10))
            self.entry_destino.grid(row=row, column=1, pady=8)
        
        row += 1
        
        # Observaciones
        ttk.Label(main_frame, text="Observaciones:", font=('Arial', 10, 'bold')).grid(
            row=row, column=0, sticky=tk.W, pady=8, padx=(0, 10))
        self.entry_observaciones = ttk.Entry(main_frame, width=38, font=('Arial', 10))
        self.entry_observaciones.grid(row=row, column=1, pady=8)
        
        # Botones
        btn_frame = tk.Frame(main_frame, bg='#ecf0f1')
        btn_frame.grid(row=row+1, column=0, columnspan=2, pady=25)
        
        color = "#27ae60" if self.tipo == 'entrada' else "#f39c12"
        tk.Button(btn_frame, text=f"💾 Registrar {self.tipo.title()}", 
                 command=self.registrar_movimiento, bg=color, fg="white",
                 font=('Arial', 10, 'bold'), relief='flat', padx=20, pady=8,
                 cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="❌ Cancelar", command=self.window.destroy,
                 bg="#e74c3c", fg="white", font=('Arial', 10, 'bold'),
                 relief='flat', padx=20, pady=8, cursor='hand2').pack(side=tk.LEFT, padx=5)
    
    def cargar_productos(self):
        """Cargar productos en el combobox"""
        cursor = self.main_window.conn.cursor()
        cursor.execute("SELECT codigo, nombre FROM productos ORDER BY codigo")
        productos = [f"{row[0]} - {row[1]}" for row in cursor.fetchall()]
        self.combo_producto['values'] = productos
    
    def registrar_movimiento(self):
        """Registrar el movimiento"""
        try:
            # Validar campos
            if not all([self.combo_producto.get(), self.entry_cantidad.get()]):
                messagebox.showerror("Error", "Completa los campos obligatorios")
                return
            
            # Obtener código del producto
            producto_texto = self.combo_producto.get()
            codigo_producto = producto_texto.split(' - ')[0]
            
            cantidad = int(self.entry_cantidad.get())
            if cantidad <= 0:
                raise ValueError("La cantidad debe ser mayor a 0")
            
            cursor = self.main_window.conn.cursor()
            
            # Obtener ID y saldo actual del producto
            cursor.execute("SELECT id, saldo FROM productos WHERE codigo = ?", (codigo_producto,))
            resultado = cursor.fetchone()
            if not resultado:
                messagebox.showerror("Error", "Producto no encontrado")
                return
            
            producto_id, saldo_actual = resultado
            
            # Validar stock para salidas
            if self.tipo == 'salida' and cantidad > saldo_actual:
                messagebox.showerror("Error", f"Stock insuficiente. Disponible: {saldo_actual}")
                return
            
            # Calcular nuevo saldo
            nuevo_saldo = saldo_actual + cantidad if self.tipo == 'entrada' else saldo_actual - cantidad
            
            # Registrar movimiento
            factura = getattr(self, 'entry_factura', None)
            proveedor = getattr(self, 'entry_proveedor', None)
            destino = getattr(self, 'entry_destino', None)
            
            cursor.execute('''
                INSERT INTO movimientos 
                (producto_id, tipo, cantidad, fecha, factura, proveedor, destino, observaciones)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                producto_id, self.tipo, cantidad, date.today(),
                factura.get() if factura else None,
                proveedor.get() if proveedor else None,
                destino.get() if destino else None,
                self.entry_observaciones.get()
            ))
            
            # Actualizar saldo del producto
            cursor.execute("UPDATE productos SET saldo = ? WHERE id = ?", (nuevo_saldo, producto_id))
            
            self.main_window.conn.commit()
            messagebox.showinfo("Éxito", f"{self.tipo.title()} registrada correctamente")
            
            # Actualizar tabla principal
            self.main_window.cargar_inventario()
            
            # Cerrar ventana
            self.window.destroy()
            
        except ValueError as e:
            messagebox.showerror("Error", f"Cantidad inválida: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al registrar movimiento: {e}")


class AlertasWindow:
    """Ventana para mostrar alertas de stock"""
    
    def __init__(self, parent, main_window):
        self.parent = parent
        self.main_window = main_window
        
        # Crear ventana
        self.window = tk.Toplevel(parent)
        self.window.title("⚠️ Alertas de Stock")
        self.window.geometry("700x500")
        self.window.configure(bg='#f0f0f0')
        
        # Centrar y hacer modal
        self.center_window()
        self.window.transient(parent)
        self.window.grab_set()
        
        # Crear interfaz
        self.create_interface()
        self.cargar_alertas()
    
    def center_window(self):
        """Centrar ventana"""
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (700 // 2)
        y = (self.window.winfo_screenheight() // 2) - (500 // 2)
        self.window.geometry(f"700x500+{x}+{y}")
    
    def create_interface(self):
        """Crear interfaz"""
        # Frame principal
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Título
        title = tk.Label(main_frame, text="⚠️ Alertas de Stock Bajo", 
                        font=('Arial', 20, 'bold'), bg='#f0f0f0', fg='#e74c3c')
        title.grid(row=0, column=0, columnspan=4, pady=(0, 25))
        
        # Tabla de alertas
        alertas_frame = ttk.LabelFrame(main_frame, text="🚨 Productos con Stock Bajo o Agotado", padding="15")
        alertas_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        columns = ('Código', 'Producto', 'Saldo', 'Stock Mín.', 'Estado')
        self.tree_alertas = ttk.Treeview(alertas_frame, columns=columns, show='headings', height=12)
        
        for col in columns:
            self.tree_alertas.heading(col, text=col)
            if col == 'Producto':
                self.tree_alertas.column(col, width=250, anchor='w')
            else:
                self.tree_alertas.column(col, width=100, anchor='center')
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(alertas_frame, orient=tk.VERTICAL, command=self.tree_alertas.yview)
        self.tree_alertas.configure(yscrollcommand=scrollbar.set)
        
        self.tree_alertas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Botones
        btn_frame = tk.Frame(main_frame, bg='#f0f0f0')
        btn_frame.grid(row=2, column=0, columnspan=4, pady=20)
        
        tk.Button(btn_frame, text="🔄 Actualizar", command=self.cargar_alertas,
                 bg="#3498db", fg="white", font=('Arial', 10, 'bold'),
                 relief='flat', padx=20, pady=8, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="❌ Cerrar", command=self.window.destroy,
                 bg="#95a5a6", fg="white", font=('Arial', 10, 'bold'),
                 relief='flat', padx=20, pady=8, cursor='hand2').pack(side=tk.RIGHT, padx=5)
        
        # Configurar grid weights
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        alertas_frame.columnconfigure(0, weight=1)
        alertas_frame.rowconfigure(0, weight=1)
    
    def cargar_alertas(self):
        """Cargar alertas en la tabla"""
        # Limpiar tabla
        for item in self.tree_alertas.get_children():
            self.tree_alertas.delete(item)
        
        cursor = self.main_window.conn.cursor()
        cursor.execute('''
            SELECT codigo, nombre, saldo, stock_minimo 
            FROM productos 
            WHERE saldo <= stock_minimo 
            ORDER BY saldo ASC
        ''')
        
        alertas = cursor.fetchall()
        
        if not alertas:
            # Mostrar mensaje si no hay alertas
            self.tree_alertas.insert('', tk.END, values=(
                "", "¡Sin alertas! Todos los productos tienen stock adecuado", "", "", "🟢 OK"
            ))
            return
        
        for codigo, nombre, saldo, stock_min in alertas:
            if saldo == 0:
                estado = "🔴 AGOTADO"
                tags = ('agotado',)
            else:
                estado = "🟡 STOCK BAJO"
                tags = ('bajo',)
            
            self.tree_alertas.insert('', tk.END, values=(
                codigo, nombre, saldo, stock_min, estado
            ), tags=tags)
        
        # Configurar colores
        self.tree_alertas.tag_configure('agotado', background='#ffe6e6', foreground='#c0392b')
        self.tree_alertas.tag_configure('bajo', background='#fff3e0', foreground='#e67e22')


class EstadisticasWindow:
    """Ventana para mostrar estadísticas detalladas"""
    
    def __init__(self, parent, main_window):
        self.parent = parent
        self.main_window = main_window
        
        # Crear ventana
        self.window = tk.Toplevel(parent)
        self.window.title("📊 Estadísticas Detalladas")
        self.window.geometry("800x600")
        self.window.configure(bg='#f0f0f0')
        
        # Centrar y hacer modal
        self.center_window()
        self.window.transient(parent)
        self.window.grab_set()
        
        # Crear interfaz
        self.create_interface()
        self.cargar_estadisticas()
    
    def center_window(self):
        """Centrar ventana"""
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (800 // 2)
        y = (self.window.winfo_screenheight() // 2) - (600 // 2)
        self.window.geometry(f"800x600+{x}+{y}")
    
    def create_interface(self):
        """Crear interfaz de estadísticas"""
        # Frame principal
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Título
        title = tk.Label(main_frame, text="📊 Estadísticas Detalladas del Inventario", 
                        font=('Arial', 18, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        title.grid(row=0, column=0, columnspan=2, pady=(0, 25))
        
        # Frame izquierdo - Estadísticas generales
        stats_frame = ttk.LabelFrame(main_frame, text="📈 Resumen General", padding="15")
        stats_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        
        self.stats_text = tk.Text(stats_frame, width=35, height=20, 
                                 font=('Consolas', 10), bg='#ffffff',
                                 relief='solid', bd=1)
        self.stats_text.pack(fill=tk.BOTH, expand=True)
        
        # Frame derecho - Estadísticas por categoría
        category_frame = ttk.LabelFrame(main_frame, text="📦 Por Categorías", padding="15")
        category_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(10, 0))
        
        # Tabla para categorías
        columns = ('Categoría', 'Productos', 'Valor Total', 'Stock Bajo', 'Sin Stock')
        self.tree_categories = ttk.Treeview(category_frame, columns=columns, 
                                          show='headings', height=8)
        
        for col in columns:
            self.tree_categories.heading(col, text=col)
            if col == 'Categoría':
                self.tree_categories.column(col, width=100, anchor='w')
            else:
                self.tree_categories.column(col, width=80, anchor='center')
        
        self.tree_categories.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Top 10 productos más valiosos
        top_frame = ttk.LabelFrame(category_frame, text="💰 Top 10 Más Valiosos", padding="10")
        top_frame.pack(fill=tk.BOTH, expand=True)
        
        self.top_text = tk.Text(top_frame, width=35, height=8,
                               font=('Consolas', 9), bg='#ffffff',
                               relief='solid', bd=1)
        self.top_text.pack(fill=tk.BOTH, expand=True)
        
        # Botones
        btn_frame = tk.Frame(main_frame, bg='#f0f0f0')
        btn_frame.grid(row=2, column=0, columnspan=2, pady=20)
        
        tk.Button(btn_frame, text="🔄 Actualizar", command=self.cargar_estadisticas,
                 bg="#3498db", fg="white", font=('Arial', 10, 'bold'),
                 relief='flat', padx=20, pady=8, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="📊 Exportar", command=self.exportar_estadisticas,
                 bg="#27ae60", fg="white", font=('Arial', 10, 'bold'),
                 relief='flat', padx=20, pady=8, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="❌ Cerrar", command=self.window.destroy,
                 bg="#95a5a6", fg="white", font=('Arial', 10, 'bold'),
                 relief='flat', padx=20, pady=8, cursor='hand2').pack(side=tk.RIGHT, padx=5)
        
        # Configurar grid weights
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
    
    def cargar_estadisticas(self):
        """Cargar todas las estadísticas"""
        cursor = self.main_window.conn.cursor()
        
        # Limpiar contenido
        self.stats_text.delete(1.0, tk.END)
        for item in self.tree_categories.get_children():
            self.tree_categories.delete(item)
        self.top_text.delete(1.0, tk.END)
        
        # === ESTADÍSTICAS GENERALES ===
        stats_content = "=== RESUMEN GENERAL ===\n\n"
        
        # Total productos
        cursor.execute("SELECT COUNT(*) FROM productos")
        total_productos = cursor.fetchone()[0]
        stats_content += f"📦 Total Productos: {total_productos}\n\n"
        
        # Productos por estado
        cursor.execute("SELECT COUNT(*) FROM productos WHERE saldo > stock_minimo")
        normal = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM productos WHERE saldo <= stock_minimo AND saldo > 0")
        bajo = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM productos WHERE saldo = 0")
        agotado = cursor.fetchone()[0]
        
        stats_content += "=== ESTADO DEL STOCK ===\n"
        stats_content += f"🟢 Stock Normal: {normal} ({normal/total_productos*100:.1f}%)\n"
        stats_content += f"🟡 Stock Bajo: {bajo} ({bajo/total_productos*100:.1f}%)\n"
        stats_content += f"🔴 Agotado: {agotado} ({agotado/total_productos*100:.1f}%)\n\n"
        
        # Valor total del inventario
        cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos")
        valor_total = cursor.fetchone()[0] or 0
        stats_content += f"💰 Valor Total: ${valor_total:,.0f}\n\n"
        
        # Productos con mayor rotación (simulado)
        cursor.execute("""
            SELECT nombre, saldo, valor_unitario, (saldo * valor_unitario) as valor_total
            FROM productos 
            WHERE saldo > 0
            ORDER BY valor_total DESC
            LIMIT 5
        """)
        productos_valor = cursor.fetchall()
        
        stats_content += "=== TOP 5 POR VALOR ===\n"
        for i, (nombre, saldo, valor_unit, valor_total) in enumerate(productos_valor, 1):
            nombre_corto = (nombre[:25] + '...') if len(nombre) > 25 else nombre
            stats_content += f"{i}. {nombre_corto}\n"
            stats_content += f"   ${valor_total:,.0f} ({saldo} x ${valor_unit:,.0f})\n"
        
        # Unidades más comunes
        stats_content += "\n=== UNIDADES DE MEDIDA ===\n"
        cursor.execute("SELECT unidad, COUNT(*) as count FROM productos GROUP BY unidad ORDER BY count DESC")
        unidades = cursor.fetchall()
        for unidad, count in unidades:
            stats_content += f"{unidad}: {count} productos\n"
        
        self.stats_text.insert(tk.END, stats_content)
        
        # === ESTADÍSTICAS POR CATEGORÍA ===
        categorias = ['almacen', 'postcosecha', 'quimicos']
        for categoria in categorias:
            cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria = ?", (categoria,))
            total = cursor.fetchone()[0]
            
            cursor.execute("SELECT SUM(saldo * valor_unitario) FROM productos WHERE categoria = ?", (categoria,))
            valor = cursor.fetchone()[0] or 0
            
            cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria = ? AND saldo <= stock_minimo AND saldo > 0", (categoria,))
            bajo = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria = ? AND saldo = 0", (categoria,))
            sin_stock = cursor.fetchone()[0]
            
            self.tree_categories.insert('', tk.END, values=(
                categoria.upper(), total, f"${valor:,.0f}", bajo, sin_stock
            ))
        
        # === TOP 10 MÁS VALIOSOS ===
        cursor.execute("""
            SELECT nombre, categoria, saldo, valor_unitario, (saldo * valor_unitario) as valor_total
            FROM productos 
            WHERE saldo > 0
            ORDER BY valor_total DESC
            LIMIT 10
        """)
        top_productos = cursor.fetchall()
        
        top_content = "Producto - Valor Total\n" + "="*40 + "\n"
        for i, (nombre, categoria, saldo, valor_unit, valor_total) in enumerate(top_productos, 1):
            nombre_corto = (nombre[:22] + '...') if len(nombre) > 22 else nombre
            top_content += f"{i:2d}. {nombre_corto}\n"
            top_content += f"    [{categoria.upper()}] ${valor_total:,.0f}\n"
        
        self.top_text.insert(tk.END, top_content)
    
    def exportar_estadisticas(self):
        """Exportar estadísticas a archivo de texto"""
        try:
            archivo = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")],
                initialvalue=f"estadisticas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            
            if archivo:
                with open(archivo, 'w', encoding='utf-8') as f:
                    f.write("ESTADÍSTICAS DETALLADAS DEL INVENTARIO\n")
                    f.write(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                    f.write("="*50 + "\n\n")
                    f.write(self.stats_text.get(1.0, tk.END))
                    f.write("\n" + "="*50 + "\n")
                    f.write("TOP 10 PRODUCTOS MÁS VALIOSOS\n")
                    f.write("="*50 + "\n")
                    f.write(self.top_text.get(1.0, tk.END))
                
                messagebox.showinfo("Éxito", f"Estadísticas exportadas: {archivo}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")