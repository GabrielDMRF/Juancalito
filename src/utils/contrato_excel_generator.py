#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Instalación y Configuración
Sistema de Generación Automática de Contratos Excel
"""

import os
import sys
import shutil
import subprocess
from pathlib import Path
from datetime import datetime

def print_header():
    """Mostrar header del instalador"""
    print("=" * 70)
    print("🚀 INSTALADOR DE GENERADOR DE CONTRATOS EXCEL")
    print("=" * 70)
    print("Este script configurará automáticamente el sistema para generar")
    print("contratos Excel usando la plantilla 'CONTRATO EXCEL FLORE JUNCALITO.xlsx'")
    print("=" * 70)
    print()

def check_dependencies():
    """Verificar e instalar dependencias"""
    print("📦 Verificando dependencias...")
    
    # Lista de dependencias requeridas
    dependencies = [
        'openpyxl',
        'python-docx',
        'sqlalchemy',
        'pillow'
    ]
    
    missing_deps = []
    
    for dep in dependencies:
        try:
            __import__(dep.replace('-', '_'))
            print(f"  {dep} - OK")
        except ImportError:
            print(f"  ❌ {dep} - FALTA")
            missing_deps.append(dep)
    
    if missing_deps:
        print(f"\n🔧 Instalando dependencias faltantes: {', '.join(missing_deps)}")
        try:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install'] + missing_deps)
            print("Dependencias instaladas exitosamente")
        except subprocess.CalledProcessError as e:
            print(f"❌ Error instalando dependencias: {e}")
            return False
    
    return True

def create_directory_structure():
    """Crear estructura de directorios necesaria"""
    print("\n📁 Creando estructura de directorios...")
    
    directories = [
        'src/utils',
        'src/views',
        'plantillas',
        'empleados_data',
        'database'
    ]
    
    for directory in directories:
        path = Path(directory)
        path.mkdir(parents=True, exist_ok=True)
        print(f"  {directory}")
    
    return True

def copy_template_if_exists():
    """Buscar y copiar plantilla si existe en el directorio actual"""
    print("\n📋 Buscando plantilla de contrato...")
    
    # Posibles nombres de la plantilla
    template_names = [
        'CONTRATO EXCEL FLORE JUNCALITO.xlsx',
        'contrato_excel_flore_juncalito.xlsx',
        'plantilla_contrato.xlsx',
        'contrato_plantilla.xlsx'
    ]
    
    plantillas_dir = Path('plantillas')
    found_template = False
    
    for template_name in template_names:
        if Path(template_name).exists():
            destination = plantillas_dir / 'CONTRATO EXCEL FLORE JUNCALITO.xlsx'
            shutil.copy2(template_name, destination)
            print(f"  Plantilla copiada: {template_name} -> {destination}")
            found_template = True
            break
    
    if not found_template:
        print("  ⚠️  Plantilla no encontrada en el directorio actual")
        print("     Coloca manualmente 'CONTRATO EXCEL FLORE JUNCALITO.xlsx' en la carpeta 'plantillas/'")
    
    return found_template

def create_sample_data():
    """Crear datos de ejemplo para pruebas"""
    print("\n🧪 Verificando datos de prueba...")
    
    try:
        # Importar modelos
        sys.path.append('src')
        from models.database import create_tables, get_db, Empleado, Contrato
        
        # Crear tablas
        create_tables()
        
        db = get_db()
        
        # Verificar si ya hay empleados
        empleados_count = db.query(Empleado).count()
        contratos_count = db.query(Contrato).count()
        
        print(f"  📊 Empleados existentes: {empleados_count}")
        print(f"  📊 Contratos existentes: {contratos_count}")
        
        if empleados_count == 0:
            print("  🔧 Creando empleado de ejemplo...")
            empleado_ejemplo = Empleado(
                nombre_completo="Juan Pérez García",
                cedula="12345678",
                telefono="3001234567",
                email="juan.perez@empresa.com",
                direccion="Calle 123 #45-67",
                area_trabajo="produccion",
                cargo="operario",
                salario_base=1423500,
                estado=True
            )
            db.add(empleado_ejemplo)
            db.commit()
            print("  Empleado de ejemplo creado")
        
        if contratos_count == 0 and empleados_count > 0:
            print("  🔧 Creando contrato de ejemplo...")
            empleado = db.query(Empleado).first()
            
            contrato_ejemplo = Contrato(
                numero_contrato="CT-2024-001",
                empleado_id=empleado.id,
                tipo_contrato_id=1,  # Asumiendo que 1 es el ID para "temporal"
                fecha_inicio=datetime.now().date(),
                salario_base=1423500,
                subsidio_transporte=140606,
                estado="borrador"
            )
            db.add(contrato_ejemplo)
            db.commit()
            print("  Contrato de ejemplo creado")
        
        return True
        
    except Exception as e:
        print(f"  ❌ Error creando datos de ejemplo: {e}")
        return False

def create_test_script():
    """Crear script de prueba"""
    print("\n🧪 Creando script de prueba...")
    
    test_script = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Prueba - Generador de Contratos Excel
"""

import sys
import os
sys.path.append('src')

def test_excel_generator():
    """Probar generador de Excel"""
    try:
        print("🧪 Probando generador de contratos Excel...")
        
        # Importar generador
        from utils.contrato_excel_generator import ContratoExcelGenerator
        from models.database import get_db, Empleado, Contrato
        
        # Obtener datos de prueba
        db = get_db()
        empleado = db.query(Empleado).first()
        contrato = db.query(Contrato).first()
        
        if not empleado or not contrato:
            print("❌ No hay datos de prueba. Ejecuta el instalador primero.")
            return False
        
        # Probar generador
        generator = ContratoExcelGenerator()
        result = generator.generar_contrato_excel(contrato, empleado)
        
        if result['success']:
            print(f"Contrato Excel generado exitosamente:")
            print(f"   Archivo: {result['file_path']}")
            return True
        else:
            print(f"❌ Error generando contrato: {result['message']}")
            return False
            
    except Exception as e:
        print(f"❌ Error en prueba: {e}")
        return False

def test_integration():
    """Probar integración completa"""
    try:
        print("🔧 Probando integración con sistema principal...")
        
        # Importar sistema principal
        from views.main_window import MainWindow
        import tkinter as tk
        
        print("Sistema principal importado correctamente")
        
        # Importar vista de contratos
        from views.contratos_view import ContratosWindow
        print("Vista de contratos importada correctamente")
        
        return True
        
    except Exception as e:
        print(f"❌ Error en integración: {e}")
        return False

if __name__ == "__main__":
    print("🚀 PROBADOR DEL SISTEMA DE CONTRATOS EXCEL")
    print("=" * 50)
    
    success = True
    
    if test_excel_generator():
        print("Generador de Excel: OK")
    else:
        print("❌ Generador de Excel: FALLO")
        success = False
    
    if test_integration():
        print("Integración: OK")
    else:
        print("❌ Integración: FALLO")
        success = False
    
    if success:
        print("\\n🎉 ¡TODOS LOS TESTS PASARON!")
        print("El sistema está listo para generar contratos Excel")
    else:
        print("\\n❌ Algunos tests fallaron")
        print("Revisa la configuración y dependencias")
'''
    
    with open('test_contratos_excel.py', 'w', encoding='utf-8') as f:
        f.write(test_script)
    
    print("  Script de prueba creado: test_contratos_excel.py")
    return True

def create_readme():
    """Crear documentación README"""
    print("\n📚 Creando documentación...")
    
    readme_content = '''# Generador Automático de Contratos Excel

## 🎯 Descripción
Sistema integrado para generar contratos laborales en formato Excel usando plantillas personalizadas.

## 📋 Características
- Generación automática de contratos Excel
- Relleno automático de todos los campos
- Integración completa con el sistema existente
- Guardado organizado en carpetas por empleado
- Interfaz profesional y fácil de usar

## 🚀 Instalación

### 1. Dependencias
```bash
pip install openpyxl python-docx sqlalchemy pillow
```

### 2. Estructura de Archivos
```
proyecto/
├── src/
│   ├── utils/
│   │   └── contrato_excel_generator.py
│   └── views/
│       └── contratos_view.py
├── plantillas/
│   └── CONTRATO EXCEL FLORE JUNCALITO.xlsx
└── empleados_data/
    └── {empleado}/
        └── contratos/
```

### 3. Plantilla
Coloca tu plantilla Excel como:
`plantillas/CONTRATO EXCEL FLORE JUNCALITO.xlsx`

## 🎮 Uso

1. Ejecutar sistema principal: `python src/main.py`
2. Ir a "Gestión de Contratos"
3. Seleccionar un contrato
4. Hacer clic en "📊 Generar Excel"
5. ¡Listo! El contrato se genera automáticamente

## 🔧 Pruebas
```bash
python test_contratos_excel.py
```

## 📁 Archivos Generados
Los contratos se guardan en:
`empleados_data/{nombre_empleado}_{cedula}/contratos/contrato_excel_{cedula}_{timestamp}.xlsx`

## 🆘 Solución de Problemas

### Plantilla no encontrada
- Verifica que `CONTRATO EXCEL FLORE JUNCALITO.xlsx` esté en `plantillas/`
- Usa el botón "📁 Seleccionar Plantilla" en la interfaz

### Error de dependencias
```bash
pip install --upgrade openpyxl python-docx
```

### Error de permisos
- Verifica permisos de escritura en `empleados_data/`
- Ejecuta como administrador si es necesario

## 🎉 ¡Sistema Listo!
Tu sistema ahora puede generar contratos Excel automáticamente.
'''
    
    with open('README_CONTRATOS_EXCEL.md', 'w', encoding='utf-8') as f:
        f.write(readme_content)
    
    print("  Documentación creada: README_CONTRATOS_EXCEL.md")
    return True

def main():
    """Función principal del instalador"""
    print_header()
    
    success_steps = []
    
    # Paso 1: Verificar dependencias
    if check_dependencies():
        success_steps.append("Dependencias")
    else:
        success_steps.append("❌ Dependencias")
    
    # Paso 2: Crear estructura
    if create_directory_structure():
        success_steps.append("Directorios")
    else:
        success_steps.append("❌ Directorios")
    
    # Paso 3: Buscar plantilla
    if copy_template_if_exists():
        success_steps.append("Plantilla")
    else:
        success_steps.append("⚠️ Plantilla")
    
    # Paso 4: Datos de ejemplo
    if create_sample_data():
        success_steps.append("Datos de ejemplo")
    else:
        success_steps.append("⚠️ Datos de ejemplo")
    
    # Paso 5: Script de prueba
    if create_test_script():
        success_steps.append("Script de prueba")
    else:
        success_steps.append("❌ Script de prueba")
    
    # Paso 6: Documentación
    if create_readme():
        success_steps.append("Documentación")
    else:
        success_steps.append("❌ Documentación")
    
    # Mostrar resumen
    print("\n" + "=" * 70)
    print("📊 RESUMEN DE INSTALACIÓN")
    print("=" * 70)
    for step in success_steps:
        print(f"  {step}")
    
    print("\n🎯 PRÓXIMOS PASOS:")
    print("1. Copia el código del generador a: src/utils/contrato_excel_generator.py")
    print("2. Reemplaza src/views/contratos_view.py con la versión integrada")
    print("3. Coloca tu plantilla Excel en: plantillas/CONTRATO EXCEL FLORE JUNCALITO.xlsx")
    print("4. Ejecuta: python test_contratos_excel.py")
    print("5. ¡Disfruta generando contratos automáticamente!")
    
    print("\n🎉 ¡INSTALACIÓN COMPLETADA!")
    print("=" * 70)

if __name__ == "__main__":
    main()

def abrir_generador_contratos_excel(parent, contratos_window, contrato):
    """Abrir ventana del generador de contratos Excel"""
    import tkinter as tk
    from tkinter import messagebox, filedialog
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    import os
    import sys
    
    # Agregar path para importar modelos
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from models.database import Empleado
    
    def generar_contrato_excel():
        """Generar contrato Excel usando la plantilla original"""
        try:
            # Obtener empleado
            empleado = contratos_window.db.query(Empleado).filter(Empleado.id == contrato.empleado_id).first()
            if not empleado:
                messagebox.showerror("Error", "Empleado no encontrado")
                return
            
            # Ruta de la plantilla
            template_path = "templates_contratos/CONTRATO EXCEL FLORE JUNCALITO.xlsx"
            if not os.path.exists(template_path):
                messagebox.showerror("Error", f"Plantilla no encontrada: {template_path}")
                return
            
            # Cargar la plantilla
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Función para convertir fechas a español
            def fecha_a_espanol(fecha):
                """Convertir fecha a formato español"""
                if not fecha:
                    return "NO DEFINIDA"
                
                meses_espanol = {
                    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
                }
                
                dia = fecha.day
                mes = meses_espanol[fecha.month]
                anio = fecha.year
                
                return f"{dia} DE {mes} DE {anio}"
            
            # Función para convertir números a letras (versión mejorada)
            def numero_a_letras(numero):
                """Convertir número a letras"""
                if numero is None or numero == 0:
                    return "CERO PESOS"
                
                # Convertir a entero
                numero = int(numero)
                
                # Diccionarios para conversión
                unidades = ["", "UNO", "DOS", "TRES", "CUATRO", "CINCO", "SEIS", "SIETE", "OCHO", "NUEVE"]
                decenas = ["", "DIEZ", "VEINTE", "TREINTA", "CUARENTA", "CINCUENTA", "SESENTA", "SETENTA", "OCHENTA", "NOVENTA"]
                centenas = ["", "CIENTO", "DOSCIENTOS", "TRESCIENTOS", "CUATROCIENTOS", "QUINIENTOS", "SEISCIENTOS", "SETECIENTOS", "OCHOCIENTOS", "NOVECIENTOS"]
                
                if numero == 0:
                    return "CERO PESOS"
                elif numero == 100:
                    return "CIEN PESOS"
                elif numero < 1000:
                    # Manejar números menores a 1000
                    if numero < 100:
                        if numero < 20:
                            if numero == 11:
                                return "ONCE PESOS"
                            elif numero == 12:
                                return "DOCE PESOS"
                            elif numero == 13:
                                return "TRECE PESOS"
                            elif numero == 14:
                                return "CATORCE PESOS"
                            elif numero == 15:
                                return "QUINCE PESOS"
                            elif numero == 16:
                                return "DIECISÉIS PESOS"
                            elif numero == 17:
                                return "DIECISIETE PESOS"
                            elif numero == 18:
                                return "DIECIOCHO PESOS"
                            elif numero == 19:
                                return "DIECINUEVE PESOS"
                            else:
                                return unidades[numero] + " PESOS"
                        else:
                            if numero % 10 == 0:
                                return decenas[numero // 10] + " PESOS"
                            else:
                                return decenas[numero // 10] + " Y " + unidades[numero % 10] + " PESOS"
                    else:
                        # Números entre 100 y 999
                        centena = numero // 100
                        resto = numero % 100
                        
                        if resto == 0:
                            return centenas[centena] + " PESOS"
                        else:
                            return centenas[centena] + " " + numero_a_letras(resto).replace(" PESOS", "") + " PESOS"
                else:
                    # Para números grandes, usar formato simplificado
                    return f"{numero:,} PESOS"
            
            # Reemplazar variables con llaves en toda la plantilla
            print(f"Buscando y reemplazando variables en la plantilla...")
            
            # Diccionario de variables disponibles
            variables = {
                # Datos del empleado
                '{NOMBRE_EMPLEADO}': empleado.nombre_completo or "NO DEFINIDO",
                '{CEDULA_EMPLEADO}': empleado.cedula or "NO DEFINIDO",
                '{DIRECCION_EMPLEADO}': empleado.direccion or "DIRECCIÓN NO ESPECIFICADA",
                '{TELEFONO_EMPLEADO}': empleado.telefono or "NO DEFINIDO",
                '{EMAIL_EMPLEADO}': empleado.email or "NO DEFINIDO",
                '{CARGO_EMPLEADO}': empleado.cargo or "OPERARIO OFICIOS VARIOS",
                '{AREA_TRABAJO}': empleado.area_trabajo or "NO DEFINIDO",
                
                # Datos del contrato
                '{NUMERO_CONTRATO}': contrato.numero_contrato or "NO DEFINIDO",
                '{TIPO_CONTRATO}': contrato.tipo_contrato_rel.nombre if contrato.tipo_contrato_rel else "NO DEFINIDO",
                '{SALARIO_NUMEROS}': f"$ {contrato.salario_contrato:,}" if contrato.salario_contrato else "NO DEFINIDO",
                '{SALARIO_LETRAS}': numero_a_letras(contrato.salario_contrato) if contrato.salario_contrato else "NO DEFINIDO",
                '{FECHA_INICIO}': fecha_a_espanol(contrato.fecha_inicio) if contrato.fecha_inicio else "NO DEFINIDA",
                '{FECHA_FIN}': fecha_a_espanol(contrato.fecha_fin) if contrato.fecha_fin else "NO DEFINIDA",
                '{ESTADO_CONTRATO}': contrato.estado or "NO DEFINIDO",
                
                # Fechas del sistema
                '{FECHA_GENERACION}': datetime.now().strftime('%d/%m/%Y'),
                '{FECHA_ACTUAL}': datetime.now().strftime('%d/%m/%Y'),
                '{HORA_GENERACION}': datetime.now().strftime('%H:%M:%S'),
                
                # Datos de la empresa (puedes personalizar)
                '{NOMBRE_EMPRESA}': "FLORES JUNCALITO S.A.S",
                '{DIRECCION_EMPRESA}': "CALLE 19* C N. 88-07",
                '{CIUDAD_EMPRESA}': "EL ROSAL CUNDINAMARCA",
                
                # NUEVAS VARIABLES BASADAS EN LA PLANTILLA
                '{NOMBRE_EMPLEADOR}': "FLORES JUNCALITO S.A.S",
                '{DIRECCION_EMPLEADOR}': "CALLE 19* C N. 88-07",
                '{LUGAR_NACIMIENTO}': empleado.lugar_nacimiento or "BOGOTÁ, COLOMBIA",
                '{FECHA_NACIMIENTO}': fecha_a_espanol(empleado.fecha_nacimiento) if empleado.fecha_nacimiento else "01 DE ENERO DE 1990",
                '{NACIONALIDAD}': empleado.nacionalidad or "COLOMBIANA",
                '{TIPO_SALARIO}': "ORDINARIO",
                '{PERIODOS_PAGO}': "MENSUALES",
                '{LUGAR_LABORES}': "FLORES JUNCALITO S.A.S",
                '{CIUDAD_CONTRATACION}': "BARRIO SAN JOSE - EL ROSAL CUNDINAMARCA",
                '{TIPO_TERMINO_CONTRATO}': "FIJO",
                
                # Variables adicionales para flexibilidad
                '{SALARIO_CONTRATO}': f"$ {contrato.salario_contrato:,}" if contrato.salario_contrato else "NO DEFINIDO",
                '{FECHA_INICIO_LABORES}': fecha_a_espanol(contrato.fecha_inicio) if contrato.fecha_inicio else "NO DEFINIDA",
                '{VENCE_EL_DIA}': fecha_a_espanol(contrato.fecha_fin) if contrato.fecha_fin else "NO DEFINIDA"
            }
            
            # Buscar y reemplazar variables en todas las celdas
            variables_encontradas = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        valor_original = cell.value
                        valor_nuevo = valor_original
                        
                        # Reemplazar todas las variables encontradas
                        for variable, valor in variables.items():
                            if variable in valor_original:
                                valor_nuevo = valor_nuevo.replace(variable, str(valor))
                                variables_encontradas += 1
                                print(f"   Reemplazado: {variable} -> {valor}")
                        
                        # Actualizar la celda si hubo cambios
                        if valor_nuevo != valor_original:
                            cell.value = valor_nuevo
            
            print(f"   📊 Total de variables reemplazadas: {variables_encontradas}")
            
            # También reemplazar referencias directas a celdas específicas (como H24, K24)
            print(f"🔄 Reemplazando referencias directas a celdas...")
            
            # Mapeo de celdas específicas basado en las referencias que mencionaste
            celdas_especificas = {
                'H24': numero_a_letras(contrato.salario_contrato) if contrato.salario_contrato else "NO DEFINIDO",
                'K24': contrato.fecha_fin.strftime('%d/%m/%Y') if contrato.fecha_fin else "NO DEFINIDA",
                'E24': empleado.nombre_completo or "NO DEFINIDO",
                'F24': empleado.cedula or "NO DEFINIDO",
                'G24': str(contrato.salario_contrato) if contrato.salario_contrato else "NO DEFINIDO",
                'I24': contrato.fecha_inicio.strftime('%d/%m/%Y') if contrato.fecha_inicio else "NO DEFINIDA",
                'L24': empleado.cargo or "NO DEFINIDO",
                'M24': empleado.direccion or "NO DEFINIDO"
            }
            
            # Buscar y reemplazar referencias a BASE DE DATOS
            referencias_encontradas = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value)
                        
                        # Buscar referencias como ='BASE DE DATOS'!H24
                        if "'BASE DE DATOS'!" in cell_value:
                            # Extraer la referencia de celda (ej: H24, K24)
                            import re
                            match = re.search(r"'BASE DE DATOS'!([A-Z]+\d+)", cell_value)
                            if match:
                                celda_ref = match.group(1)
                                if celda_ref in celdas_especificas:
                                    cell.value = celdas_especificas[celda_ref]
                                    referencias_encontradas += 1
                                    print(f"   Reemplazada referencia {celda_ref} -> {celdas_especificas[celda_ref]}")
            
            print(f"   Total de referencias reemplazadas: {referencias_encontradas}")
            
            if variables_encontradas == 0 and referencias_encontradas == 0:
                print("   No se encontraron variables ni referencias para reemplazar")
                print("   Sugerencia: Agrega variables como {NOMBRE_EMPLEADO} o referencias a BASE DE DATOS")
            
            # Función para limpiar nombres de archivo
            def limpiar_nombre_archivo(texto):
                if not texto:
                    return "sin_nombre"
                # Remover caracteres problemáticos para nombres de archivo
                caracteres_prohibidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
                texto_limpio = str(texto)
                for char in caracteres_prohibidos:
                    texto_limpio = texto_limpio.replace(char, '_')
                texto_limpio = texto_limpio.replace(" ", "_")
                return texto_limpio
            
            # Crear carpeta del empleado si no existe
            nombre_seguro = limpiar_nombre_archivo(empleado.nombre_completo)
            cedula_segura = limpiar_nombre_archivo(empleado.cedula)
            empleado_dir = os.path.join("empleados_data", f"{nombre_seguro}_{cedula_segura}")
            contratos_dir = os.path.join(empleado_dir, "contratos")
            
            os.makedirs(contratos_dir, exist_ok=True)
            
            # Guardar archivo en la carpeta del empleado
            filename = f"contrato_{limpiar_nombre_archivo(empleado.cedula)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(contratos_dir, filename)
            
            wb.save(filepath)
            wb.close()
            
            # Actualizar contrato en base de datos
            contrato.archivo_path = filepath
            contratos_window.db.commit()
            
            messagebox.showinfo("Éxito", f"Contrato Excel generado exitosamente:\n{filepath}")
            
            # Actualizar lista de contratos
            contratos_window.cargar_contratos()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generando contrato Excel: {e}")
            print(f"Error generando contrato Excel: {e}")
    
    # Crear ventana del generador
    window = tk.Toplevel(parent)
    window.title("Generador de Contrato Excel")
    window.geometry("400x300")
    window.configure(bg='#f8f9fa')
    
    # Centrar ventana
    window.update_idletasks()
    x = (window.winfo_screenwidth() // 2) - (400 // 2)
    y = (window.winfo_screenheight() // 2) - (300 // 2)
    window.geometry(f"400x300+{x}+{y}")
    
    # Contenido de la ventana
    tk.Label(window, text="Generador de Contrato Excel", 
            font=('Segoe UI', 16, 'bold'), bg='#f8f9fa', fg='#2c3e50').pack(pady=20)
    
    tk.Label(window, text=f"Empleado: {contrato.empleado.nombre_completo if contrato.empleado else 'No encontrado'}", 
            font=('Segoe UI', 12), bg='#f8f9fa').pack(pady=10)
    
    tk.Label(window, text=f"Contrato: {contrato.numero_contrato or 'Sin número'}", 
            font=('Segoe UI', 12), bg='#f8f9fa').pack(pady=10)
    
    tk.Button(window, text="Generar Contrato Excel", command=generar_contrato_excel,
             bg="#27ae60", fg="white", font=('Segoe UI', 12, 'bold'),
             relief='flat', bd=0, padx=20, pady=10, cursor='hand2').pack(pady=20)
    
    tk.Button(window, text="Cerrar", command=window.destroy,
             bg="#e74c3c", fg="white", font=('Segoe UI', 12, 'bold'),
             relief='flat', bd=0, padx=20, pady=10, cursor='hand2').pack(pady=10)