import os
from datetime import datetime
from tkinter import messagebox

# Intentar importar openpyxl y estilos necesarios
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    messagebox.showerror("Error", 
        "Falta la librería 'openpyxl'.\nInstálala usando: pip install openpyxl")
    raise

def safe_filename(nombre):
    """Convertir una cadena en un nombre de archivo seguro (sin caracteres inválidos)"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        nombre = nombre.replace(char, '_')
    # Permitir solo alfanuméricos, espacios, guiones y guiones bajos
    nombre = ''.join(c for c in nombre if c.isalnum() or c in (' ', '-', '_')).strip()
    # Reemplazar espacios por guiones bajos
    nombre = nombre.replace(' ', '_')
    # Limitar longitud máxima
    return nombre[:50]

def auto_generate_employee_excel(empleado):
    """
    Genera automáticamente el archivo Excel personal de un empleado, creando 
    la carpeta del empleado y subcarpetas (contratos, documentos, nomina, historiales).
    El Excel incluye datos básicos del empleado y varias hojas para seguimiento.
    """
    try:
        # Definir ruta base y carpeta del empleado
        base_dir = os.path.join(os.getcwd(), "empleados_data")
        nombre_seguro = safe_filename(empleado.nombre_completo)
        cedula_segura = safe_filename(empleado.cedula)
        folder_name = f"{nombre_seguro}_{cedula_segura}"
        employee_folder = os.path.join(base_dir, folder_name)
        # Crear carpeta del empleado y subcarpetas necesarias
        os.makedirs(employee_folder, exist_ok=True)
        for sub in ["contratos", "documentos", "nomina", "historiales"]:
            os.makedirs(os.path.join(employee_folder, sub), exist_ok=True)
        # Crear libro de Excel nuevo
        wb = openpyxl.Workbook()
        # ===== HOJA 1: DATOS PERSONALES =====
        ws1 = wb.active
        ws1.title = "Datos Personales"
        # Estilos comunes
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(size=18, bold=True, color="2F4F4F")
        data_font = Font(size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        # Función local para limpiar texto (eliminar caracteres problemáticos)
        def limpiar_texto(texto):
            if texto is None:
                return "No definido"
            texto_str = str(texto)
            texto_limpio = texto_str.encode('ascii', 'ignore').decode('ascii')
            if not texto_limpio.strip():
                return "No definido"
            return texto_limpio.strip()
        # Título principal
        ws1.merge_cells('A1:D1')
        ws1['A1'] = "INFORMACION DEL EMPLEADO"
        ws1['A1'].font = title_font
        ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # Fecha de generación del documento
        ws1.merge_cells('A2:D2')
        ws1['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws1['A2'].alignment = Alignment(horizontal='center')
        ws1['A2'].font = Font(size=10, italic=True)
        # Datos personales y laborales
        row = 4
        datos_personales = [
            ("DATOS BASICOS", ""),
            ("Nombre Completo:", limpiar_texto(empleado.nombre_completo)),
            ("Cedula:", limpiar_texto(empleado.cedula)),
            ("Telefono:", limpiar_texto(empleado.telefono)),
            ("Email:", limpiar_texto(empleado.email)),
            ("Direccion:", limpiar_texto(empleado.direccion)),
            ("", ""),  # Línea en blanco
            ("DATOS LABORALES", ""),
            ("Area de Trabajo:", limpiar_texto(empleado.area_trabajo)),
            ("Cargo:", limpiar_texto(empleado.cargo)),
            ("Salario Base:", f"${empleado.salario_base:,}" if empleado.salario_base else "No definido"),
            ("Estado:", "Activo" if empleado.estado else "Inactivo"),
            ("Fecha de Ingreso:", empleado.fecha_creacion.strftime('%d/%m/%Y') if getattr(empleado, 'fecha_creacion', None) else "No definida")
        ]
        for etiqueta, valor in datos_personales:
            if etiqueta in ["DATOS BASICOS", "DATOS LABORALES"]:
                # Encabezados de sección
                ws1.merge_cells(f'A{row}:D{row}')
                ws1[f'A{row}'] = limpiar_texto(etiqueta)
                ws1[f'A{row}'].font = header_font
                ws1[f'A{row}'].fill = header_fill
                ws1[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
                ws1[f'A{row}'].border = border
            elif etiqueta == "":
                # Línea en blanco, no se escribe nada
                pass
            else:
                # Etiqueta y valor de dato personal
                ws1[f'A{row}'] = limpiar_texto(etiqueta)
                ws1[f'B{row}'] = limpiar_texto(valor)
                ws1[f'A{row}'].font = Font(size=11, bold=True)
                ws1[f'B{row}'].font = data_font
                ws1[f'A{row}'].border = border
                ws1[f'B{row}'].border = border
            row += 1
        # Ajustar ancho de columnas en hoja 1
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 15
        # ===== HOJA 2: SEGUIMIENTO =====
        ws2 = wb.create_sheet("Seguimiento")
        ws2.merge_cells('A1:F1')
        ws2['A1'] = "SEGUIMIENTO Y NOTAS"
        ws2['A1'].font = title_font
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # Encabezados de la tabla de seguimiento
        headers_seguimiento = ['Fecha', 'Tipo', 'Descripcion', 'Usuario', 'Estado', 'Observaciones']
        for col_index, header in enumerate(headers_seguimiento, start=1):
            cell = ws2.cell(row=3, column=col_index, value=limpiar_texto(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Fila de ejemplo en seguimiento
        ejemplo_seguimiento = [
            datetime.now().strftime('%d/%m/%Y'),
            'CREACION',
            'Empleado registrado en el sistema',
            'Sistema',
            'COMPLETADO',
            'Registro inicial automático'
        ]
        for col_index, valor in enumerate(ejemplo_seguimiento, start=1):
            cell = ws2.cell(row=4, column=col_index, value=limpiar_texto(valor))
            cell.font = data_font
            cell.border = border
        # Ajustar anchos de columnas en hoja 2
        widths_seg = [12, 15, 30, 15, 12, 25]
        for i, width in enumerate(widths_seg, start=1):
            ws2.column_dimensions[chr(64 + i)].width = width
        # ===== HOJA 3: CAPACITACIONES =====
        ws3 = wb.create_sheet("Capacitaciones")
        ws3.merge_cells('A1:E1')
        ws3['A1'] = "REGISTRO DE CAPACITACIONES"
        ws3['A1'].font = title_font
        ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
        headers_capacitaciones = ['Fecha', 'Capacitacion', 'Instructor', 'Duracion (hrs)', 'Estado']
        for col_index, header in enumerate(headers_capacitaciones, start=1):
            cell = ws3.cell(row=3, column=col_index, value=limpiar_texto(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Instrucciones para capacitaciones
        ws3['A5'] = "INSTRUCCIONES:"
        ws3['A5'].font = Font(size=12, bold=True, color="8B0000")
        instrucciones = [
            "• Registrar todas las capacitaciones recibidas",
            "• Actualizar el estado segun el progreso",
            "• Incluir certificaciones obtenidas",
            "• Revisar mensualmente"
        ]
        for i, instruccion in enumerate(instrucciones, start=6):
            ws3[f'A{i}'] = limpiar_texto(instruccion)
            ws3[f'A{i}'].font = Font(size=10)
        # Ajustar anchos de columnas en hoja 3
        for i, width in enumerate([12, 25, 20, 12, 15], start=1):
            ws3.column_dimensions[chr(64 + i)].width = width
        # ===== HOJA 4: EVALUACIONES =====
        ws4 = wb.create_sheet("Evaluaciones")
        ws4.merge_cells('A1:F1')
        ws4['A1'] = "EVALUACIONES DE DESEMPEÑO"
        ws4['A1'].font = title_font
        ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
        headers_eval = ['Fecha', 'Periodo', 'Puntuacion', 'Fortalezas', 'Areas de Mejora', 'Evaluador']
        for col_index, header in enumerate(headers_eval, start=1):
            cell = ws4.cell(row=3, column=col_index, value=limpiar_texto(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Próxima evaluación (ejemplo)
        ws4['A5'] = "Proxima evaluacion:"
        ws4['B5'] = (datetime.now().replace(
                        year=datetime.now().year + 1 if datetime.now().month > 9 else datetime.now().year,
                        month=datetime.now().month + 3 if datetime.now().month <= 9 else datetime.now().month - 9
                     )).strftime('%d/%m/%Y')
        ws4['A5'].font = Font(size=11, bold=True)
        ws4['B5'].font = Font(size=11, color="FF0000")
        # Ajustar anchos de columnas en hoja 4
        widths_eval = [12, 15, 12, 25, 25, 15]
        for i, width in enumerate(widths_eval, start=1):
            ws4.column_dimensions[chr(64 + i)].width = width
        # Guardar archivo Excel en la subcarpeta documentos
        excel_filename = f"empleado_{cedula_segura}.xlsx"
        excel_path = os.path.join(employee_folder, "documentos", excel_filename)
        wb.save(excel_path)
        print(f"✅ Excel informativo generado: {excel_path}")
        return True
    except Exception as e:
        print(f"❌ Error generando Excel para empleado: {e}")
        messagebox.showerror("Error", f"Error generando Excel del empleado: {e}")
        return False
